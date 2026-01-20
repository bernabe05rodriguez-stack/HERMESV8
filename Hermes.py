# -*- coding: utf-8 -*-
"""
HΞЯMΞS V1 - Envío automático de mensajes de WhatsApp
Autor: Berna - 2025
Con procesador de Excel/CSV integrado

Código limpiado y optimizado.
(Incluye Tooltip en el título principal y corrección de encoding)

--- MODIFICADO (para incluir MODO GRUPO en Fidelizado) ---
--- FIX 6 (Fix path with spaces issue, Use keyevents for text input) ---
"""

import subprocess
import time
import random
import math
import tkinter as tk
import customtkinter as ctk
import tkinter.font as tkfont
import tkinter.ttk as ttk
from tkinter import filedialog, messagebox
import os
import threading
from datetime import datetime, timedelta
import sys
import csv
import io
import urllib.parse
import shlex # Import shlex for better command splitting
import tempfile
import shutil
import adbutils
import uiautomator2 as u2
import pytesseract
try:
    import pywinstyles
except ImportError:
    pywinstyles = None
from PIL import Image, ImageDraw, ImageFilter
import re
import xml.etree.ElementTree as ET


# --- Clase auxiliar para evitar errores de Tkinter con campos numéricos vacíos ---
class SafeIntVar(tk.IntVar):
    def get(self):
        try:
            return super().get()
        except (tk.TclError, ValueError):
            return 0

# --- Función para encontrar archivos en modo compilado ---
def resource_path(relative_path):
    """ Obtiene la ruta absoluta al recurso, funciona para desarrollo y para PyInstaller """
    try:
        # PyInstaller crea una carpeta temporal y guarda la ruta en _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

# --- Constante para el directorio base ---
BASE_DIR = resource_path(".")

# --- INICIO: Mapeo de caracteres a keycodes ADB ---
# (Simplificado, solo incluye caracteres comunes. Se puede expandir)
# Definición de sets GSM para cálculo de SMS
GSM_EXTENDED_CHARS = set("^{}\\[~]|€")
# ASCII printable (32-126) + \n + \r + €
# Nota: La ñ y acentos NO están aquí, forzando Unicode según requerimiento
GSM_BASIC_CHARS = set(chr(i) for i in range(32, 127)) | set("\n\r") | set("€")

KEYCODE_MAP = {
    '0': 'KEYCODE_0', '1': 'KEYCODE_1', '2': 'KEYCODE_2', '3': 'KEYCODE_3',
    '4': 'KEYCODE_4', '5': 'KEYCODE_5', '6': 'KEYCODE_6', '7': 'KEYCODE_7',
    '8': 'KEYCODE_8', '9': 'KEYCODE_9',
    'a': 'KEYCODE_A', 'b': 'KEYCODE_B', 'c': 'KEYCODE_C', 'd': 'KEYCODE_D',
    'e': 'KEYCODE_E', 'f': 'KEYCODE_F', 'g': 'KEYCODE_G', 'h': 'KEYCODE_H',
    'i': 'KEYCODE_I', 'j': 'KEYCODE_J', 'k': 'KEYCODE_K', 'l': 'KEYCODE_L',
    'm': 'KEYCODE_M', 'n': 'KEYCODE_N', 'o': 'KEYCODE_O', 'p': 'KEYCODE_P',
    'q': 'KEYCODE_Q', 'r': 'KEYCODE_R', 's': 'KEYCODE_S', 't': 'KEYCODE_T',
    'u': 'KEYCODE_U', 'v': 'KEYCODE_V', 'w': 'KEYCODE_W', 'x': 'KEYCODE_X',
    'y': 'KEYCODE_Y', 'z': 'KEYCODE_Z',
    ' ': 'KEYCODE_SPACE', '.': 'KEYCODE_PERIOD', ',': 'KEYCODE_COMMA',
    '!': 'KEYCODE_1', '?': 'KEYCODE_SLASH', # SHIFT + /
    '@': 'KEYCODE_AT', '#': 'KEYCODE_POUND', '$': 'KEYCODE_4', # SHIFT + 4
    '%': 'KEYCODE_5', '^': 'KEYCODE_6', '&': 'KEYCODE_7', '*': 'KEYCODE_8',
    '(': 'KEYCODE_9', ')': 'KEYCODE_0', '-': 'KEYCODE_MINUS', '_': 'KEYCODE_MINUS', # SHIFT + -
    '+': 'KEYCODE_PLUS', '=': 'KEYCODE_EQUALS', '/': 'KEYCODE_SLASH',
    '\\': 'KEYCODE_BACKSLASH', '\n': 'KEYCODE_ENTER', # Nueva línea es Enter
    # Símbolos comunes con SHIFT (esto puede variar según el layout del teclado virtual)
    ':': 'KEYCODE_SEMICOLON', # SHIFT + ;
    '"': 'KEYCODE_APOSTROPHE', # SHIFT + '
    # ... se pueden añadir más según sea necesario
}
NEEDS_SHIFT = "!@#$%^&*()_+?:\"" + "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
# --- FIN: Mapeo ---

# Verificar dependencias
try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    print("\n"+"="*50+"\nERROR: Falta 'openpyxl'. Ejecuta INSTALAR.bat.\n"+"="*50)
    input("\nEnter para salir...")
    sys.exit(1)
try:
    from PIL import Image, ImageTk
except ImportError:
    print("\n"+"="*50+"\nERROR: Falta 'Pillow'. Ejecuta INSTALAR.bat.\n"+"="*50)
    input("\nEnter para salir...")
    sys.exit(1)
try:
    import customtkinter
except ImportError:
    print("\n"+"="*50+"\nERROR: Falta 'customtkinter'. Ejecuta INSTALAR.bat.\n"+"="*50)
    input("\nEnter para salir...")
    sys.exit(1)

# --- Funciones de color ---
def _clamp(value):
    """Asegura que un valor esté entre 0 y 255."""
    return max(0, min(255, int(value)))

def lighten_color(color, factor=0.1):
    """Aclara un color hexadecimal."""
    color = color.lstrip('#')
    if len(color) == 3:
        color = "".join([c*2 for c in color])
    if len(color) != 6:
        return color
    try:
        r, g, b = int(color[0:2], 16), int(color[2:4], 16), int(color[4:6], 16)
    except ValueError:
        return color
    r = _clamp(r + (255 - r) * factor)
    g = _clamp(g + (255 - g) * factor)
    b = _clamp(b + (255 - b) * factor)
    return f"#{int(r):02x}{int(g):02x}{int(b):02x}"

def darken_color(color, factor=0.1):
    """Oscurece un color hexadecimal."""
    color = color.lstrip('#')
    if len(color) == 3:
        color = "".join([c*2 for c in color])
    if len(color) != 6:
        return color
    try:
        r, g, b = int(color[0:2], 16), int(color[2:4], 16), int(color[4:6], 16)
    except ValueError:
        return color
    r = _clamp(r * (1 - factor))
    g = _clamp(g * (1 - factor))
    b = _clamp(b * (1 - factor))
    return f"#{int(r):02x}{int(g):02x}{int(b):02x}"

def format_currency_value(value):
    """Formatea un valor numérico como moneda en formato argentino ($###.###,##)."""
    if value is None:
        return ''

    try:
        text = str(value).strip()
        if not text:
            return ''

        text = (text
                .replace('$', '')
                .replace(' ', '')
                .replace('\u202f', '')
                .replace('\u00a0', '')
                .replace('−', '-')
                )

        if ',' in text and '.' in text:
            if text.rfind(',') > text.rfind('.'):
                text = text.replace('.', '').replace(',', '.')
            else:
                text = text.replace(',', '')
        elif ',' in text:
            text = text.replace('.', '').replace(',', '.')
        elif text.count('.') > 1:
            parts = text.split('.')
            text = ''.join(parts[:-1]) + '.' + parts[-1]

        amount = float(text)
        sign = '-' if amount < 0 else ''
        formatted = f"{abs(amount):,.2f}"
        formatted = formatted.replace(',', '¤').replace('.', ',').replace('¤', '.')
        return f"{sign}${formatted}"
    except Exception:
        return str(value)

# --- INICIO MODIFICACIÓN: Clase para Tooltips (CORREGIDA) ---
class Tooltip:
    """
    Crea un tooltip (mensaje flotante) para un widget de CustomTkinter.
    Se instancia como: Tooltip(widget, "Texto del tooltip", font_info)
    """
    def __init__(self, widget, text, font_info):
        self.widget = widget
        self.text = text
        self.font_info = font_info
        self.tooltip_window = None
        self.widget.bind("<Enter>", self.show_tooltip)
        self.widget.bind("<Leave>", self.hide_tooltip)

    def show_tooltip(self, event): # <--- CORRECCIÓN: Se usa el 'event'
        if self.tooltip_window:
            return

        # Crear la ventana Toplevel
        self.tooltip_window = tk.Toplevel(self.widget)
        self.tooltip_window.wm_overrideredirect(True) # Sin bordes/barra de título

        # --- INICIO DE LA CORRECCIÓN ---
        # Usar las coordenadas del mouse (event.x_root) en lugar de
        # las coordenadas del widget (widget.winfo_rootx()), que
        # pueden ser incorrectas (0,0) si la ventana se está iniciando.
        # Se posiciona 15px a la derecha y 10px abajo del cursor.
        x = event.x_root + 15
        y = event.y_root + 10
        # --- FIN DE LA CORRECCIÓN ---

        # Ajustar si se sale de la pantalla (simple check)
        if x + 400 > self.widget.winfo_screenwidth():
            x = self.widget.winfo_screenwidth() - 410

        self.tooltip_window.wm_geometry(f"+{int(x)}+{int(y)}")

        # Añadir el label de CustomTkinter dentro
        label = ctk.CTkLabel(self.tooltip_window,
                             text=self.text,
                             font=self.font_info,
                             fg_color=("#333333", "#444444"), # Color oscuro
                             text_color="white",
                             corner_radius=6,
                             justify='left',
                             wraplength=400, # Ancho máximo del texto
                             padx=10, pady=10)
        label.pack()

        self.tooltip_window.update_idletasks()
        self.tooltip_window.lift() # Asegurarse de que esté al frente

    def hide_tooltip(self, event): # <--- CORRECCIÓN: Se usa el 'event'
        if self.tooltip_window:
            self.tooltip_window.destroy()
            self.tooltip_window = None

# --- Clase principal de la aplicación ---
# --- Clase auxiliar para las estrellas ---
class Star:
    def __init__(self, w, h):
        self.w = w
        self.h = h
        # Inicializa una estrella en una posición aleatoria (coordenadas polares)
        self.r = random.uniform(0, w/2)  # Radio (distancia al centro)
        self.a = random.uniform(0, 2*math.pi)  # Ángulo
        # Tasa de cambio de radio (velocidad) - Ajustado para pixels/frame visibles (0.5x speed)
        self.rs = random.uniform(0.1, 0.5)
        # Tasa de cambio de ángulo (rotación) (0.5x speed)
        self.ar = random.uniform(0.001, 0.003)
        self.sz = random.randint(1, 2)  # Tamaño

    def move(self, mx=None, my=None):
        # Mueve la estrella, aumentando su radio y ángulo
        self.r += self.rs
        self.a += self.ar

        # Obtener coordenadas cartesianas actuales
        cx = self.w/2 + math.cos(self.a) * self.r
        cy = self.h/2 + math.sin(self.a) * self.r

        # Lógica de atracción del mouse (si está cerca)
        if mx is not None and my is not None:
            dx = mx - cx
            dy = my - cy
            dist = math.hypot(dx, dy)
            threshold = 200  # Radio de influencia en píxeles

            if dist < threshold:
                # Factor de atracción (ajustar para "de a poco")
                factor = 0.05
                cx += dx * factor
                cy += dy * factor

                # Recalcular coordenadas polares basadas en la nueva posición atraída
                # Esto hace que la animación fluya desde el nuevo punto
                self.r = math.hypot(cx - self.w/2, cy - self.h/2)
                self.a = math.atan2(cy - self.h/2, cx - self.w/2)

        # Si la estrella sale del límite, la reinicia al centro
        if self.r > self.w/2:
            self.r = 0
            self.a = random.uniform(0, 2*math.pi)

    def get_coords(self):
        # Convierte coordenadas polares (r, a) a cartesianas (x, y)
        x = self.w/2 + math.cos(self.a) * self.r
        y = self.h/2 + math.sin(self.a) * self.r
        return x, y, self.sz

# --- Clase principal de la aplicación ---
class Hermes:
    def __init__(self, root):
        self.root = root
        self.root.title("HΞЯMΞS V1")
        self.root.geometry("1500x900")
        self.root.state('zoomed')
        self.root.resizable(True, True)
        self.root.minsize(1500, 900)

        # Variables para la animación de estrellas
        self.stars = []
        self.starfield_running = False
        self.starfield_canvas = None
        self.starfield_after_id = None
        self.mouse_x = None
        self.mouse_y = None
        self.mouse_pressed = False

        # Aplicar estilos de Windows 11 si está disponible
        if pywinstyles:
            try:
                pywinstyles.apply_style(self.root, "mica")
            except Exception as e:
                print(f"Error aplicando estilo Windows: {e}")

        # Variables de estado
        self.adb_path = tk.StringVar(value="")
        self.delay_min = SafeIntVar(value=3)
        self.delay_max = SafeIntVar(value=5)
        self.wait_after_open = SafeIntVar(value=7)
        self.pause_sends_count = SafeIntVar(value=10)
        self.pause_sends_delay_min = SafeIntVar(value=30)
        self.pause_sends_delay_max = SafeIntVar(value=60)

        # Variables para el nuevo SMS
        self.sms_mode_active = False
        self.sms_delay_min = SafeIntVar(value=10)
        self.sms_delay_max = SafeIntVar(value=15)
        self.sms_simultaneous_mode = tk.BooleanVar(value=False)
        self.sms_enable_call = tk.BooleanVar(value=False)
        self.sms_call_duration = SafeIntVar(value=10)

        # Variables para el modo Llamadas
        self.calls_delay_min = SafeIntVar(value=5)
        self.calls_delay_max = SafeIntVar(value=10)
        self.calls_duration = SafeIntVar(value=15)
        self.calls_whatsapp_mode = tk.StringVar(value="Business")
        self.calls_simultaneous_mode = tk.BooleanVar(value=False)
        self.calls_selected_columns = []
        self.calls_phone_columns_vars = {} # Map col_name -> BooleanVar

        self.excel_file = ""
        self.links = []
        self.link_retry_map = {}
        self.devices = []

        self.is_running = False
        self.is_paused = False
        self.should_stop = False
        self.pause_lock = threading.Lock()

        self.total_messages = 0
        self.sent_count = 0
        self.failed_count = 0
        self.current_index = 0
        self.start_time = None
        self.task_times = []  # Lista de tiempos de cada tarea para promediado
        self.last_task_time = None  # Tiempo de inicio de la última tarea

        # --- INICIO MODIFICACIÓN: Variables de Fidelizado (Modo Bucles Blast V2) ---
        self.manual_inputs_numbers = [] # Almacena números
        self.manual_inputs_groups = [] # Almacena links de grupo
        self.manual_paired_messages = [] # Almacena los mensajes pareados (para Modo Grupo y Mixto)
        self.manual_messages_numbers = [] # Almacena los mensajes .txt para números
        self.manual_messages_groups = [] # Almacena los mensajes .txt para grupos
        
        self.manual_mode = False # Flag general de Fidelizado
        self.group_mode = False # Flag para MODO GRUPO (puro)

        # Registro de actividad para modo Fidelizado (Números)
        self.fidelizado_activity_records = []
        self.log_view_mode = "log"

        # Estado: "NUMEROS", "GRUPOS", "MIXTO" o None (modo tradicional Excel/CSV)
        self.fidelizado_mode = None
        self.mixto_variant = tk.IntVar(value=1)  # Variante del modo mixto: 1, 2 o 3
        
        # Índice de inicio aleatorio para rotación de mensajes
        self.mensaje_start_index = 0
        
        # --- INICIO MODIFICACIÓN: Variables para el nuevo Modo Números Automático ---
        self.fidelizado_numeros_mode = tk.StringVar(value="Uno a uno")
        self.detected_phone_lines = [] # Almacenará {"device": str, "type": "WA/WB", "number": str}
        self.manual_cycles_var = SafeIntVar(value=1) # NUEVO: Para los ciclos del modo "Uno a uno"
        # --- FIN MODIFICACIÓN ---

        self.manual_loops = 1

        self.fidelizado_delay_min = SafeIntVar(value=10)
        self.fidelizado_delay_max = SafeIntVar(value=15)
        # --- FIX: Añadir variables separadas para el retardo entre envíos ---
        self.fidelizado_send_delay_min = SafeIntVar(value=10)
        self.fidelizado_send_delay_max = SafeIntVar(value=15)

        self.simultaneous_mode = tk.BooleanVar(value=False) # Flag para modo simultáneo (usando BooleanVar para UI)
        self.auto_join_groups = tk.BooleanVar(value=True) # Permite auto-unirse a grupos cuando se detecta la pantalla de invitación

        # Estado de widgets bloqueados
        self._blocked_widgets_state = {}

        # Variables para la vista del log
        self.log_detailed_view = False # False = simple, True = detallado
        self.log_history = [] # Almacena tuplas de (full_message, tag)

        # Variables de tiempo para Modo Grupos Dual
        self.report_data = []
        self.wait_after_write = SafeIntVar(value=2)  # Tiempo después de escribir antes del primer Enter
        self.wait_after_first_enter = SafeIntVar(value=2)  # Tiempo de espera tras el primer Enter
        self.wait_between_enters = SafeIntVar(value=3)  # Tiempo de espera tras presionar Enter
        self.wait_between_messages = SafeIntVar(value=2)  # Tiempo entre Business y Normal
        self.whatsapp_mode = tk.StringVar(value="Todas")  # Qué WhatsApp usar: Normal, Business, Ambos
        self.whatsapp_mode.trace_add('write', self.update_per_whatsapp_stat)
        self.traditional_send_mode = tk.StringVar(value="Business")  # Modo de envío tradicional: Business, Normal, Ambos, TODOS

        self.raw_data = []
        self.columns = []
        self.selected_columns = []
        self.phone_columns = []

        self.fidelizado_unlocked = True
        self.fidelizado_unlock_btn = None
        self.detect_numbers_btn = None
        self.last_detection_snapshot = {}
        self.numbers_editor_window = None
        self.numbers_editor_entries = {}
        self.numbers_editor_closed_event = threading.Event()
        self.numbers_editor_closed_event.set()
        self.numbers_editor_start_requested = False
        self.dark_mode = True  # Estado del modo oscuro (Predeterminado: True)

        # Paleta de colores
        self.colors_light = {
            'blue': '#4285F4', 'green': '#1DB954', 'orange': '#FB923C',
            'bg': '#e8e8e8', 'bg_card': '#ffffff', 'bg_header': '#ffffff',
            'bg_log': '#282c34',
            'log_text': '#abb2bf', 'log_success': '#98c379', 'log_error': '#e06c75',
            'log_warning': '#d19a66', 'log_info': '#61afef',
            'text': '#202124', 'text_light': '#5f6368', 'text_header_buttons': '#ffffff', 'text_header': '#000000', 'log_title_color': '#ffffff',
            'action_detect': '#2563EB', 'action_excel': '#FB923C',
            'action_mode': '#0EA5E9', 'action_fidelizador': '#111827', 'action_start': '#16A34A',
            'action_pause': '#FB923C', 'action_cancel': '#DC2626'
        }

        self.colors_dark = {
            'blue': '#5B9FFF', 'green': '#1ED760', 'orange': '#FFA45C',
            'bg': '#000000', 'bg_card': '#1a1a1a', 'bg_header': '#1a1a1a',
            'bg_log': '#1a1a1a',
            'log_text': '#ffffff', 'log_success': '#98c379', 'log_error': '#e06c75',
            'log_warning': '#d19a66', 'log_info': '#61afef',
            'text': '#ffffff', 'text_light': '#cccccc', 'text_header_buttons': '#ffffff', 'text_header': '#ffffff', 'log_title_color': '#ffffff',
            'action_detect': '#5B9FFF', 'action_excel': '#FFA45C',
            'action_mode': '#38BDF8', 'action_fidelizador': '#e4e6eb', 'action_start': '#22C55E',
            'action_pause': '#FFA45C', 'action_cancel': '#EF4444'
        }
        
        self.colors = self.colors_dark.copy()

        self.hover_colors = {k: darken_color(v, 0.18) for k, v in self.colors.items() if k.startswith('action_')}

        # Fuentes
        self.fonts = {
            'header': ('Big Russian', 76, 'bold'),
            'card_title': ('Inter', 18, 'bold'),
            'button': ('Inter', 13, 'bold'),
            'button_small': ('Inter', 12, 'bold'),
            'stat_value': ('Inter', 40, 'bold'),
            'stat_label': ('Inter', 13),
            'setting_label': ('Inter', 12),
            'log_title': ('Inter', 16, 'bold'),
            'log_text': ('Consolas', 12),
            'progress_label': ('Inter', 11),
            'progress_value': ('Inter', 16, 'bold'),
            'time_label': ('Inter', 10),
            'dialog_title': ('Inter', 16, 'bold'),
            'dialog_text': ('Inter', 12)
        }

        self.auto_detect_adb()
        self.setup_ui()

    def _center_toplevel(self, window, width, height):
        """Centrar una ventana toplevel en la pantalla."""
        try:
            window.update_idletasks()
            sw = window.winfo_screenwidth()
            sh = window.winfo_screenheight()
            x = (sw // 2) - (width // 2)
            y = (sh // 2) - (height // 2)
            window.geometry(f"{width}x{height}+{x}+{y}")
        except tk.TclError:
            pass

    def open_tutorial_window(self):
        """Abre una ventana flotante con instrucciones (Tutorial)."""
        if hasattr(self, 'tutorial_window') and self.tutorial_window is not None and self.tutorial_window.winfo_exists():
            self.tutorial_window.lift()
            self.tutorial_window.focus_force()
            return

        self.tutorial_window = ctk.CTkToplevel(self.root)
        self.tutorial_window.title("Tutorial")
        self.tutorial_window.geometry("420x320")
        self.tutorial_window.attributes('-topmost', True)

        # Intentar centrar
        self._center_toplevel(self.tutorial_window, 420, 320)

        # Usar un textbox de solo lectura para el contenido
        textbox = ctk.CTkTextbox(
            self.tutorial_window,
            font=('Inter', 14),
            text_color=self.colors['text'],
            fg_color=self.colors['bg_card'],
            wrap="word",
            corner_radius=0,
            border_width=0
        )
        textbox.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        text_content = (
            "PASO 1 - Preparar el dispositivo.\n\n"
            "1.  Ve a -Ajustes- ⚙️.\n"
            "2.  En la lupa, escribe -Compilación- y toca el resultado -7 veces- seguidas.\n"
            "3.  Vuelve al buscador, escribe -Depuración por USB- y actívala."
        )

        textbox.insert("0.0", text_content)
        textbox.configure(state="disabled")

    def setup_ui(self):
        # Configurar fondo de la ventana principal
        self.root.configure(fg_color=self.colors['bg'])
        
        # 1. Header (Guardado en self.header_frame, NO SE EMPAQUETA AÚN)
        self.header_frame = ctk.CTkFrame(self.root, fg_color=self.colors['bg_header'], height=140, corner_radius=30)
        # self.header_frame.pack(fill=tk.X, pady=(10, 10), padx=10) # MOVIDO A enter_app_mode
        self.header_frame.pack_propagate(False)

        hc = ctk.CTkFrame(self.header_frame, fg_color=self.colors['bg_header'])
        hc.pack(expand=True, fill=tk.X, padx=40)


        # Logo Izquierdo
        try:
            l_img_path = os.path.join(BASE_DIR, 'logo_left.png')
            l_img = Image.open(l_img_path).resize((150, 150), Image.Resampling.LANCZOS)
            l_pho = ctk.CTkImage(light_image=l_img, dark_image=l_img, size=(150, 150))
            self.header_logo_left = ctk.CTkLabel(hc, image=l_pho, text="")
            self.header_logo_left.pack(side=tk.LEFT, padx=(0, 20))
        except Exception as e:
            print(f"Error cargando logo_left: {e}")
            self.header_logo_left = ctk.CTkLabel(hc, text="🦶", font=('Inter', 60), fg_color="transparent")
            self.header_logo_left.pack(side=tk.LEFT, padx=(0, 20))

        # Logo Derecho
        try:
            r_img_path = os.path.join(BASE_DIR, 'logo_right.png')
            r_img = Image.open(r_img_path).resize((150, 150), Image.Resampling.LANCZOS)
            r_pho = ctk.CTkImage(light_image=r_img, dark_image=r_img, size=(150, 150))
            self.header_logo_right = ctk.CTkLabel(hc, image=r_pho, text="")
            self.header_logo_right.pack(side=tk.RIGHT, padx=(20, 0))
        except Exception as e:
            print(f"Error cargando logo_right: {e}")
            self.header_logo_right = ctk.CTkLabel(hc, text="🦶", font=('Inter', 60), fg_color="transparent")
            self.header_logo_right.pack(side=tk.RIGHT, padx=(20, 0))

        # Título
        self.header_title_label = ctk.CTkLabel(hc, text="HΞЯMΞS", font=self.fonts['header'],
                                   fg_color="transparent",
                                   text_color=self.colors['text_header'],
                                   cursor="hand2") # Añadir cursor para indicar que es interactivo
        self.header_title_label.pack(side=tk.LEFT, fill=tk.X, expand=True, anchor='center')

        # Tooltip para el título
        tooltip_text = (
            "Hermes fue el mensajero de los dioses en la mitología griega. \n"
            "Hijo de Zeus, simbolizaba la comunicación, la rapidez y el ingenio. \n"
            "También protegía a los viajeros y guiaba las almas al inframundo.\n\n"
            "Programa pensado y creado por \n"
            "BERNABE GABRIEL RODRIGUEZ, y FRANCISCO JOSE RODRIGUEZ."
        )
        tooltip_font = self.fonts.get('dialog_text', ('Inter', 12))
        self.hermes_tooltip = Tooltip(widget=self.header_title_label, text=tooltip_text, font_info=tooltip_font)

        # 2. Contenedor principal scrollable (Guardado en self.main_content_frame)
        self.main_content_frame = ctk.CTkFrame(self.root, fg_color="transparent")
        # self.main_content_frame.pack(fill=tk.BOTH, expand=True, padx=40, pady=(0, 20)) # MOVIDO

        self.scroll_frame = ctk.CTkScrollableFrame(self.main_content_frame, fg_color="transparent")
        self.scroll_frame.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)
        self.scroll_frame.grid_columnconfigure(0, weight=618, uniform='main_panels')
        self.scroll_frame.grid_columnconfigure(1, weight=382, uniform='main_panels')
        self.scroll_frame.grid_rowconfigure(0, weight=1)
        self.main_layout = self.scroll_frame

        # 3. Paneles
        left = ctk.CTkFrame(self.main_layout, fg_color="transparent")
        right = ctk.CTkFrame(self.main_layout, fg_color="transparent")
        self.left_panel = left
        self.right_panel = right
        self._current_main_layout = None

        self.root.bind("<Configure>", self._on_main_configure)
        self.setup_right(right) # FIX: Inicializar el panel derecho primero para que exista el log_text
        self.setup_left(left)

        # NO llamar a _update_main_layout todavía porque los widgets no están packeados
        # self._update_main_layout(self.root.winfo_width())

        # Configurar atajos globales
        self.setup_global_shortcuts()

        # INICIAR EN EL MENÚ DE INICIO
        self.setup_start_menu()

    def init_starfield(self, width, height, color="white"):
        """Inicializa las estrellas para la animación y crea sus items en el canvas."""
        self.stars = []
        # Limpiar canvas solo de estrellas, por si acaso se llama de nuevo
        if hasattr(self, 'starfield_canvas') and self.starfield_canvas.winfo_exists():
            self.starfield_canvas.delete("star")

            for _ in range(500):
                star = Star(width, height)
                x, y, sz = star.get_coords()
                # Crear item en canvas y guardar ID en el objeto Star
                star.item_id = self.starfield_canvas.create_oval(x, y, x+sz, y+sz, fill=color, outline=color, tags="star")
                self.stars.append(star)

    def animate_starfield(self):
        """Actualiza y dibuja el campo de estrellas moviendo los items existentes."""
        if not self.starfield_running or not self.starfield_canvas or not self.starfield_canvas.winfo_exists():
            return

        # Obtener dimensiones actuales (por si redimensionan la ventana)
        w = self.starfield_canvas.winfo_width()
        h = self.starfield_canvas.winfo_height()

        # Si el canvas es muy pequeño (al iniciar), usar dimensiones por defecto
        if w < 10 or h < 10:
            w, h = 1500, 900

        # Mover estrellas existentes (Optimizado: sin delete/create)
        for star in self.stars:
            star.w = w  # Actualizar límites por si cambia el tamaño
            star.h = h

            # Solo aplicar atracción si el mouse está presionado
            if self.mouse_pressed:
                star.move(self.mouse_x, self.mouse_y)
            else:
                star.move(None, None)

            x, y, sz = star.get_coords()

            if hasattr(star, 'item_id'):
                self.starfield_canvas.coords(star.item_id, x, y, x+sz, y+sz)

        # Programar siguiente frame (aprox 60 FPS -> 16ms)
        self.starfield_after_id = self.root.after(20, self.animate_starfield)

    def stop_starfield(self):
        """Detiene la animación de estrellas."""
        self.starfield_running = False
        if self.starfield_after_id:
            self.root.after_cancel(self.starfield_after_id)
            self.starfield_after_id = None

    def setup_start_menu(self):
        """Crea y muestra el menú de inicio (Diseño Espacial/Minimalista)."""
        if hasattr(self, 'start_menu_frame') and self.start_menu_frame.winfo_exists():
            self.start_menu_frame.pack(fill=tk.BOTH, expand=True)
            self.starfield_running = True
            self.animate_starfield()
            return

        # Determinar colores según el modo
        is_dark = getattr(self, 'dark_mode', False)
        bg_color = "black" if is_dark else "white"
        fg_color = "white" if is_dark else "black"

        # Configurar frame principal
        self.start_menu_frame = ctk.CTkFrame(self.root, fg_color=bg_color)
        self.start_menu_frame.pack(fill=tk.BOTH, expand=True)

        # Crear Canvas para el fondo de estrellas
        self.starfield_canvas = tk.Canvas(self.start_menu_frame, bg=bg_color, highlightthickness=0)
        self.starfield_canvas.place(relwidth=1, relheight=1) # Llenar todo el fondo

        # Inicializar estrellas con dimensiones fijas iniciales
        init_w = self.root.winfo_width()
        init_h = self.root.winfo_height()
        if init_w < 100: init_w = 1500
        if init_h < 100: init_h = 900

        self.init_starfield(init_w, init_h, color=fg_color)
        self.starfield_running = True
        self.animate_starfield()

        # Título principal HHERMES (Agrandado) - Ahora en Canvas
        header_font = list(self.fonts.get('header', ('Big Russian', 76, 'bold')))
        header_font[1] = 100 # Aumentar tamaño de fuente

        self.title_text_id = self.starfield_canvas.create_text(
            0, 0, # Se posiciona en _update_start_menu_layout
            text="HΞЯMΞS",
            font=tuple(header_font),
            fill=fg_color,
            anchor='center'
        )

        # Botón Modo Oscuro (Emoji en Canvas)
        self.dark_mode_text_id = self.starfield_canvas.create_text(
            0, 0, # Se posiciona en _update_start_menu_layout
            text="🌙" if is_dark else "☀️",
            font=('Inter', 34),
            fill=fg_color,
            anchor='center'
        )
        self.starfield_canvas.tag_bind(self.dark_mode_text_id, "<Button-1>", lambda _event: self.toggle_dark_mode())
        self.starfield_canvas.tag_bind(self.dark_mode_text_id, "<Enter>", lambda _event: self.starfield_canvas.config(cursor="hand2"))
        self.starfield_canvas.tag_bind(self.dark_mode_text_id, "<Leave>", lambda _event: self.starfield_canvas.config(cursor=""))

        # --- LOGOS EN CANVAS (Transparencia Real) ---

        # Cargar Imágenes
        try:
            img_size_normal = (300, 300)
            img_size_hover = (330, 330)

            # WSP Images
            wsp_path = os.path.join(BASE_DIR, "WSP alas.png")
            if os.path.exists(wsp_path):
                pil_wsp = Image.open(wsp_path)
                self.img_wsp_normal = ImageTk.PhotoImage(pil_wsp.resize(img_size_normal, Image.Resampling.LANCZOS))
                self.img_wsp_hover = ImageTk.PhotoImage(pil_wsp.resize(img_size_hover, Image.Resampling.LANCZOS))
            else:
                self.img_wsp_normal = None

            # SMS Images
            sms_path = os.path.join(BASE_DIR, "SMS alas.png")
            if os.path.exists(sms_path):
                pil_sms = Image.open(sms_path)
                self.img_sms_normal = ImageTk.PhotoImage(pil_sms.resize(img_size_normal, Image.Resampling.LANCZOS))
                self.img_sms_hover = ImageTk.PhotoImage(pil_sms.resize(img_size_hover, Image.Resampling.LANCZOS))
            else:
                self.img_sms_normal = None

        except Exception as e:
            print(f"Error cargando imágenes del menú: {e}")
            self.img_wsp_normal = None
            self.img_sms_normal = None

        # Crear items en Canvas
        self.canvas_item_wsp = None
        self.canvas_item_sms = None

        if self.img_wsp_normal:
            self.canvas_item_wsp = self.starfield_canvas.create_image(0, 0, image=self.img_wsp_normal, anchor='center')
            # Bindings WSP
            self.starfield_canvas.tag_bind(self.canvas_item_wsp, '<Button-1>', lambda e: self.enter_app_mode("whatsapp"))
            self.starfield_canvas.tag_bind(self.canvas_item_wsp, '<Enter>', lambda e: self._on_hover_start_logo(self.canvas_item_wsp, self.img_wsp_hover))
            self.starfield_canvas.tag_bind(self.canvas_item_wsp, '<Leave>', lambda e: self._on_leave_start_logo(self.canvas_item_wsp, self.img_wsp_normal))

        if self.img_sms_normal:
            self.canvas_item_sms = self.starfield_canvas.create_image(0, 0, image=self.img_sms_normal, anchor='center')
            # Bindings SMS
            self.starfield_canvas.tag_bind(self.canvas_item_sms, '<Button-1>', lambda e: self.enter_app_mode("sms"))
            self.starfield_canvas.tag_bind(self.canvas_item_sms, '<Enter>', lambda e: self._on_hover_start_logo(self.canvas_item_sms, self.img_sms_hover))
            self.starfield_canvas.tag_bind(self.canvas_item_sms, '<Leave>', lambda e: self._on_leave_start_logo(self.canvas_item_sms, self.img_sms_normal))

        # Copyright Text
        self.copyright_text_id = self.starfield_canvas.create_text(
            0, 0, # Se posiciona en _update_start_menu_layout
            text="Copyright © 2025 Hermes Inc. All rights reserved.",
            font=('Inter', 8),
            fill=fg_color,
            anchor='center'
        )

        # Bind resize event to center logos
        self.starfield_canvas.bind('<Configure>', self._update_start_menu_layout)
        # Bind mouse movement for star attraction
        self.starfield_canvas.bind('<Motion>', self._on_mouse_move)
        self.starfield_canvas.bind('<Button-1>', self._on_mouse_down)
        self.starfield_canvas.bind('<ButtonRelease-1>', self._on_mouse_up)

    def _on_mouse_move(self, event):
        """Actualiza las coordenadas del mouse para la animación de estrellas."""
        self.mouse_x = event.x
        self.mouse_y = event.y

    def _on_mouse_down(self, event):
        self.mouse_pressed = True

    def _on_mouse_up(self, event):
        self.mouse_pressed = False

    def _on_hover_start_logo(self, tag, hover_img):
        """Efecto Hover: Cambia imagen y cursor."""
        self.starfield_canvas.itemconfig(tag, image=hover_img)
        self.starfield_canvas.config(cursor="hand2")

    def _on_leave_start_logo(self, tag, normal_img):
        """Efecto Leave: Restaura imagen y cursor."""
        self.starfield_canvas.itemconfig(tag, image=normal_img)
        self.starfield_canvas.config(cursor="")

    def _update_start_menu_layout(self, event=None):
        """Centra los logos en el canvas cuando cambia el tamaño de la ventana."""
        if not hasattr(self, 'starfield_canvas') or not self.starfield_canvas.winfo_exists():
            return

        w = self.starfield_canvas.winfo_width()
        h = self.starfield_canvas.winfo_height()

        # Posiciones relativas
        # WSP a la izquierda (35%), SMS a la derecha (65%)
        # Centrados verticalmente (55% para dar espacio al titulo grande)
        y_pos = h * 0.55

        if self.canvas_item_wsp:
            self.starfield_canvas.coords(self.canvas_item_wsp, w * 0.35, y_pos)

        if self.canvas_item_sms:
            self.starfield_canvas.coords(self.canvas_item_sms, w * 0.65, y_pos)

        if hasattr(self, 'copyright_text_id') and self.copyright_text_id:
            self.starfield_canvas.coords(self.copyright_text_id, w * 0.5, h - 30)

        if hasattr(self, 'title_text_id') and self.title_text_id:
            # Posicionar título al 30% de la altura (aprox donde estaba el header)
            self.starfield_canvas.coords(self.title_text_id, w * 0.5, 125)

        if hasattr(self, 'dark_mode_text_id') and self.dark_mode_text_id:
            # Posicionar botón dark mode alineado con el título, a la derecha
            self.starfield_canvas.coords(self.dark_mode_text_id, w - 100, 125)

    def enter_app_mode(self, mode):
        """Transición del menú de inicio a la aplicación principal."""
        self.stop_starfield()
        self.start_menu_frame.pack_forget()

        # Mostrar UI principal
        self.header_frame.pack(fill=tk.X, pady=(10, 10), padx=10)
        self.main_content_frame.pack(fill=tk.BOTH, expand=True, padx=40, pady=(0, 20))

        # Forzar actualización de layout
        self.root.update_idletasks()
        self._update_main_layout(self.root.winfo_width())

        if mode == "whatsapp":
            self.show_traditional_view()
            self.log("Modo WhatsApp activado", 'info')
        elif mode == "sms":
            self.show_sms_view()
            self.log("Modo SMS activado", 'info')

    def return_to_start_menu(self):
        """Vuelve al menú de inicio ocultando la UI principal."""
        self.header_frame.pack_forget()
        self.main_content_frame.pack_forget()
        self.setup_start_menu() # Recrear o repackear

    def _on_main_configure(self, event):
        # Solo actualizar si la UI principal está visible
        if hasattr(self, 'main_content_frame') and self.main_content_frame.winfo_ismapped():
            self._update_main_layout(self.root.winfo_width())

    def _update_main_layout(self, width=None):
        """Cambia entre vista de 2 columnas o 1 columna (apilada) si la ventana es muy angosta."""
        if not hasattr(self, 'left_panel') or not hasattr(self, 'right_panel'):
            return
        if not width:
            width = self.root.winfo_width() - 80 # 80 por el padding

        mode = 'stacked' if width < 1100 else 'columns'

        if self._current_main_layout == mode:
            return

        self.left_panel.update_idletasks()
        self.right_panel.update_idletasks()

        if mode == 'columns':
            self.main_layout.grid_columnconfigure(0, weight=618, uniform='main_panels', minsize=0)
            self.main_layout.grid_columnconfigure(1, weight=382, uniform='main_panels', minsize=0)
            self.main_layout.grid_rowconfigure(1, weight=0)
            self.left_panel.grid(row=0, column=0, sticky='nsew', padx=(0, 10), pady=0)
            self.right_panel.grid(row=0, column=1, sticky='nsew', padx=(10, 0), pady=0)
        else: # mode == 'stacked'
            self.main_layout.grid_columnconfigure(0, weight=1, uniform='main_panels', minsize=0)
            self.main_layout.grid_columnconfigure(1, weight=0, minsize=0)
            self.main_layout.grid_rowconfigure(1, weight=1)
            self.left_panel.grid(row=0, column=0, sticky='nsew', padx=0, pady=0)
            self.right_panel.grid(row=1, column=0, sticky='nsew', padx=0, pady=0)

        # Forzar estado correcto del panel de registro/tabla tras un redibujado
        self._enforce_log_state()

    def _enforce_log_state(self):
        """Asegura que la visibilidad del log/tabla sea correcta tras redimensionar."""
        if not hasattr(self, 'log_table_container') or not hasattr(self, 'log_text'):
            return

        # Si el botón de tabla está oculto o deshabilitado (modo tradicional/SMS), forzar vista de log
        # Usamos la visibilidad del botón como proxy del estado, o el modo fidelizado
        is_numeros_mode = (self.fidelizado_mode == "NUMEROS")

        if not is_numeros_mode:
            # Forzar modo Log Texto y ocultar Tabla
            if self.log_table_container.winfo_ismapped():
                self.log_table_container.grid_remove()
            if not self.log_text.winfo_ismapped():
                self.log_text.grid(row=0, column=0, sticky="nsew")

            # Asegurar que el botón toggle esté oculto
            if hasattr(self, 'toggle_log_view_btn') and self.toggle_log_view_btn.winfo_ismapped():
                self.toggle_log_view_btn.pack_forget()
        else:
            # Estamos en modo Números, respetar la preferencia del usuario (log_view_mode)
            if hasattr(self, 'toggle_log_view_btn') and not self.toggle_log_view_btn.winfo_ismapped():
                self.toggle_log_view_btn.pack(side=tk.RIGHT, padx=(10, 0))
                self.toggle_log_view_btn.configure(state=tk.NORMAL)

            if self.log_view_mode == "table":
                if not self.log_table_container.winfo_ismapped():
                    self.log_text.grid_remove()
                    self.log_table_container.grid(row=0, column=0, sticky="nsew")
            else:
                if not self.log_text.winfo_ismapped():
                    self.log_table_container.grid_remove()
                    self.log_text.grid(row=0, column=0, sticky="nsew")

    def _iniciar_ver_pantalla(self):
        """Handles the logic for the 'Ver Pantalla' button."""
        if not self.devices:
            messagebox.showerror("Error", "No hay dispositivos detectados. Por favor, detecta los dispositivos primero.", parent=self.root)
            return

        if len(self.devices) == 1:
            self._lanzar_scrcpy(self.devices[0])
        else:
            self._open_device_selection_window()

    def _open_device_selection_window(self):
        """Opens a window to select a device for screen mirroring."""
        selection_window = ctk.CTkToplevel(self.root)
        selection_window.title("Seleccionar Dispositivo")
        selection_window.transient(self.root)
        selection_window.grab_set()

        # Center the window
        width, height = 300, len(self.devices) * 50 + 40
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - (width // 2)
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - (height // 2)
        selection_window.geometry(f"{width}x{height}+{x}+{y}")

        ctk.CTkLabel(selection_window, text="Selecciona un dispositivo:", font=self.fonts['button']).pack(pady=10)

        def on_select(device_id):
            selection_window.destroy()
            self._lanzar_scrcpy(device_id)

        for device_id in self.devices:
            btn = ctk.CTkButton(selection_window, text=device_id, command=lambda d=device_id: on_select(d),
                                font=self.fonts['button'], height=40)
            btn.pack(fill=tk.X, padx=20, pady=5)

    def _lanzar_scrcpy(self, device_id):
        """Launches scrcpy for a specific device."""
        scrcpy_path = os.path.join(BASE_DIR, "scrcpy-win64-v3.2", "scrcpy.exe")

        if not os.path.exists(scrcpy_path):
            self.log(f"Error: No se encontró scrcpy.exe en la ruta esperada: {scrcpy_path}", 'error')
            messagebox.showerror("Error", "No se encontró 'scrcpy.exe'.\nAsegúrate de que la carpeta 'scrcpy-win64-v3.2' está en el directorio del programa.", parent=self.root)
            return

        self.log(f"Iniciando scrcpy para el dispositivo: {device_id}", 'info')

        try:
            # Use Popen to run scrcpy in a non-blocking way
            command = [scrcpy_path, "-s", device_id]
            subprocess.Popen(command)
        except Exception as e:
            self.log(f"Error al iniciar scrcpy: {e}", 'error')
            messagebox.showerror("Error", f"No se pudo iniciar scrcpy:\n{e}", parent=self.root)

    def setup_left(self, parent):
        # Contenedor principal para las vistas
        self.views_container = ctk.CTkFrame(parent, fg_color="transparent")
        self.views_container.pack(fill=tk.X, expand=False, anchor="n")

        # Las vistas ahora se inicializan bajo demanda (lazy loading)
        self.traditional_view_frame = None
        self.fidelizado_view_frame = None
        self.sms_view_frame = None
        self.calls_view_frame = None

    def show_traditional_view(self):
        """
        ################################################################################
        # BLOQUE WHATSAPP (TRADICIONAL Y FIDELIZADO)
        ################################################################################
        Muestra la vista tradicional, creándola si es la primera vez."""
        # Guardar datos de los textboxes para persistencia (si la vista Fidelizado ya existe)
        if self.fidelizado_view_frame and hasattr(self, 'fidelizado_groups_text'):
            self.manual_inputs_groups = [line.strip() for line in self.fidelizado_groups_text.get("1.0", tk.END).splitlines() if line.strip()]
        if self.fidelizado_view_frame and hasattr(self, 'fidelizado_numbers_text'):
            self.manual_inputs_numbers = [line.strip() for line in self.fidelizado_numbers_text.get("1.0", tk.END).splitlines() if line.strip()]
            self.manual_messages_groups = self.manual_messages_numbers

        self.sms_mode_active = False
        if self.fidelizado_view_frame:
            self.fidelizado_view_frame.pack_forget()
        if self.sms_view_frame:
            self.sms_view_frame.pack_forget()

        # Crear la vista si no existe
        if self.traditional_view_frame is None:
            self.traditional_view_frame = ctk.CTkFrame(self.views_container, fg_color="transparent")
            self.setup_traditional_view(self.traditional_view_frame)

        # Ocultar vista de Llamadas si está visible
        if self.calls_view_frame:
            self.calls_view_frame.pack_forget()

        self.traditional_view_frame.pack(fill=tk.X, expand=False, anchor="n")
        self.update_per_whatsapp_stat()
        self._apply_fidelizado_layout_styles(False)
        self.set_activity_table_enabled(False)

    def show_calls_view(self):
        """Muestra la vista de Llamadas por WhatsApp."""
        self.sms_mode_active = False
        if self.traditional_view_frame:
            self.traditional_view_frame.pack_forget()
        if self.fidelizado_view_frame:
            self.fidelizado_view_frame.pack_forget()
        if self.sms_view_frame:
            self.sms_view_frame.pack_forget()

        if self.calls_view_frame is None:
            self.calls_view_frame = ctk.CTkFrame(self.views_container, fg_color="transparent")
            self.setup_calls_view(self.calls_view_frame)

        self.calls_view_frame.pack(fill=tk.X, expand=False, anchor="n")
        self.update_per_whatsapp_stat()
        self._apply_fidelizado_layout_styles(False)
        self.set_activity_table_enabled(False)

    def show_fidelizado_view(self):
        """Muestra la vista de Fidelizado, creándola si es la primera vez."""
        self.sms_mode_active = False
        if self.traditional_view_frame:
            self.traditional_view_frame.pack_forget()
        if self.sms_view_frame:
            self.sms_view_frame.pack_forget()
        if self.calls_view_frame:
            self.calls_view_frame.pack_forget()

        # Crear la vista si no existe
        if self.fidelizado_view_frame is None:
            self.fidelizado_view_frame = ctk.CTkFrame(self.views_container, fg_color="transparent")
            self.setup_fidelizado_view(self.fidelizado_view_frame)

        self._populate_fidelizado_inputs() # Repoblar datos al mostrar la vista
        self.fidelizado_view_frame.pack(fill=tk.X, expand=False, anchor="n")
        self.update_per_whatsapp_stat()
        self._apply_fidelizado_layout_styles(True)
        self.set_activity_table_enabled(self.fidelizado_mode == "NUMEROS")

    def show_sms_view(self):
        """
        ################################################################################
        # BLOQUE SMS
        ################################################################################
        Activa la vista de envío por SMS, creándola si es la primera vez."""
        self.sms_mode_active = True
        self.manual_mode = False
        self.fidelizado_mode = None

        if self.links and not all(link.lower().startswith("sms:") for link in self.links):
            self.links = []
            self.link_retry_map = {}
            self.total_messages = 0
            self.update_stats()
            self.log("SMS activo: limpia enlaces previos para evitar envíos erróneos.", 'warning')

        if self.traditional_view_frame:
            self.traditional_view_frame.pack_forget()
        if self.fidelizado_view_frame:
            self.fidelizado_view_frame.pack_forget()
        if self.calls_view_frame:
            self.calls_view_frame.pack_forget()

        # Crear la vista si no existe
        if self.sms_view_frame is None:
            self.sms_view_frame = ctk.CTkFrame(self.views_container, fg_color="transparent")
            self.setup_sms_view(self.sms_view_frame)

        self.sms_view_frame.pack(fill=tk.X, expand=False, anchor="n")
        self.update_per_whatsapp_stat()
        self._apply_fidelizado_layout_styles(False)
        self.set_activity_table_enabled(False)

    def _apply_fidelizado_layout_styles(self, active):
        """Ajusta bordes y altura de tarjetas cuando se activa el modo Fidelizado."""
        if hasattr(self, 'fidelizado_main_card'):
            self.fidelizado_main_card.configure(
                corner_radius=32 if active else 30,
                border_width=1 if active else 0,
                border_color=self._section_border_color()
            )

        if hasattr(self, 'log_card'):
            self.log_card.configure(height=0)
            self.log_card.pack_configure(**self.log_card_pack_defaults)
            self.log_card.pack_propagate(True)

            if hasattr(self, 'log_card_header') and hasattr(self, 'log_card_header_padding'):
                self.log_card_header.grid_configure(**self.log_card_header_padding)
            if hasattr(self, 'log_card_body') and hasattr(self, 'log_card_body_padding'):
                self.log_card_body.grid_configure(**self.log_card_body_padding)

    def setup_traditional_view(self, parent):
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(0, weight=0)

        content = ctk.CTkFrame(
            parent,
            fg_color=self.colors['bg_card'],
            corner_radius=30,
            border_width=1,
            border_color=self._section_border_color()
        )
        content.grid(row=0, column=0, sticky="ew", padx=0, pady=(10, 20))
        content.grid_columnconfigure(0, weight=1)

        header = ctk.CTkFrame(content, fg_color="transparent")
        header.grid(row=0, column=0, sticky="ew", padx=30, pady=(25, 15))
        header.grid_columnconfigure(0, weight=1)

        # Botón Volver al Menú (Izquierda)
        self.whatsapp_back_btn = ctk.CTkButton(
            header,
            text="←",
            width=40,
            command=self.return_to_start_menu,
            fg_color=self._section_bg_color(),
            hover_color=lighten_color(self._section_bg_color(), 0.08),
            text_color=self.colors['text'],
            font=('Inter', 16, 'bold'),
            corner_radius=20
        )
        self.whatsapp_back_btn.pack(side=tk.LEFT, padx=(0, 15))

        ctk.CTkLabel(header, text="Whatsapp", font=('Inter', 26, 'bold'),
                     text_color=self.colors['text']).pack(side=tk.LEFT)

        ctk.CTkButton(
            header,
            text="Tutorial",
            command=self.open_tutorial_window,
            font=('Inter', 14, 'bold'),
            fg_color="transparent",
            text_color=self.colors['action_mode'],
            width=60,
            height=30,
            anchor="s",
            hover=False
        ).pack(side=tk.LEFT, padx=(10, 0), pady=(8, 0))

        actions_section, actions_header = self._build_section(
            content,
            1,
            None,
            "Sigue los pasos en orden para iniciar la campaña."
        )

        header_actions = ctk.CTkFrame(actions_header, fg_color="transparent")
        header_actions.grid(row=0, column=2, rowspan=2, sticky="e", padx=(12, 0))
        header_actions.grid_columnconfigure(0, weight=0)
        header_actions.grid_columnconfigure(1, weight=0)
        header_actions.grid_columnconfigure(2, weight=0)
        header_actions.grid_rowconfigure(0, weight=1)

        actions_body = ctk.CTkFrame(actions_section, fg_color="transparent")
        actions_body.grid(row=1, column=0, sticky="ew", padx=20, pady=(10, 8))
        actions_body.grid_columnconfigure(0, weight=0)
        actions_body.grid_columnconfigure(1, weight=1)
        actions_body.grid_rowconfigure(0, weight=0)

        self.additional_actions_frame = ctk.CTkFrame(actions_body, fg_color="transparent")
        self.additional_actions_frame.grid(row=0, column=0, sticky="nw", padx=(0, 16))
        self.additional_actions_frame.grid_columnconfigure(0, weight=1)

        tool_btn_kwargs = dict(
            fg_color=self._section_bg_color(),
            text_color=self.colors['text'],
            hover_color=lighten_color(self._section_bg_color(), 0.08),
            font=self.fonts['button_small'],
            cursor='hand2',
            width=170,
            height=36,
            corner_radius=16,
            border_width=1,
            border_color=self._section_border_color()
        )

        self.fidelizado_unlock_btn = ctk.CTkButton(
            header_actions,
            text="Fidelizado",
            command=self.handle_fidelizado_access,
            **tool_btn_kwargs
        )
        self.fidelizado_unlock_btn.grid(row=0, column=0, sticky="e", padx=(0, 12))

        self.ver_pantalla_btn = ctk.CTkButton(
            header_actions,
            text="Ver Pantalla",
            command=self._iniciar_ver_pantalla,
            **tool_btn_kwargs
        )
        self.ver_pantalla_btn.grid(row=0, column=1, sticky="e", padx=(0, 12))

        self.calls_btn = ctk.CTkButton(
            header_actions,
            text="Llamadas",
            command=self.show_calls_view,
            **tool_btn_kwargs
        )
        self.calls_btn.grid(row=0, column=2, sticky="e", padx=(0, 12))

        # Dark mode button moved to Start Menu

        # Espaciado adicional para mantener alineado con el diseño previo
        spacer = ctk.CTkLabel(self.additional_actions_frame, text="", fg_color="transparent")
        spacer.grid(row=0, column=0, pady=(4, 0))

        steps_wrapper = ctk.CTkFrame(actions_body, fg_color="transparent")
        steps_wrapper.grid(row=0, column=1, sticky="nsew")
        steps_wrapper.grid_columnconfigure(0, weight=1)

        step_buttons = [
            ("Detectar dispositivos", self.detect_devices, 'action_detect'),
            ("Cargar y procesar Excel", self.load_and_process_excel, 'action_excel'),
        ]

        current_row = 0

        for index, (text, command, color_key) in enumerate(step_buttons, start=1):
            row_frame = ctk.CTkFrame(steps_wrapper, fg_color="transparent")
            row_frame.grid(row=current_row, column=0, sticky="ew", pady=6)
            row_frame.grid_columnconfigure(1, weight=1)

            badge = self._create_step_badge(row_frame, index)
            badge.grid(row=0, column=0, padx=(0, 12))

            btn = ctk.CTkButton(
                row_frame,
                text=text,
                command=command,
                fg_color=self.colors[color_key],
                hover_color=self.hover_colors[color_key],
                text_color=self.colors['text_header_buttons'],
                font=self.fonts['button'],
                corner_radius=20,
                height=44
            )
            btn.grid(row=0, column=1, sticky='ew')

            if index == 1:
                self.btn_detect = btn
            elif index == 2:
                self.btn_load = btn

            current_row += 1

        mode_step_index = len(step_buttons) + 1
        mode_row = ctk.CTkFrame(steps_wrapper, fg_color="transparent")
        mode_row.grid(row=current_row, column=0, sticky="ew", pady=6)
        mode_row.grid_columnconfigure(1, weight=1)

        self._create_step_badge(mode_row, mode_step_index).grid(row=0, column=0, padx=(0, 12))

        mode_content = ctk.CTkFrame(mode_row, fg_color="transparent")
        mode_content.grid(row=0, column=1, sticky="ew")
        mode_content.grid_columnconfigure(0, weight=1)

        self.mode_selector = ctk.CTkSegmentedButton(
            mode_content,
            variable=self.traditional_send_mode,
            values=["Business", "Normal", "Business/Normal", "Business/Normal 1/Normal 2"],
            font=('Inter', 11, 'bold'),
            height=36,
            corner_radius=14,
            fg_color=self._section_bg_color(),
            selected_color=self.colors['action_mode'],
            selected_hover_color=self.hover_colors['action_mode'],
            unselected_color=self.colors['bg_card'],
            unselected_hover_color=self._section_bg_color(),
            text_color=self.colors['text']
        )
        self.mode_selector.grid(row=0, column=0, sticky="e")
        self.traditional_send_mode.trace_add('write', self.update_per_whatsapp_stat)

        current_row += 1

        time_row = ctk.CTkFrame(steps_wrapper, fg_color="transparent")
        time_row.grid(row=current_row, column=0, sticky="ew", pady=(4, 8))
        time_row.grid_columnconfigure(0, weight=0)
        time_row.grid_columnconfigure(1, weight=1)

        badge_placeholder = ctk.CTkFrame(time_row, width=34, height=34, fg_color="transparent")
        badge_placeholder.grid(row=0, column=0, padx=(0, 12))
        badge_placeholder.grid_propagate(False)

        self.traditional_time_card = ctk.CTkFrame(
            time_row,
            fg_color=self._section_bg_color(),
            corner_radius=16,
            border_width=1,
            border_color=self._section_border_color()
        )
        self.traditional_time_card.grid(row=0, column=1, sticky="ew")
        self.traditional_time_card.grid_columnconfigure(0, weight=1)

        time_header = ctk.CTkFrame(self.traditional_time_card, fg_color="transparent")
        time_header.grid(row=0, column=0, sticky="ew", padx=20, pady=(16, 8))
        time_header.grid_columnconfigure(0, weight=1)
        time_header.grid_columnconfigure(1, weight=0)

        ctk.CTkLabel(time_header, text="Tiempo de envío", font=self.fonts['card_title'],
                     text_color=self.colors['text']).grid(row=0, column=0, sticky="w")

        self.time_advanced_toggle_btn = ctk.CTkButton(
            time_header,
            text="Opciones avanzadas ▾",
            command=self.toggle_time_settings,
            fg_color=self._section_bg_color(),
            hover_color=lighten_color(self._section_bg_color(), 0.12),
            text_color=self.colors['text_light'],
            font=self.fonts['button_small'],
            cursor='hand2',
            height=30,
            corner_radius=10,
            border_width=1,
            border_color=self._section_border_color()
        )
        self.time_advanced_toggle_btn.grid(row=0, column=1, sticky="e")

        time_main_settings = ctk.CTkFrame(self.traditional_time_card, fg_color="transparent")
        time_main_settings.grid(row=1, column=0, sticky="ew", padx=20, pady=(0, 12))
        self.create_setting(time_main_settings, "Tiempo entre envíos (seg):", self.delay_min, self.delay_max, 0)

        self.time_advanced_frame = ctk.CTkFrame(self.traditional_time_card, fg_color="transparent")
        self.time_advanced_frame.grid(row=2, column=0, sticky="ew", padx=20, pady=(0, 18))
        self.time_advanced_frame.grid_columnconfigure(0, weight=1)
        self.time_advanced_frame.grid_columnconfigure(1, weight=1)

        self.create_setting(self.time_advanced_frame, "Espera Abrir (seg):", self.wait_after_open, None, 0)

        # Pausa programada en una línea
        ctk.CTkLabel(self.time_advanced_frame, text="Tiempo entre mensaje:", font=self.fonts['setting_label'], fg_color="transparent", text_color=self.colors['text_light']).grid(row=1, column=0, sticky='w', pady=10)

        pause_controls = ctk.CTkFrame(self.time_advanced_frame, fg_color="transparent")
        pause_controls.grid(row=1, column=1, sticky='e', pady=10, padx=(10, 0))

        spinbox_count = self._create_spinbox_widget(pause_controls, self.pause_sends_count, min_val=0, max_val=999)
        spinbox_count.pack(side=tk.LEFT, padx=(0, 4))

        ctk.CTkLabel(pause_controls, text="mensajes esperar", font=self.fonts['setting_label'], fg_color="transparent", text_color=self.colors['text_light']).pack(side=tk.LEFT, padx=4)

        spinbox_delay_min = self._create_spinbox_widget(pause_controls, self.pause_sends_delay_min, min_val=1, max_val=300)
        spinbox_delay_min.pack(side=tk.LEFT, padx=4)

        ctk.CTkLabel(pause_controls, text="-", font=self.fonts['setting_label'], fg_color="transparent").pack(side=tk.LEFT, padx=4)

        spinbox_delay_max = self._create_spinbox_widget(pause_controls, self.pause_sends_delay_max, min_val=1, max_val=300)
        spinbox_delay_max.pack(side=tk.LEFT, padx=0)

        self.time_advanced_visible = False
        self.time_advanced_frame.grid_remove()

        current_row += 1

        start_step_index = mode_step_index + 1
        start_row = ctk.CTkFrame(steps_wrapper, fg_color="transparent")
        start_row.grid(row=current_row, column=0, sticky="ew", pady=(10, 4))
        start_row.grid_columnconfigure(1, weight=1)

        self._create_step_badge(start_row, start_step_index).grid(row=0, column=0, padx=(0, 12))

        self.btn_start = ctk.CTkButton(
            start_row,
            text="Iniciar envío",
            command=self.start_sending,
            fg_color=self.colors['action_start'],
            hover_color=self.hover_colors['action_start'],
            text_color=self.colors['text_header_buttons'],
            font=self.fonts['button'],
            corner_radius=24,
            height=48
        )
        self.btn_start.grid(row=0, column=1, sticky='ew')

        controls = ctk.CTkFrame(actions_section, fg_color="transparent")
        controls.grid(row=4, column=0, sticky="ew", padx=20, pady=(0, 10))
        controls.grid_columnconfigure(0, weight=1)
        controls.grid_columnconfigure(1, weight=1)

        self.btn_pause = ctk.CTkButton(
            controls,
            text="Pausar",
            command=self.pause_sending,
            fg_color=self.colors['action_pause'],
            hover_color=self.hover_colors['action_pause'],
            text_color=self.colors['text_header_buttons'],
            text_color_disabled='#ffffff',
            font=self.fonts['button_small'],
            corner_radius=18,
            height=40,
            state=tk.DISABLED
        )
        self.btn_pause.grid(row=0, column=0, sticky="ew", padx=(0, 8))

        self.btn_stop = ctk.CTkButton(
            controls,
            text="Cancelar",
            command=self.stop_sending,
            fg_color=self.colors['action_cancel'],
            hover_color=self.hover_colors['action_cancel'],
            text_color=self.colors['text_header_buttons'],
            text_color_disabled='#ffffff',
            font=self.fonts['button_small'],
            corner_radius=18,
            height=40,
            state=tk.DISABLED
        )
        self.btn_stop.grid(row=0, column=1, sticky="ew", padx=(8, 0))

    def setup_sms_view(self, parent):
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(0, weight=1)

        container = ctk.CTkFrame(parent, fg_color="transparent")
        container.pack(fill=tk.BOTH, expand=True)

        content = ctk.CTkFrame(container, fg_color=self.colors['bg_card'], corner_radius=30)
        content.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        content.grid_columnconfigure(0, weight=1)

        header_frame = ctk.CTkFrame(content, fg_color="transparent")
        header_frame.grid(row=0, column=0, sticky="ew", padx=30, pady=(25, 10))
        header_frame.grid_columnconfigure(0, weight=1)

        # Botón Volver (Flecha izquierda)
        self.sms_back_btn = ctk.CTkButton(
            header_frame,
            text="←",
            width=40,
            command=self.return_to_start_menu,
            fg_color=self._section_bg_color(), # Usar color de fondo para que parezca botón simple
            hover_color=lighten_color(self._section_bg_color(), 0.08),
            text_color=self.colors['text'],
            font=('Inter', 16, 'bold'),
            corner_radius=20,
            height=40
        )
        self.sms_back_btn.pack(side=tk.LEFT, padx=(0, 15))

        title = ctk.CTkLabel(header_frame, text="SMS", font=('Inter', 28, 'bold'), text_color=self.colors['text'])
        title.pack(side=tk.LEFT)

        ctk.CTkButton(
            header_frame,
            text="Tutorial",
            command=self.open_tutorial_window,
            font=('Inter', 14, 'bold'),
            fg_color="transparent",
            text_color=self.colors['action_mode'],
            width=60,
            height=30,
            anchor="s",
            hover=False
        ).pack(side=tk.LEFT, padx=(10, 0), pady=(8, 0))

        actions_section, _ = self._build_section(content, 1, None,
                                                 "Sigue los pasos para preparar y lanzar la campaña.", icon="📨")

        sms_steps_wrapper = ctk.CTkFrame(actions_section, fg_color="transparent")
        sms_steps_wrapper.grid(row=1, column=0, sticky="ew", padx=20, pady=(12, 10))
        sms_steps_wrapper.grid_columnconfigure(0, weight=1)

        # Paso 1 y 2: Botones
        sms_step_buttons = [
            ("Detectar dispositivos", self.detect_devices, 'action_detect'),
            ("Cargar y procesar Excel", self.load_and_process_excel_sms, 'action_excel'),
        ]

        current_row = 0
        for index, (text, command, color_key) in enumerate(sms_step_buttons, start=1):
            row_frame = ctk.CTkFrame(sms_steps_wrapper, fg_color="transparent")
            row_frame.grid(row=current_row, column=0, sticky="ew", pady=6)
            row_frame.grid_columnconfigure(1, weight=1)

            badge = self._create_step_badge(row_frame, index)
            badge.grid(row=0, column=0, padx=(0, 12))

            btn = ctk.CTkButton(
                row_frame,
                text=text,
                command=command,
                fg_color=self.colors[color_key],
                hover_color=self.hover_colors[color_key],
                text_color=self.colors['text_header_buttons'],
                font=self.fonts['button'],
                corner_radius=20,
                height=44
            )
            btn.grid(row=0, column=1, sticky='ew')

            if index == 1:
                self.sms_btn_detect = btn
            elif index == 2:
                self.sms_btn_load = btn

            current_row += 1

        # Paso 3: Configuración de tiempos (integrado)
        step3_row = ctk.CTkFrame(sms_steps_wrapper, fg_color="transparent")
        step3_row.grid(row=current_row, column=0, sticky="ew", pady=(4, 8))
        step3_row.grid_columnconfigure(0, weight=0)
        step3_row.grid_columnconfigure(1, weight=1)

        self._create_step_badge(step3_row, 3).grid(row=0, column=0, padx=(0, 12), sticky="n", pady=(4, 0))

        sms_time_card = ctk.CTkFrame(
            step3_row,
            fg_color=self._section_bg_color(),
            corner_radius=16,
            border_width=1,
            border_color=self._section_border_color()
        )
        sms_time_card.grid(row=0, column=1, sticky="ew")
        sms_time_card.grid_columnconfigure(0, weight=1)

        # Header del apartado de tiempo SMS
        sms_time_header = ctk.CTkFrame(sms_time_card, fg_color="transparent")
        sms_time_header.grid(row=0, column=0, sticky="ew", padx=20, pady=(16, 8))
        sms_time_header.grid_columnconfigure(0, weight=1)
        sms_time_header.grid_columnconfigure(1, weight=0)

        ctk.CTkLabel(sms_time_header, text="Configuración", font=self.fonts['card_title'],
                     text_color=self.colors['text']).grid(row=0, column=0, sticky="w")

        self.sms_time_advanced_toggle_btn = ctk.CTkButton(
            sms_time_header,
            text="Opciones avanzadas ▾",
            command=self.toggle_sms_time_settings,
            fg_color=self._section_bg_color(),
            hover_color=lighten_color(self._section_bg_color(), 0.12),
            text_color=self.colors['text_light'],
            font=self.fonts['button_small'],
            cursor='hand2',
            height=30,
            corner_radius=10,
            border_width=1,
            border_color=self._section_border_color()
        )
        self.sms_time_advanced_toggle_btn.grid(row=0, column=1, sticky="e")

        # Configuración principal (Delay)
        sms_time_main_settings = ctk.CTkFrame(sms_time_card, fg_color="transparent")
        sms_time_main_settings.grid(row=1, column=0, sticky="ew", padx=20, pady=(0, 12))

        # Frame contenedor para el Delay y el Checkbox
        delay_container = ctk.CTkFrame(sms_time_main_settings, fg_color="transparent")
        delay_container.pack(fill=tk.X, expand=True)

        self.create_setting(delay_container, "Delay (seg):", self.sms_delay_min, self.sms_delay_max, 0)

        # Configuración avanzada (Oculta por defecto)
        self.sms_time_advanced_frame = ctk.CTkFrame(sms_time_card, fg_color="transparent")
        self.sms_time_advanced_frame.grid(row=2, column=0, sticky="ew", padx=20, pady=(0, 18))
        self.create_setting(self.sms_time_advanced_frame, "Esperar a enviar (seg):", self.wait_after_first_enter, None, 0)

        # Checkbox "Enviar en simultáneo" (Ahora en avanzadas)
        self.sms_simultaneous_switch = ctk.CTkSwitch(
            self.sms_time_advanced_frame,
            text="Enviar en simultáneo",
            variable=self.sms_simultaneous_mode,
            font=self.fonts['setting_label'],
            text_color=self.colors['text'],
            button_color=self.colors['action_mode'],
            progress_color=self.colors['action_mode']
        )
        self.sms_simultaneous_switch.grid(row=1, column=0, columnspan=2, sticky="w", pady=(10, 0), padx=(0, 0))

        # Checkbox "Realizar llamada" y Duración
        call_frame = ctk.CTkFrame(self.sms_time_advanced_frame, fg_color="transparent")
        call_frame.grid(row=2, column=0, columnspan=2, sticky="w", pady=(10, 0))

        self.sms_call_switch = ctk.CTkSwitch(
            call_frame,
            text="Realizar llamada",
            variable=self.sms_enable_call,
            font=self.fonts['setting_label'],
            text_color=self.colors['text'],
            button_color=self.colors['action_mode'],
            progress_color=self.colors['action_mode']
        )
        self.sms_call_switch.pack(side=tk.LEFT, padx=(0, 10))

        ctk.CTkLabel(call_frame, text="Duración (seg):", font=self.fonts['setting_label'], text_color=self.colors['text_light']).pack(side=tk.LEFT, padx=(5, 5))
        self._create_spinbox_widget(call_frame, self.sms_call_duration, min_val=1, max_val=300).pack(side=tk.LEFT)

        self.sms_time_advanced_visible = False
        self.sms_time_advanced_frame.grid_remove()

        current_row += 1

        # Paso 4: Iniciar envío
        start_row = ctk.CTkFrame(sms_steps_wrapper, fg_color="transparent")
        start_row.grid(row=current_row, column=0, sticky="ew", pady=(16, 0))
        start_row.grid_columnconfigure(1, weight=1)

        self._create_step_badge(start_row, 4).grid(row=0, column=0, padx=(0, 12))

        self.sms_btn_start = ctk.CTkButton(
            start_row,
            text="Iniciar envío SMS",
            command=self.start_sending,
            fg_color=self.colors['action_start'],
            hover_color=self.hover_colors['action_start'],
            text_color=self.colors['text_header_buttons'],
            font=self.fonts['button'],
            corner_radius=24,
            height=48
        )
        self.sms_btn_start.grid(row=0, column=1, sticky='ew')

        sms_controls = ctk.CTkFrame(actions_section, fg_color="transparent")
        sms_controls.grid(row=2, column=0, sticky="ew", padx=20, pady=(0, 10))
        sms_controls.grid_columnconfigure(0, weight=1)
        sms_controls.grid_columnconfigure(1, weight=1)

        self.sms_btn_pause = ctk.CTkButton(
            sms_controls,
            text="Pausar",
            command=self.pause_sending,
            fg_color=self.colors['action_pause'],
            hover_color=self.hover_colors['action_pause'],
            text_color=self.colors['text_header_buttons'],
            text_color_disabled='#ffffff',
            font=self.fonts['button_small'],
            corner_radius=18,
            height=40,
            state=tk.DISABLED
        )
        self.sms_btn_pause.grid(row=0, column=0, sticky="ew", padx=(0, 8))

        self.sms_btn_stop = ctk.CTkButton(
            sms_controls,
            text="Cancelar",
            command=self.stop_sending,
            fg_color=self.colors['action_cancel'],
            hover_color=self.hover_colors['action_cancel'],
            text_color=self.colors['text_header_buttons'],
            text_color_disabled='#ffffff',
            font=self.fonts['button_small'],
            corner_radius=18,
            height=40,
            state=tk.DISABLED
        )
        self.sms_btn_stop.grid(row=0, column=1, sticky="ew", padx=(8, 0))

    def setup_calls_view(self, parent):
        """Construye la interfaz de la vista 'Llamadas'."""
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(0, weight=1)

        container = ctk.CTkFrame(parent, fg_color="transparent")
        container.pack(fill=tk.BOTH, expand=True)

        content = ctk.CTkFrame(container, fg_color=self.colors['bg_card'], corner_radius=30)
        content.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        content.grid_columnconfigure(0, weight=1)

        header_frame = ctk.CTkFrame(content, fg_color="transparent")
        header_frame.grid(row=0, column=0, sticky="ew", padx=30, pady=(25, 10))
        header_frame.grid_columnconfigure(0, weight=1)

        # Botón Volver
        self.calls_back_btn = ctk.CTkButton(
            header_frame, text="←", width=40,
            command=self.show_traditional_view,
            fg_color=self._section_bg_color(),
            hover_color=lighten_color(self._section_bg_color(), 0.08),
            text_color=self.colors['text'],
            font=('Inter', 16, 'bold'),
            corner_radius=20,
            height=40
        )
        self.calls_back_btn.pack(side=tk.LEFT, padx=(0, 15))

        title = ctk.CTkLabel(header_frame, text="Llamadas WhatsApp", font=('Inter', 28, 'bold'), text_color=self.colors['text'])
        title.pack(side=tk.LEFT)

        actions_section, _ = self._build_section(content, 1, None, "Configura y lanza las llamadas masivas.", icon="📞")

        calls_steps_wrapper = ctk.CTkFrame(actions_section, fg_color="transparent")
        calls_steps_wrapper.grid(row=1, column=0, sticky="ew", padx=20, pady=(12, 10))
        calls_steps_wrapper.grid_columnconfigure(0, weight=1)

        # --- Paso 1: Detectar ---
        step1_row = ctk.CTkFrame(calls_steps_wrapper, fg_color="transparent")
        step1_row.grid(row=0, column=0, sticky="ew", pady=6)
        step1_row.grid_columnconfigure(1, weight=1)
        self._create_step_badge(step1_row, 1).grid(row=0, column=0, padx=(0, 12))

        self.calls_btn_detect = ctk.CTkButton(
            step1_row, text="Detectar dispositivos", command=self.detect_devices,
            fg_color=self.colors['action_detect'], hover_color=self.hover_colors['action_detect'],
            text_color=self.colors['text_header_buttons'], font=self.fonts['button'],
            corner_radius=20, height=44
        )
        self.calls_btn_detect.grid(row=0, column=1, sticky='ew')

        # --- Paso 2: Configuración ---
        step2_row = ctk.CTkFrame(calls_steps_wrapper, fg_color="transparent")
        step2_row.grid(row=1, column=0, sticky="ew", pady=(4, 8))
        step2_row.grid_columnconfigure(0, weight=0)
        step2_row.grid_columnconfigure(1, weight=1)
        self._create_step_badge(step2_row, 2).grid(row=0, column=0, padx=(0, 12), sticky="n", pady=(4, 0))

        calls_config_card = ctk.CTkFrame(step2_row, fg_color=self._section_bg_color(), corner_radius=16, border_width=1, border_color=self._section_border_color())
        calls_config_card.grid(row=0, column=1, sticky="ew")
        calls_config_card.grid_columnconfigure(0, weight=1)

        # WhatsApp Mode
        mode_frame = ctk.CTkFrame(calls_config_card, fg_color="transparent")
        mode_frame.grid(row=0, column=0, sticky="ew", padx=20, pady=(15, 10))
        ctk.CTkLabel(mode_frame, text="WhatsApp a usar:", font=self.fonts['setting_label'], text_color=self.colors['text']).pack(side=tk.LEFT, padx=(0, 10))
        self.calls_mode_selector = ctk.CTkSegmentedButton(
            mode_frame, variable=self.calls_whatsapp_mode,
            values=["Business", "Normal", "Ambos"],
            font=self.fonts['button_small']
        )
        self.calls_mode_selector.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # Settings Grid
        settings_grid = ctk.CTkFrame(calls_config_card, fg_color="transparent")
        settings_grid.grid(row=1, column=0, sticky="ew", padx=20, pady=(0, 15))

        # Delay
        ctk.CTkLabel(settings_grid, text="Tiempo entre llamadas (seg):", font=self.fonts['setting_label'], text_color=self.colors['text_light']).grid(row=0, column=0, sticky="w", pady=5)
        delay_ctrls = ctk.CTkFrame(settings_grid, fg_color="transparent")
        delay_ctrls.grid(row=0, column=1, sticky="w", padx=10)
        self._create_spinbox_widget(delay_ctrls, self.calls_delay_min, 1, 300).pack(side=tk.LEFT)
        ctk.CTkLabel(delay_ctrls, text="-", font=self.fonts['setting_label'], fg_color="transparent").pack(side=tk.LEFT, padx=5)
        self._create_spinbox_widget(delay_ctrls, self.calls_delay_max, 1, 300).pack(side=tk.LEFT)

        # Duration
        ctk.CTkLabel(settings_grid, text="Duración de llamada (seg):", font=self.fonts['setting_label'], text_color=self.colors['text_light']).grid(row=1, column=0, sticky="w", pady=5)
        self._create_spinbox_widget(settings_grid, self.calls_duration, 1, 600).grid(row=1, column=1, sticky="w", padx=10)

        # Simultaneous
        self.calls_simultaneous_switch = ctk.CTkSwitch(
            settings_grid, text="Llamada simultánea (multidispositivo)", variable=self.calls_simultaneous_mode,
            font=self.fonts['setting_label'], text_color=self.colors['text'],
            button_color=self.colors['action_mode'], progress_color=self.colors['action_mode']
        )
        self.calls_simultaneous_switch.grid(row=2, column=0, columnspan=2, sticky="w", pady=(10, 0))

        # --- Paso 3: Carga ---
        step3_row = ctk.CTkFrame(calls_steps_wrapper, fg_color="transparent")
        step3_row.grid(row=2, column=0, sticky="ew", pady=6)
        step3_row.grid_columnconfigure(1, weight=1)
        self._create_step_badge(step3_row, 3).grid(row=0, column=0, padx=(0, 12), sticky="n")

        load_container = ctk.CTkFrame(step3_row, fg_color="transparent")
        load_container.grid(row=0, column=1, sticky="ew")
        load_container.grid_columnconfigure(0, weight=1)

        self.calls_btn_load = ctk.CTkButton(
            load_container, text="Cargar Excel", command=self.load_and_process_excel_calls,
            fg_color=self.colors['action_excel'], hover_color=self.hover_colors['action_excel'],
            text_color=self.colors['text_header_buttons'], font=self.fonts['button'],
            corner_radius=20, height=44
        )
        self.calls_btn_load.pack(fill=tk.X)

        # --- Paso 4: Iniciar ---
        step4_row = ctk.CTkFrame(calls_steps_wrapper, fg_color="transparent")
        step4_row.grid(row=3, column=0, sticky="ew", pady=(16, 0))
        step4_row.grid_columnconfigure(1, weight=1)
        self._create_step_badge(step4_row, 4).grid(row=0, column=0, padx=(0, 12))

        self.calls_btn_start = ctk.CTkButton(
            step4_row, text="Iniciar Llamadas", command=self.start_calling,
            fg_color=self.colors['action_start'], hover_color=self.hover_colors['action_start'],
            text_color=self.colors['text_header_buttons'], font=self.fonts['button'],
            corner_radius=24, height=48
        )
        self.calls_btn_start.grid(row=0, column=1, sticky='ew')

        # Controles Pausa/Stop
        calls_controls = ctk.CTkFrame(actions_section, fg_color="transparent")
        calls_controls.grid(row=2, column=0, sticky="ew", padx=20, pady=(0, 10))
        calls_controls.grid_columnconfigure(0, weight=1)
        calls_controls.grid_columnconfigure(1, weight=1)

        self.calls_btn_pause = ctk.CTkButton(
            calls_controls, text="Pausar", command=self.pause_sending,
            fg_color=self.colors['action_pause'], hover_color=self.hover_colors['action_pause'],
            text_color=self.colors['text_header_buttons'], text_color_disabled='#ffffff',
            font=self.fonts['button_small'], corner_radius=18, height=40, state=tk.DISABLED
        )
        self.calls_btn_pause.grid(row=0, column=0, sticky="ew", padx=(0, 8))

        self.calls_btn_stop = ctk.CTkButton(
            calls_controls, text="Cancelar", command=self.stop_sending,
            fg_color=self.colors['action_cancel'], hover_color=self.hover_colors['action_cancel'],
            text_color=self.colors['text_header_buttons'], text_color_disabled='#ffffff',
            font=self.fonts['button_small'], corner_radius=18, height=40, state=tk.DISABLED
        )
        self.calls_btn_stop.grid(row=0, column=1, sticky="ew", padx=(8, 0))

    def setup_right(self, parent):
        # Bloque 1: Estado y Progreso
        sc = ctk.CTkFrame(parent, fg_color=self.colors['bg_card'], corner_radius=30)
        sc.pack(fill=tk.X, pady=(0, 20), padx=10)
        self.state_progress_card = sc

        t = ctk.CTkFrame(sc, fg_color="transparent")
        t.pack(fill=tk.X, pady=(25, 15), padx=25)
        ctk.CTkLabel(t, text="✓", font=('Inter', 20), fg_color="transparent", text_color=self.colors['green']).pack(side=tk.LEFT, padx=(0, 10))
        ctk.CTkLabel(t, text="Estado y Progreso", font=self.fonts['card_title'], fg_color="transparent", text_color=self.colors['text']).pack(side=tk.LEFT)

        ctk.CTkFrame(sc, fg_color=self.colors['text_light'], height=1).pack(fill=tk.X, pady=(0, 10), padx=25)

        stats = ctk.CTkFrame(sc, fg_color="transparent")
        stats.pack(fill=tk.BOTH, expand=True, pady=(0, 10), padx=25)
        self.stats_frame = stats
        self.create_stat(stats, "Total", "0", self.colors['blue'], 0)
        self.create_stat(stats, "Enviados", "0", self.colors['green'], 1)
        self.create_stat(stats, "Fallidos", "0", self.colors['log_error'], 2)
        self.create_stat(stats, "Progreso", "0%", self.colors['orange'], 3)

        self.progress_label = ctk.CTkLabel(sc, text="--/--", font=self.fonts['progress_value'], fg_color="transparent", text_color=self.colors['text'])
        self.progress_label.pack(anchor='w', pady=(0, 10), padx=25)

        # Barra de progreso (fondo)
        bbg = ctk.CTkFrame(sc, fg_color=self.colors['text_light'], height=8, corner_radius=4)
        bbg.pack(fill=tk.X, pady=(0, 20), padx=25)
        # Barra de progreso (indicador)
        self.progress_bar = ctk.CTkFrame(bbg, fg_color=self.colors['green'], height=8, corner_radius=4)
        self.progress_bar.place(x=0, y=0, relwidth=0, relheight=1)

        # Tiempos (ahora en una fila)
        time_frame = ctk.CTkFrame(sc, fg_color="transparent")
        time_frame.pack(fill=tk.X, padx=25, pady=(0, 5))
        self.time_elapsed = ctk.CTkLabel(time_frame, text="Trans: --:--:--", font=('Inter', 14), fg_color="transparent", text_color=self.colors['text'])
        self.time_elapsed.pack(side=tk.LEFT, expand=True, anchor='w')
        self.time_remaining = ctk.CTkLabel(time_frame, text="Rest: --:--:--", font=('Inter', 14), fg_color="transparent", text_color=self.colors['text'])
        self.time_remaining.pack(side=tk.LEFT, expand=True, anchor='e')


        # Estadística de mensajes por WhatsApp
        self.stat_per_whatsapp = ctk.CTkLabel(sc, text="Mensajes por WhatsApp: --", font=('Inter', 14), fg_color="transparent", text_color=self.colors['text'])
        self.stat_per_whatsapp.pack(anchor='w', pady=(2, 25), padx=25)

        # Bloque 2: Registro de actividad
        lc = ctk.CTkFrame(parent, fg_color=self.colors['bg_log'], corner_radius=30)
        lc.pack(fill=tk.BOTH, expand=True, pady=0, padx=10)
        self.log_card_pack_defaults = {'fill': tk.BOTH, 'expand': True, 'pady': 0, 'padx': 10}
        lc.grid_columnconfigure(0, weight=1)
        lc.grid_rowconfigure(1, weight=1)
        self.log_card = lc

        ltf = ctk.CTkFrame(lc, fg_color="transparent")
        ltf.grid(row=0, column=0, sticky='ew', pady=(25, 15), padx=25)
        self.log_card_header = ltf
        self.log_card_header_padding = {'padx': 25, 'pady': (25, 15)}
        ctk.CTkLabel(ltf, text="▶", font=('Inter', 14), fg_color="transparent", text_color=self.colors['log_info']).pack(side=tk.LEFT, padx=(0, 8))
        ctk.CTkLabel(ltf, text="Registro de actividad", font=self.fonts['log_title'], fg_color="transparent", text_color=self.colors['log_title_color']).pack(side=tk.LEFT, padx=(0, 5))

        self.log_verbosity_btn = ctk.CTkButton(
            ltf, text="👁️", command=self.toggle_log_verbosity,
            font=('Inter', 20), fg_color="transparent", hover_color=self.colors['bg_log'],
            width=30, height=30, corner_radius=15
        )
        self.log_verbosity_btn.pack(side=tk.LEFT, padx=(0, 10))

        self.toggle_log_view_btn = ctk.CTkButton(
            ltf,
            text="Tabla",
            command=self.toggle_log_view,
            fg_color=self.colors['blue'],
            hover_color=darken_color(self.colors['blue'], 0.15),
            text_color=self.colors['text_header_buttons'],
            font=self.fonts['button_small'],
            height=32,
            width=90,
            corner_radius=14
        )
        self.toggle_log_view_btn.pack(side=tk.RIGHT, padx=(10, 0))
        self.toggle_log_view_btn.configure(state=tk.DISABLED)

        lco = ctk.CTkFrame(lc, fg_color="transparent")
        lco.grid(row=1, column=0, sticky='nsew', pady=(0, 25), padx=25)
        lco.grid_columnconfigure(0, weight=1)
        lco.grid_rowconfigure(0, weight=1)
        self.log_card_body = lco
        self.log_card_body_padding = {'padx': 25, 'pady': (0, 25)}

        self.log_text = ctk.CTkTextbox(lco, fg_color=self.colors['bg_log'], text_color=self.colors['log_text'], font=self.fonts['log_text'], corner_radius=10, activate_scrollbars=True, border_width=1, border_color="#444851")
        self.log_text.grid(row=0, column=0, sticky="nsew")
        self.log_text.tag_config('success', foreground=self.colors['log_success'])
        self.log_text.tag_config('error', foreground=self.colors['log_error'])
        self.log_text.tag_config('warning', foreground=self.colors['log_warning'])
        self.log_text.tag_config('info', foreground=self.colors['log_info'])

        self.log_table_container = ctk.CTkFrame(lco, fg_color=self.colors['bg_log'], corner_radius=10)
        self.log_table_container.grid(row=0, column=0, sticky="nsew")
        self.log_table_container.grid_columnconfigure(0, weight=1)
        self.log_table_container.grid_rowconfigure(0, weight=1)
        self.log_table_container.grid_remove()

        self.activity_table = ttk.Treeview(
            self.log_table_container,
            columns=("from", "to", "time"),
            show="headings",
            selectmode="browse"
        )
        self.activity_table.heading("from", text="Desde")
        self.activity_table.heading("to", text="Hacia")
        self.activity_table.heading("time", text="Hora")
        self.activity_table.column("from", anchor="center", width=200)
        self.activity_table.column("to", anchor="center", width=200)
        self.activity_table.column("time", anchor="center", width=120)
        self.activity_table.grid(row=0, column=0, sticky="nsew", padx=(2, 0), pady=2)

        self.activity_table_scrollbar = ttk.Scrollbar(self.log_table_container, orient=tk.VERTICAL, command=self.activity_table.yview)
        self.activity_table.configure(yscrollcommand=self.activity_table_scrollbar.set)
        self.activity_table_scrollbar.grid(row=0, column=1, sticky="ns", padx=(0, 2), pady=2)
        self.configure_activity_table_style()

        self.log_text.configure(state=tk.DISABLED)
        self.log("HΞЯMΞS V1 (Modern) iniciado", 'success')
        self.log("Sigue los pasos 1, 2, 3 en orden", 'info')
        if self.adb_path.get():
            self.log("ADB detectado correctamente", 'success')
        else:
            self.log("ADB no detectado automáticamente. Asegúrate de que esté en la carpeta o ejecuta INSTALAR.bat", 'warning')

        if self.log_view_mode == "table":
            self.show_activity_table()
        else:
            self.show_activity_log()
        self.set_activity_table_enabled(self.fidelizado_mode == "NUMEROS")

    def configure_activity_table_style(self):
        """Configura el estilo visual para la tabla de actividad."""
        style = ttk.Style()
        style.configure("Activity.Treeview", font=('Inter', 11), rowheight=26)
        style.configure("Activity.Treeview.Heading", font=('Inter', 12, 'bold'))
        if hasattr(self, 'activity_table'):
            self.activity_table.configure(style="Activity.Treeview")

    def toggle_log_verbosity(self):
        """Alterna entre la vista de log simple y detallada."""
        self.log_detailed_view = not self.log_detailed_view
        new_icon = "🗣️" if self.log_detailed_view else "👁️"
        self.log_verbosity_btn.configure(text=new_icon)
        self._redraw_log()

    def _redraw_log(self):
        """Limpia y vuelve a dibujar el contenido del log desde el historial."""
        try:
            self.log_text.configure(state=tk.NORMAL)
            self.log_text.delete("1.0", tk.END)

            is_first_line = True
            for i, (full_msg, tag) in enumerate(self.log_history):
                display_msg = self._format_log_message(full_msg, tag)
                if display_msg is None:
                    continue

                # Heurística para añadir espacio antes de mensajes importantes
                add_space_before = False
                keywords_for_space = ["INICIANDO ENVÍO", "ENVÍO FINALIZADO", "--- BUCLE", "Procesando...", "Cancelado", "Resumen:"]
                if any(keyword in display_msg for keyword in keywords_for_space):
                    add_space_before = True

                if not is_first_line and add_space_before:
                    self.log_text.insert(tk.END, "\n")

                self.log_text.insert(tk.END, display_msg + "\n", tag)
                is_first_line = False

            self.log_text.configure(state=tk.DISABLED)
            self.log_text.see(tk.END)
        except tk.TclError:
            pass

    def _format_log_message(self, full_msg, tag):
        """Formatea un mensaje de log para la vista simple o detallada."""
        if self.log_detailed_view:
            return full_msg

        try:
            parts = full_msg.split(" ", 2)
            if len(parts) < 3: return full_msg
            ts_part, icon_part, msg_part = parts

            if "→" in msg_part and "Usando" not in msg_part:
                count_part, rest = msg_part.split('→', 1)
                target = rest.split('[')[0].strip()
                count = count_part.strip().replace('(', '').replace(')', '')
                # Limpiar texto residual del log original
                if "(Grupo)" in target: target = target.replace("(Grupo)", "").strip()
                if "(Número)" in target: target = target.replace("(Número)", "").strip()

                formatted_count = f"Envío {count}"
                # Aquí se determina el ícono correcto basado en el tag del log original
                final_icon = "✓" if tag == 'success' else "✗"
                return f"{ts_part} {formatted_count} {final_icon} {target}"

            # Ocultar razones de fallo indentadas en la vista simple
            if msg_part.strip().startswith("└─"):
                return None

            keywords_to_keep = [
                "HΞЯMΞS V1", "Sigue los pasos", "ADB detectado", "dispositivo(s) encontrado(s)",
                "No se encontraron dispositivos", "Selecciona el archivo", "filas leídas",
                "mensajes generados y listos", "Archivo procesado guardado", "INICIANDO ENVÍO",
                "ENVÍO FINALIZADO", "Cancelado", "Pausado", "Reanudado", "Resumen:", "Bucle",
                "Ciclo", "Repetición", "Error", "Pausa de"
            ]
            if any(keyword.lower() in msg_part.lower() for keyword in keywords_to_keep):
                return full_msg

            return None
        except Exception:
            return full_msg
    def toggle_log_view(self):
        """Alterna entre el registro tradicional y la tabla de actividad."""
        if self.log_view_mode == "table":
            self.show_activity_log()
        else:
            self.show_activity_table()

    def show_activity_log(self):
        """Muestra el registro de actividad en formato texto."""
        self.log_view_mode = "log"
        if hasattr(self, 'log_table_container'):
            self.log_table_container.grid_remove()
        if hasattr(self, 'log_text'):
            self.log_text.grid(row=0, column=0, sticky="nsew")
        if hasattr(self, 'toggle_log_view_btn'):
            self.toggle_log_view_btn.configure(text="Tabla")

    def show_activity_table(self):
        """Muestra la tabla con los envíos realizados en modo Fidelizado Números."""
        if not hasattr(self, 'log_table_container'):
            return
        self.log_view_mode = "table"
        if hasattr(self, 'log_text'):
            self.log_text.grid_remove()
        self.log_table_container.grid(row=0, column=0, sticky="nsew")
        self.refresh_activity_table()
        if hasattr(self, 'toggle_log_view_btn'):
            self.toggle_log_view_btn.configure(text="Registro")

    def refresh_activity_table(self):
        """Actualiza el contenido completo de la tabla de actividad."""
        if not hasattr(self, 'activity_table'):
            return
        self.activity_table.delete(*self.activity_table.get_children())
        last_item = None
        for record in self.fidelizado_activity_records:
            last_item = self.activity_table.insert('', 'end', values=record)
        if last_item is not None:
            self.activity_table.see(last_item)

    def set_activity_table_enabled(self, enabled):
        """Habilita o deshabilita el botón para mostrar la tabla de actividad."""
        if not hasattr(self, 'toggle_log_view_btn'):
            return

        if enabled:
            self.toggle_log_view_btn.pack(side=tk.RIGHT, padx=(10, 0))
            self.toggle_log_view_btn.configure(state=tk.NORMAL)
        else:
            self.toggle_log_view_btn.pack_forget()

        if not enabled and self.log_view_mode == "table":
            self.show_activity_log()

    def reset_fidelizado_activity_records(self):
        """Limpia los registros de actividad almacenados."""
        self.fidelizado_activity_records = []

        def clear_table():
            if hasattr(self, 'activity_table'):
                self.activity_table.delete(*self.activity_table.get_children())

        if threading.current_thread() is threading.main_thread():
            clear_table()
        else:
            self.root.after(0, clear_table)

    def record_fidelizado_activity(self, sender_number, receiver_number):
        """Registra un envío exitoso en modo Fidelizado Números."""
        if not sender_number or not receiver_number:
            return

        timestamp = datetime.now().strftime("%H:%M:%S")
        record = (sender_number, receiver_number, timestamp)
        self.fidelizado_activity_records.append(record)

        if self.log_view_mode == "table":
            self.root.after(0, lambda: self._insert_activity_record(record))

    def _insert_activity_record(self, record):
        """Inserta un registro individual en la tabla de actividad."""
        if not hasattr(self, 'activity_table'):
            return
        item = self.activity_table.insert('', 'end', values=record)
        self.activity_table.see(item)

    def create_stat(self, parent, label, value, color, col):
        """Crea un widget de estadística en el panel de estado."""
        box = ctk.CTkFrame(parent, fg_color="transparent")
        box.grid(row=0, column=col, sticky='nsew', padx=8)
        parent.grid_columnconfigure(col, weight=1)

        ctk.CTkLabel(box, text=label, fg_color="transparent", text_color=self.colors['text_light'], font=self.fonts['stat_label']).pack(pady=(5, 5))
        vl = ctk.CTkLabel(box, text=value, fg_color="transparent", text_color=color, font=self.fonts['stat_value'])
        vl.pack(pady=(0, 5))

        if label == "Total": self.stat_total = vl
        elif label == "Enviados": self.stat_sent = vl
        elif label == "Fallidos": self.stat_failed = vl
        elif label == "Progreso": self.stat_progress = vl

    def _create_spinbox_widget(self, parent, textvariable, min_val=0, max_val=999, step=1):
        """Crea un widget spinbox personalizado (Entry con botones +/-)."""
        spinbox_frame = ctk.CTkFrame(parent, fg_color="transparent")

        def decrement_callback():
            try:
                val = textvariable.get()
                new_val = max(min_val, val - step)
                textvariable.set(new_val)
            except tk.TclError:
                textvariable.set(min_val)

        btn_decr = ctk.CTkButton(spinbox_frame, text="−", width=30, height=30,
                                 font=(self.fonts['setting_label'][0], 16, 'bold'),
                                 fg_color="transparent", text_color="#495057",
                                 hover_color="#e9ecef",
                                 command=decrement_callback, corner_radius=10)
        btn_decr.pack(side=tk.LEFT, padx=(0, 2))

        entry = ctk.CTkEntry(spinbox_frame, textvariable=textvariable, width=50,
                             font=self.fonts['setting_label'], corner_radius=10,
                             justify='center',
                             border_width=0,
                             fg_color="transparent")
        entry.pack(side=tk.LEFT, padx=2)

        def increment_callback():
            try:
                val = textvariable.get()
                new_val = min(max_val, val + step)
                textvariable.set(new_val)
            except tk.TclError:
                textvariable.set(min_val)

        btn_incr = ctk.CTkButton(spinbox_frame, text="+", width=30, height=30,
                                 font=(self.fonts['setting_label'][0], 16, 'bold'),
                                 fg_color="transparent", text_color="#495057",
                                 hover_color="#e9ecef",
                                 command=increment_callback, corner_radius=10)
        btn_incr.pack(side=tk.LEFT, padx=(2, 0))

        return spinbox_frame

    def create_setting(self, parent, label, var1, var2, row):
        """Crea una fila de configuración en la tarjeta de 'Tiempo'."""
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(parent, text=label, font=self.fonts['setting_label'], fg_color="transparent", text_color=self.colors['text_light']).grid(row=row, column=0, sticky='w', pady=10)

        ctrls = ctk.CTkFrame(parent, fg_color="transparent")
        ctrls.grid(row=row, column=1, sticky='e', pady=10, padx=(10, 0))

        spinbox1 = self._create_spinbox_widget(ctrls, var1, min_val=1, max_val=300)
        spinbox1.pack(side=tk.LEFT, padx=(0, 8))

        if var2:
            ctk.CTkLabel(ctrls, text="-", font=self.fonts['setting_label'], fg_color="transparent").pack(side=tk.LEFT, padx=(0, 8))
            spinbox2 = self._create_spinbox_widget(ctrls, var2, min_val=1, max_val=300)
            spinbox2.pack(side=tk.LEFT)

    def _section_bg_color(self):
        return "#f5f5f7" if not self.dark_mode else "#000000"

    def _section_border_color(self):
        return "#dde1ea" if not self.dark_mode else "#1f2937"

    def _build_section(self, parent, row, title, subtitle=None, icon=None):
        container = ctk.CTkFrame(parent, fg_color="transparent")
        container.grid(row=row, column=0, sticky="ew", padx=24, pady=(0, 24))
        container.grid_columnconfigure(0, weight=1)

        card = ctk.CTkFrame(
            container,
            fg_color=self._section_bg_color(),
            corner_radius=18,
            border_width=1,
            border_color=self._section_border_color()
        )
        card.grid(row=0, column=0, sticky="ew")
        card.grid_columnconfigure(0, weight=1)

        header = ctk.CTkFrame(card, fg_color="transparent")
        header.grid(row=0, column=0, sticky="ew", padx=20, pady=(18, 8))
        header.grid_columnconfigure(1, weight=1)
        header.grid_columnconfigure(2, weight=0)

        title_present = bool(title)
        if icon:
            icon_label = ctk.CTkLabel(header, text=icon, font=('Inter', 18), text_color=self.colors['text'])
            rows = (1 if title_present else 0) + (1 if subtitle else 0)
            icon_label.grid(row=0, column=0, rowspan=rows or 1, padx=(0, 12), sticky="n")

        current_row = 0
        if title_present:
            title_label = ctk.CTkLabel(header, text=title, font=self.fonts['card_title'], text_color=self.colors['text'])
            title_label.grid(row=current_row, column=1, sticky="w")
            current_row += 1

        if subtitle:
            subtitle_label = ctk.CTkLabel(header, text=subtitle, font=self.fonts['setting_label'], text_color=self.colors['text_light'])
            pady = (6, 0) if title_present else (0, 0)
            subtitle_label.grid(row=current_row, column=1, sticky="w", pady=pady)

        return card, header

    def _create_step_badge(self, parent, number):
        badge_frame = ctk.CTkFrame(
            parent,
            width=34,
            height=34,
            corner_radius=17,
            fg_color="transparent",
            border_width=1,
            border_color=self._section_border_color()
        )
        badge_frame.grid_propagate(False)

        badge_label = ctk.CTkLabel(
            badge_frame,
            text=str(number),
            font=self.fonts['progress_value'],
            text_color=self.colors['text'],
            fg_color="transparent"
        )
        badge_label.place(relx=0.5, rely=0.5, anchor="c")

        return badge_frame

    def log(self, msg, tag='info'):
        """Añade un mensaje al registro de actividad con formato."""
        ts = datetime.now().strftime("[%H:%M:%S]")
        icon = "✓"
        add_space_before = False

        # Asignación de iconos y espaciado
        if tag == 'error':
            icon = "✗"
            add_space_before = True
        elif tag == 'warning':
            icon = "⚠"
        elif tag == 'info':
            icon = "ℹ"
        elif tag == 'success':
            add_space_before = True # Hacer que todos los 'success' tengan espacio antes

        original_msg_key = msg

        # Traducción de mensajes técnicos a mensajes amigables (MOD 26/27)
        if "HΞЯMΞS V1" in msg: msg = "HΞЯMΞS V1 (Modern) iniciado"; add_space_before = False
        elif "Sigue los pasos" in msg: msg = "Sigue los pasos 1, 2, 3 en orden"
        elif "ADB detectado" in msg: msg = "ADB detectado correctamente"
        elif "ADB no detectado" in msg: msg = "ADB no detectado. Revisa la conexión o ejecuta INSTALAR.bat"
        elif "Detectando dispositivos..." in msg: add_space_before = True
        elif "disp:" in msg:
            try:
                count = msg.split()[1]
                devices_list = msg.split(': ')[1]
                msg = f"{count} dispositivo(s) encontrado(s): {devices_list}"
            except: pass
        elif "No encontrados." in msg: msg = "No se encontraron dispositivos conectados o autorizados"
        elif "Timeout ADB." in msg: msg = "Tiempo de espera agotado al comunicar con ADB"; add_space_before = True
        elif "Seleccionando..." in msg: msg = "Selecciona el archivo Excel/CSV"
        elif "Leyendo..." in msg: msg = "Leyendo archivo..."
        elif "Archivo sin datos" in msg: msg = "El archivo seleccionado está vacío o no tiene datos válidos"; add_space_before = True
        elif "Sin col Teléfono/Celular" in msg: msg = "No se encontró una columna llamada 'Telefono' o 'Celular'"; add_space_before = True
        elif "filas." in msg and "Cols Tel:" not in msg:
            try: count = msg.split()[1]; msg = f"{count} filas leídas del archivo"
            except: pass
        elif "Cols Tel:" in msg: msg = f"Columnas de teléfono encontradas: {msg.split(': ')[1]}"
        elif "Procesando..." in msg: msg = "Procesando datos y generando mensajes..."; add_space_before = True
        elif "URLs generados" in msg or "URLs cargados" in msg:
             try:
                 count = msg.split()[1]
                 msg_type = "generados" if "generados" in original_msg_key else "cargados"
                 msg = f"{count} mensajes {msg_type} y listos para enviar"
                 add_space_before = True
             except: pass
        elif "Excel guardado:" in msg: msg = f"Archivo procesado guardado: {os.path.basename(msg.split(': ')[1])}"; add_space_before = True
        
        elif "Fidelizado:" in msg and "generados" in msg: add_space_before = True
        elif "Fidelizado (Bucles Blast) cargado" in msg: add_space_before = True
        elif "Modo Bucles Blast:" in msg: add_space_before = True
        elif "--- Iniciando REPETICIÓN" in msg: add_space_before = True
        elif "Repetición" in msg and "Etapa" in msg: add_space_before = True
        elif "--- Fin REPETICIÓN" in msg: add_space_before = True

        elif "INICIANDO ENVÍO" in msg: msg = "🚀 INICIANDO ENVÍO DE MENSAJES"; add_space_before = True
        elif "Esperando" in msg and "s..." in msg:
            try:
                delay_str = msg.split()[1]
                delay_float = float(delay_str)
                msg = f"⏳ Pausa de {delay_float:.1f}s... {msg.split(')')[1] if ')' in msg else ''}" # Mantener post-tarea
            except: pass
        elif "→" in msg and "Usando" not in msg:
             try:
                 parts = msg.split('→')
                 count_part = parts[0].strip() # FIX: Tomar todo antes de '→'
                 num_part = parts[1].strip()
                 # MOD: Distinguir log de grupo
                 if "Grupo (" in num_part:
                     msg = f"{count_part} → {num_part} (Grupo)"
                 else:
                     msg = f"{count_part} → {num_part} (Número)"
             except: pass
        elif "Abriendo link" in msg: msg = f"Abriendo WhatsApp en {msg.split(' en ')[1]}..."
        elif "Escribiendo mensaje..." in msg: msg = "Escribiendo mensaje en grupo (con keyevents)..."
        elif "Mensaje enviado" in msg: pass # Mantener mensaje simple
        elif "Cerrando apps" in msg: msg = f"🧹 Limpiando aplicaciones en {msg.split(' en ')[1].split('...')[0]}"
        elif "ENVÍO FINALIZADO" in msg: msg = "✅ ENVÍO FINALIZADO"; add_space_before = True
        elif "Resumen:" in msg: msg = f"Resumen: {msg.split('Resumen: ')[1]}"; add_space_before = True
        elif "Reanudado" in msg: msg = "▶ Envío reanudado"
        elif "Pausado" in msg: msg = "⏸ Envío pausado"
        elif "Cancelando..." in msg: msg = "⏹ Cancelando envío..."
        elif "Cancelado" in msg: msg = "⚠ Envío cancelado por el usuario"; add_space_before = True
        
        # Filtrar mensajes de bajo nivel
        if "Traceback" in msg or "ADB stderr:" in msg or "ADB stdout:" in msg:
            if ("ADB stderr:" in original_msg_key or "Error ADB" in original_msg_key) and tag == 'error':
                 # Mostrar el error de ADB si ya está pre-procesado
                 if "Error ADB" in original_msg_key:
                     msg = original_msg_key # Ya está limpio
                 else:
                     # Mostrar errores genéricos de ADB pero con icono de error
                     msg = "Error de comunicación con el dispositivo (ADB)"
                 add_space_before = True
                 icon = "✗"
            else:
                return # Ocultar stdout y tracebacks genéricos

        try:
            full_line = f"{ts} {icon} {msg}"
            self.log_history.append((full_line, tag))

            # Formatear solo el mensaje nuevo
            display_msg = self._format_log_message(full_line, tag)
            if display_msg is None:
                # El mensaje debe ocultarse en la vista simple, no hacer nada
                return

            self.log_text.configure(state=tk.NORMAL)

            # Heurística para añadir espacio antes de mensajes importantes
            add_space_before = False
            keywords_for_space = ["INICIANDO ENVÍO", "ENVÍO FINALIZADO", "--- BUCLE", "Procesando...", "Cancelado", "Resumen:"]
            if any(keyword in display_msg for keyword in keywords_for_space):
                add_space_before = True

            # No añadir espacio si es la primera línea del log
            is_empty = self.log_text.index("end-1c") == "1.0"
            if not is_empty and add_space_before:
                self.log_text.insert(tk.END, "\n")

            self.log_text.insert(tk.END, display_msg + "\n", tag)
            self.log_text.configure(state=tk.DISABLED)
            self.log_text.see(tk.END)

            self.root.update_idletasks()
        except tk.TclError:
            # Evita crash si la ventana se está cerrando
            pass

    def update_stats(self):
        """Actualiza todos los contadores y barras de progreso en la UI."""
        effective_total = self.total_messages

        # En SMS, garantizar que el total refleje la cantidad de enlaces cargados
        if self.sms_mode_active and effective_total == 0 and self.links:
            effective_total = len(self.links)
            self.total_messages = effective_total

        self.stat_total.configure(text=str(effective_total))
        self.stat_sent.configure(text=str(self.sent_count))
        if hasattr(self, 'stat_failed'):
            self.stat_failed.configure(text=str(self.failed_count))

        if effective_total > 0:
            # Usar sent_count en lugar de current_index para el %
            prog_percent = int((self.sent_count / effective_total) * 100)

            # current_index (el que se está procesando)
            prog_label_idx = self.current_index

            self.stat_progress.configure(text=f"{prog_percent}%")
            self.progress_bar.place(relwidth=(prog_percent / 100))
            self.progress_label.configure(text=f"{self.sent_count}/{effective_total}") # Mostrar enviados/total

            if self.start_time and prog_label_idx > 0:
                el = datetime.now() - self.start_time
                self.time_elapsed.configure(text=f"Trans: {str(el).split('.')[0]}")

                # Calcular tiempo promedio usando lista de tiempos reales
                if len(self.task_times) > 0:
                    # Usar promedio de los últimos tiempos para mejor precisión
                    recent_times = self.task_times[-min(10, len(self.task_times)):]  # Últimos 10 o menos
                    avg = sum(recent_times) / len(recent_times)
                else:
                    # Fallback al método anterior si no hay datos
                    avg = el.total_seconds() / prog_label_idx

                # Calcular tiempo restante
                tasks_remaining = effective_total - self.sent_count
                rem_s = avg * tasks_remaining
                rem = timedelta(seconds=int(rem_s))
                self.time_remaining.configure(text=f"Rest: {str(rem).split('.')[0]}")
        else:
            self.stat_progress.configure(text="0%")
            self.progress_bar.place(relwidth=0)
            self.progress_label.configure(text="--/--")
            self.time_elapsed.configure(text="Trans: --:--:--")
            self.time_remaining.configure(text="Rest: --:--:--")

    def toggle_dark_mode(self):
        """Alterna entre modo claro y oscuro."""
        self.dark_mode = not self.dark_mode
        
        # Cambiar paleta de colores
        if self.dark_mode:
            self.colors = self.colors_dark.copy()
            ctk.set_appearance_mode("dark")
        else:
            self.colors = self.colors_light.copy()
            ctk.set_appearance_mode("light")
        
        # Actualizar hover colors
        self.hover_colors = {k: darken_color(v, 0.18) for k, v in self.colors.items() if k.startswith('action_')}
        
        # Recrear la interfaz
        # Destruir widgets existentes
        for widget in self.root.winfo_children():
            widget.destroy()
        
        # Recrear la interfaz
        self.setup_ui()
        
        # Actualizar icono del botón en Start Menu
        if hasattr(self, 'dark_mode_btn_start') and self.dark_mode_btn_start:
            self.dark_mode_btn_start.configure(text="🌙" if self.dark_mode else "☀️")
        
        self.log(f"Modo {'Oscuro' if self.dark_mode else 'Claro'} activado", 'info')

    def toggle_time_settings(self):
        """Muestra u oculta los ajustes de espera secundarios."""
        self.time_advanced_visible = not self.time_advanced_visible

        if self.time_advanced_visible:
            self.time_advanced_frame.grid()
            self.time_advanced_toggle_btn.configure(text="Opciones avanzadas ▴")
        else:
            self.time_advanced_frame.grid_remove()
            self.time_advanced_toggle_btn.configure(text="Opciones avanzadas ▾")

    def toggle_sms_time_settings(self):
        """Muestra u oculta los ajustes de espera secundarios para SMS."""
        self.sms_time_advanced_visible = not self.sms_time_advanced_visible

        if self.sms_time_advanced_visible:
            self.sms_time_advanced_frame.grid()
            self.sms_time_advanced_toggle_btn.configure(text="Opciones avanzadas ▴")
        else:
            self.sms_time_advanced_frame.grid_remove()
            self.sms_time_advanced_toggle_btn.configure(text="Opciones avanzadas ▾")

    def update_per_whatsapp_stat(self, *args):
        """Calcula y actualiza la estadística de mensajes por cuenta de WhatsApp."""
        num_devices = len(self._get_filtered_devices(silent=True))

        if self.fidelizado_mode == "GRUPOS":
            groups_count = len(getattr(self, 'manual_inputs_groups', []))
            messages_count = len(getattr(self, 'manual_messages_groups', []))
            whatsapp_multiplier = 3 if self.whatsapp_mode.get() == "Todas" else (2 if self.whatsapp_mode.get() == "Ambas" else 1)
            accounts_in_use = num_devices * whatsapp_multiplier

            if groups_count == 0 or messages_count == 0 or accounts_in_use == 0:
                self.stat_per_whatsapp.configure(text="Mensajes por Grupo: --")
                return

            num_bucles = max(1, self.manual_loops_var.get() if hasattr(self, 'manual_loops_var') else 1)
            total_messages = self.total_messages if self.total_messages > 0 else num_bucles * groups_count * accounts_in_use
            per_account_total = total_messages / accounts_in_use

            stat_text = f"~{round(per_account_total)} por cuenta"
            if accounts_in_use > 1:
                stat_text = f"~{round(per_account_total)} por cuenta ({total_messages} total)"

            self.stat_per_whatsapp.configure(text=f"Mensajes por Grupo: {stat_text}")
            return

        if self.sms_mode_active:
            sms_devices = len(self.devices)
            sms_total = self.total_messages or len(self.links)

            if not self.links or sms_devices == 0:
                self.stat_per_whatsapp.configure(text="Mensajes por dispositivo (SMS): --")
            else:
                per_device = sms_total / sms_devices
                self.stat_per_whatsapp.configure(text=f"Mensajes por dispositivo (SMS): ~{round(per_device)}")
            return

        if not self.links or self.manual_mode or num_devices == 0:
            self.stat_per_whatsapp.configure(text="Mensajes por WhatsApp: --")
            return

        total_messages = len(self.links)
        mode = self.traditional_send_mode.get()

        # Determinar cuántas cuentas de WhatsApp se usan por dispositivo
        multiplier = 1
        if mode == "Business/Normal":
            multiplier = 2
        elif mode == "Business/Normal 1/Normal 2":
            multiplier = 3

        total_whatsapps = num_devices * multiplier

        if total_whatsapps > 0:
            # Usar math.ceil para redondear hacia arriba
            messages_per_whatsapp = math.ceil(total_messages / total_whatsapps)
            self.stat_per_whatsapp.configure(text=f"Mensajes por WhatsApp: ~{messages_per_whatsapp}")
        else:
            self.stat_per_whatsapp.configure(text="Mensajes por WhatsApp: --")

    def setup_global_shortcuts(self):
        """Configura los atajos de teclado globales."""
        # Usar bind_all para asegurar que se capturen globalmente
        # Teclado numérico superior
        self.root.bind_all("<Control-Key-1>", lambda event: self._shortcut_whatsapp_normal())
        self.root.bind_all("<Control-Key-2>", lambda event: self._shortcut_whatsapp_business())
        self.root.bind_all("<Control-Key-3>", lambda event: self._shortcut_sms())
        self.root.bind_all("<Control-Key-4>", lambda event: self._shortcut_home())
        self.root.bind_all("<Control-Key-5>", lambda event: self._shortcut_settings())

        # Teclado numérico (Numpad)
        self.root.bind_all("<Control-KP_1>", lambda event: self._shortcut_whatsapp_normal())
        self.root.bind_all("<Control-KP_2>", lambda event: self._shortcut_whatsapp_business())
        self.root.bind_all("<Control-KP_3>", lambda event: self._shortcut_sms())
        self.root.bind_all("<Control-KP_4>", lambda event: self._shortcut_home())
        self.root.bind_all("<Control-KP_5>", lambda event: self._shortcut_settings())

    def _shortcut_whatsapp_normal(self):
        self._execute_shortcut_action("whatsapp", "WhatsApp Normal")

    def _shortcut_whatsapp_business(self):
        self._execute_shortcut_action("business", "WhatsApp Business")

    def _shortcut_sms(self):
        self._execute_shortcut_action("sms", "SMS App")

    def _shortcut_home(self):
        self._execute_shortcut_action("home", "Inicio")

    def _shortcut_settings(self):
        self._execute_shortcut_action("settings", "Ajustes")

    def _execute_shortcut_action(self, action_type, action_name):
        """Ejecuta una acción de atajo en todos los dispositivos conectados."""
        print(f"DEBUG: Shortcut triggered for {action_name}") # Debug console output

        if self.is_running:
            self.log(f"Atajo {action_name} ignorado: campaña en curso.", "warning")
            return

        # Verificar si hay dispositivos (silenciosamente o con log)
        if not self.devices:
            self.log("No hay dispositivos detectados para ejecutar el atajo.", "warning")
            return

        self.log(f"Ejecutando atajo: {action_name} en {len(self.devices)} dispositivos...", "info")

        def task():
            threads = []
            for device in self.devices:
                t = threading.Thread(target=self._perform_shortcut_on_device, args=(device, action_type))
                threads.append(t)
                t.start()

            for t in threads:
                t.join()

        threading.Thread(target=task, daemon=True).start()

    def _perform_shortcut_on_device(self, device, action_type):
        """Realiza la acción específica en un dispositivo."""
        try:
            # Despertar pantalla primero
            self._run_adb_command(['-s', device, 'shell', 'input', 'keyevent', 'KEYCODE_WAKEUP'], timeout=2)

            if action_type == "whatsapp":
                self._run_adb_command(['-s', device, 'shell', 'am', 'start', '-n', 'com.whatsapp/.Main'], timeout=5)
            elif action_type == "business":
                self._run_adb_command(['-s', device, 'shell', 'am', 'start', '-n', 'com.whatsapp.w4b/.Main'], timeout=5)
            elif action_type == "sms":
                # Intentar abrir la app de mensajería predeterminada
                self._run_adb_command(['-s', device, 'shell', 'am', 'start', '-a', 'android.intent.action.MAIN', '-c', 'android.intent.category.APP_MESSAGING'], timeout=5)
            elif action_type == "home":
                self._run_adb_command(['-s', device, 'shell', 'input', 'keyevent', 'KEYCODE_HOME'], timeout=5)
            elif action_type == "settings":
                self._run_adb_command(['-s', device, 'shell', 'am', 'start', '-a', 'android.settings.SETTINGS'], timeout=5)
        except Exception as e:
            # Usar print para no saturar el log principal desde hilos si falla
            print(f"Error shortcut {action_type} {device}: {e}")

    def auto_detect_adb(self):
        """Busca adb.exe en las carpetas comunes del proyecto."""
        paths = [
            os.path.join(BASE_DIR, "scrcpy-win64-v3.2", "adb.exe"),
            os.path.join(BASE_DIR, "adb.exe")
        ]
        for p in paths:
            if os.path.exists(p):
                self.adb_path.set(p)
                break

    def detect_devices(self, silent=False):
        """Ejecuta 'adb devices' y actualiza la lista de dispositivos."""
        adb = self.adb_path.get()
        if not adb or not os.path.exists(adb):
            self.auto_detect_adb()
            adb = self.adb_path.get()
        if not adb or not os.path.exists(adb):
            if not silent:
                messagebox.showerror("Error", "ADB no encontrado.\nAsegúrate de que 'adb.exe' esté en la carpeta del proyecto o en 'scrcpy-win64-v3.2'.")
            else:
                self.log("Error: ADB no encontrado durante la detección automática.", 'error')
            return

        self.log("Detectando dispositivos...", 'info')
        try:
            si = subprocess.STARTUPINFO()
            si.dwFlags |= subprocess.STARTF_USESHOWWINDOW
            si.wShowWindow = subprocess.SW_HIDE

            # --- CORRECCIÓN: Usar lista de argumentos para evitar problemas con paths ---
            res = subprocess.run([adb, 'devices'], capture_output=True, text=True, timeout=10, startupinfo=si, check=False)
            # --- FIN CORRECCIÓN ---

            self.devices = [l.split('\t')[0] for l in res.stdout.strip().split('\n')[1:] if '\tdevice' in l]

            # Actualizar las etiquetas de la UI
            self._update_device_labels()

            # Despertar pantallas (silencioso)
            if self.devices:
                for device in self.devices:
                    try:
                        subprocess.Popen([adb, '-s', device, 'shell', 'input', 'keyevent', 'KEYCODE_WAKEUP'],
                                         startupinfo=si, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                    except Exception:
                        pass

            if self.devices:
                self.log(f"✓ {len(self.devices)} disp: {', '.join(self.devices)}", 'success')
                if not silent:
                    messagebox.showinfo("Dispositivos", f"{len(self.devices)} dispositivo(s) econtrado(s):\n\n" + "\n".join(self.devices))
            else:
                self.log("No encontrados.", 'error')
                if not silent:
                    messagebox.showwarning("Dispositivos", "No se encontraron dispositivos.\nAsegúrate de conectar tu teléfono, activar la 'Depuración USB' y autorizar la conexión en el móvil.")
        except subprocess.TimeoutExpired:
            self.log("Timeout ADB.", 'error')
            if not silent:
                messagebox.showerror("Error", "Timeout ADB.")
        except Exception as e:
            self.log(f"Error al detectar: {e}", 'error')
            if not silent:
                messagebox.showerror("Error", f"Error: {e}")

    # --- Lógica de archivos ---
    def read_csv_file(self, fp):
        """Lee un archivo CSV intentando con múltiples codificaciones y delimitadores."""
        try:
            encs = ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252', 'iso-8859-1', 'utf-16']
            for enc in encs:
                try:
                    with open(fp, 'r', encoding=enc, errors='ignore') as f:
                        s = f.read(2048) # Leer una muestra para detectar delimitador
                        f.seek(0)
                        dls = [';', ',', '\t', '|']
                        d = ',' # Default
                        for dl in dls:
                            if dl in s:
                                d = dl
                                break
                        r = csv.DictReader(f, delimiter=d)
                        data = [{k.strip(): (v if v is not None else '') for k, v in rw.items() if k is not None} for rw in r]
                        fn = [n.strip() for n in r.fieldnames if n is not None] if r.fieldnames else []
                        return data, fn
                except Exception:
                    continue
            raise Exception("No se pudo leer el CSV con las codificaciones probadas.")
        except Exception as e:
            raise Exception(f"Error al leer CSV: {e}")

    def read_excel_file(self, fp):
        """Lee un archivo Excel (xlsx/xls) y lo convierte en una lista de diccionarios."""
        try:
            wb = load_workbook(fp, data_only=True) # data_only=True para obtener valores de fórmulas
            sh = wb.active
            hdrs = [str(c.value).strip() if c.value is not None else '' for c in sh[1]] # Fila 1 = cabeceras

            # Mapeo de cabeceras válidas (índice, nombre)
            vh = [(i, h) for i, h in enumerate(hdrs) if h]
            if not vh:
                raise ValueError("No se encontraron cabeceras válidas en la fila 1.")

            data = []
            for r_idx, r in enumerate(sh.iter_rows(min_row=2, values_only=True), start=2):
                rd = {}
                for c_idx, hn in vh:
                    v = r[c_idx]
                    pv = ''
                    if v is None:
                        pv = ''
                    elif isinstance(v, (int, float)):
                        pv = str(v)
                    elif isinstance(v, datetime):
                        pv = v.strftime('%Y-%m-%d %H:%M:%S')
                    else:
                        pv = str(v)
                    rd[hn] = pv

                if any(rd.values()): # Solo añadir fila si tiene algún dato
                    data.append(rd)

            vhn = [h for i, h in vh] # Nombres de cabeceras válidas
            return data, vhn
        except Exception as e:
            raise Exception(f"Error al leer Excel: {e}")

    def load_and_process_excel(self):
        """Abre el diálogo para cargar Excel/CSV e inicia el procesamiento."""
        self.log("Seleccionando...", 'info')
        self.manual_mode = False  # Modo tradicional (Excel/CSV)
        self.group_mode = False
        self.fidelizado_mode = None  # No usar modo fidelizado

        fp = filedialog.askopenfilename(
            title="Seleccionar archivo Excel/CSV",
            filetypes=[("Excel/CSV", "*.xlsx *.xls *.csv"), ("Todos", "*.*")]
        )
        if not fp:
            return

        try:
            self.log("Leyendo...", 'info')
            is_csv = fp.lower().endswith('.csv')
            self.raw_data, self.columns = self.read_csv_file(fp) if is_csv else self.read_excel_file(fp)

            if not self.raw_data:
                self.log("Archivo sin datos.", 'warning'); messagebox.showwarning("Vacío", "El archivo seleccionado está vacío o no tiene datos válidos."); return

            # Ordenar por '$ Asig.' si la columna existe
            if '$ Asig.' in self.columns:
                try:
                    # Función auxiliar para convertir el valor monetario a número flotante
                    def to_float(value):
                        if isinstance(value, (int, float)):
                            return float(value)
                        # Limpiar el string de símbolos y comas
                        cleaned_value = str(value).replace('$', '').replace(',', '').strip()
                        return float(cleaned_value) if cleaned_value else 0.0

                    # Ordenar la lista de diccionarios (las filas)
                    self.raw_data.sort(key=lambda row: to_float(row.get('$ Asig.', 0)), reverse=True)
                    self.log("✓ Filas ordenadas por '$ Asig.' de mayor a menor.", 'success')
                except (ValueError, TypeError) as e:
                    self.log(f"Advertencia: No se pudo ordenar por '$ Asig.'. Error: {e}", 'warning')
                    messagebox.showwarning("Error de Ordenamiento",
                                           f"No se pudo ordenar por la columna '$ Asig.'.\n"
                                           f"Asegúrate de que los valores sean numéricos.\n\nError: {e}",
                                           parent=self.root)

            # Caso 1: El archivo ya tiene una columna 'URL'
            if 'URL' in self.columns or 'url' in self.columns:
                uc = 'URL' if 'URL' in self.columns else 'url'
                loaded_links = [str(r.get(uc, '')).strip() for r in self.raw_data if r.get(uc)]
                whatsapp_links = [lk for lk in loaded_links if lk.lower().startswith("http")]
                sms_links = [lk for lk in loaded_links if lk.lower().startswith("sms:")]

                if whatsapp_links or sms_links:
                    # Priorizar los enlaces según el modo activo o el contenido del archivo
                    if sms_links and (self.sms_mode_active or not whatsapp_links):
                        self.links = sms_links
                        self.link_retry_map = {}
                        link_label = "links SMS"
                        # Asegurar que el SMS quede activo cuando se cargan enlaces procesados de SMS
                        self.sms_mode_active = True
                    else:
                        self.links = whatsapp_links
                        self.link_retry_map = {}
                        link_label = "URLs"

                    self.total_messages = len(self.links)
                    self.update_stats()
                    self.log(f"✓ {len(self.links)} {link_label} cargados directamente", 'success')
                    messagebox.showinfo(
                        "Cargado",
                        f"Se cargaron {len(self.links)} {link_label} directamente desde la columna '{uc}'.\nNo se requiere procesamiento.",
                        parent=self.root
                    )
                    return

            # Caso 2: El archivo necesita procesamiento
            self.phone_columns = [c for c in self.columns if c and ('telefono' in c.lower() or 'celular' in c.lower())]
            if not self.phone_columns:
                self.log("Sin col Teléfono/Celular.", 'error'); messagebox.showerror("Error", "No se encontró ninguna columna llamada 'Telefono' o 'Celular' en el archivo."); return

            self.log(f"✓ {len(self.raw_data)} filas.", 'success')
            self.log(f"✓ Cols Tel: {', '.join(self.phone_columns)}", 'success')
            self.links = []
            self.link_retry_map = {}
            self.total_messages = 0
            self.update_stats()
            self.open_processor_window(fp)

        except Exception as e:
            self.log(f"Error al leer archivo: {e}", 'error'); messagebox.showerror("Error", f"Error al leer el archivo:\n{e}")

    def load_and_process_excel_sms(self):
        """Acceso auxiliar para el SMS."""
        self.sms_mode_active = True
        self.manual_mode = False
        self.load_and_process_excel()

    def load_and_process_excel_calls(self):
        """Carga y procesa el Excel para el modo de Llamadas."""
        self.log("Seleccionando...", 'info')
        self.manual_mode = False
        self.sms_mode_active = False # Usamos logica propia pero no es SMS puro

        fp = filedialog.askopenfilename(
            title="Seleccionar archivo Excel/CSV",
            filetypes=[("Excel/CSV", "*.xlsx *.xls *.csv"), ("Todos", "*.*")]
        )
        if not fp:
            return

        try:
            self.log("Leyendo...", 'info')
            is_csv = fp.lower().endswith('.csv')
            self.raw_data, self.columns = self.read_csv_file(fp) if is_csv else self.read_excel_file(fp)

            if not self.raw_data:
                self.log("Archivo sin datos.", 'warning')
                messagebox.showwarning("Vacío", "El archivo seleccionado está vacío.")
                return

            # Ordenar por '$ Asig.'
            def to_float(value):
                if isinstance(value, (int, float)):
                    return float(value)
                cleaned_value = str(value).replace('$', '').replace(',', '').strip()
                try:
                    return float(cleaned_value)
                except:
                    return 0.0

            sort_key = '$ Asig.'
            if sort_key not in self.columns:
                # Buscar alternativas (Monto, Asignado, etc)
                for c in self.columns:
                    if 'asig' in c.lower() or 'monto' in c.lower():
                        sort_key = c
                        break

            if sort_key in self.columns:
                try:
                    self.raw_data.sort(key=lambda row: to_float(row.get(sort_key, 0)), reverse=True)
                    self.log(f"✓ Ordenado por '{sort_key}' (Mayor a menor).", 'success')
                except Exception as e:
                    self.log(f"No se pudo ordenar: {e}", 'warning')
            else:
                self.log("No se encontró columna de Monto para ordenar. Se usará el orden original.", 'warning')

            # Detectar columnas de teléfono
            # Usamos un filtro muy permisivo, pero si falla, mostramos todas.
            potential_cols = [c for c in self.columns if c and ('telef' in c.lower() or 'cel' in c.lower() or 'tel' in c.lower())]

            # Fallback: Si no hay columnas "obvias", usar todas
            if not potential_cols:
                potential_cols = [c for c in self.columns if c]
                self.log("No se detectaron columnas obvias. Mostrando todas las columnas.", 'warning')

            self.calls_selected_columns = [] # Resetear selección previa para evitar stale data

            # Abrir selector
            self.root.after(100, lambda: self._open_calls_column_selector(potential_cols))

            self.log(f"✓ {len(self.raw_data)} filas cargadas. Abre ventana de selección.", 'success')
            self.total_messages = len(self.raw_data) # Estimado inicial
            self.update_stats()

        except Exception as e:
            self.log(f"Error al leer archivo: {e}", 'error')
            messagebox.showerror("Error", f"Error al leer archivo:\n{e}")

    # --- Lógica de Fidelizado (Carga Manual) ---

    def _open_calls_column_selector(self, columns):
        """Abre un popup para seleccionar las columnas de teléfono."""
        if not columns:
             self.log("No hay columnas para seleccionar.", 'warning')
             return

        window = ctk.CTkToplevel(self.root)
        window.title("Seleccionar Columnas")
        window.geometry("400x500")
        window.transient(self.root)
        window.grab_set()
        self._center_toplevel(window, 400, 500)

        # Header
        ctk.CTkLabel(window, text="Selecciona las columnas con números:", font=self.fonts['card_title']).pack(pady=(20, 10))

        # Scrollable area
        scroll = ctk.CTkScrollableFrame(window, fg_color="transparent")
        scroll.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        selected_vars = {}
        for col in columns:
            var = tk.BooleanVar(value=True) # Default all checked as requested
            chk = ctk.CTkCheckBox(scroll, text=col, variable=var, font=self.fonts['setting_label'])
            chk.pack(anchor="w", pady=5)
            selected_vars[col] = var

        # Buttons
        btn_frame = ctk.CTkFrame(window, fg_color="transparent")
        btn_frame.pack(fill=tk.X, padx=20, pady=20)

        ctk.CTkButton(btn_frame, text="Cancelar", command=window.destroy, fg_color=self.colors['action_cancel'], hover_color=self.hover_colors['action_cancel']).pack(side=tk.LEFT, expand=True, padx=5)

        ctk.CTkButton(btn_frame, text="Aceptar", command=lambda: self._on_calls_columns_confirmed(selected_vars, window), fg_color=self.colors['action_start'], hover_color=self.hover_colors['action_start']).pack(side=tk.LEFT, expand=True, padx=5)

    def _on_calls_columns_confirmed(self, selected_vars, window):
        """Procesa la selección de columnas del popup."""
        selected = [col for col, var in selected_vars.items() if var.get()]

        if not selected:
            messagebox.showwarning("Atención", "Debes seleccionar al menos una columna.", parent=window)
            return

        self.calls_selected_columns = selected

        # Calcular el total real de números (considerando splits)
        total_real = 0
        for row in self.raw_data:
            for col in selected:
                val = str(row.get(col, '')).strip()
                if val:
                    # Usar regex para dividir por guiones (todos los tipos), comas, punto y coma, saltos de línea
                    nums = [n.strip() for n in re.split(r'[-–—,\n;]', val) if n.strip()]
                    for n in nums:
                        clean_n = ''.join(filter(str.isdigit, n))
                        if clean_n:
                            total_real += 1

        self.total_messages = total_real
        self.update_stats()

        self.log(f"Columnas seleccionadas para llamadas: {', '.join(selected)}", 'success')
        self.log(f"Total de llamadas calculadas: {total_real}", 'info')
        window.destroy()

    def _load_default_messages(self):
        """Carga los mensajes predeterminados desde Grupos.txt si existe."""
        try:
            # Buscar Grupos.txt en el directorio del script
            grupos_path = os.path.join(BASE_DIR, "Grupos.txt")
            if os.path.exists(grupos_path):
                with open(grupos_path, 'r', encoding='utf-8') as f:
                    lines = [ln.strip() for ln in f.read().splitlines() if ln.strip()]
                if lines:
                    self.manual_messages_numbers = lines
                    self.manual_messages_groups = lines
                    # Generar índice de inicio aleatorio
                    self.mensaje_start_index = random.randint(0, len(lines) - 1)
                    self.log(f"Mensajes predeterminados cargados: {len(lines)} mensajes (inicio aleatorio en posición {self.mensaje_start_index + 1})", 'success')
                    return True
        except Exception as e:
            self.log(f"No se pudieron cargar mensajes predeterminados: {e}", 'warning')
        return False

    def _load_fidelizado_messages_from_file(self):
        """Abre un diálogo para cargar un archivo .txt de mensajes para el modo Fidelizado."""
        filepath = filedialog.askopenfilename(
            title="Seleccionar archivo de Mensajes (.txt)",
            filetypes=[("Text Files", "*.txt"), ("All files", "*.*")],
            parent=self.root
        )
        if not filepath:
            return

        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                lines = [line.strip() for line in f.readlines() if line.strip()]

            if not lines:
                messagebox.showwarning("Archivo Vacío", "El archivo seleccionado está vacío.", parent=self.root)
                return

            # Actualizar mensajes para ambos modos
            self.manual_messages_numbers = lines
            self.manual_messages_groups = lines

            # Actualizar la UI
            self.fidelizado_message_count_label.configure(text=f"✅ {len(lines)} mensajes cargados")
            self.log(f"Cargados {len(lines)} nuevos mensajes para Fidelizado desde {os.path.basename(filepath)}", 'success')
            messagebox.showinfo("Éxito", f"Se cargaron {len(lines)} mensajes.", parent=self.root)

        except Exception as e:
            self.log(f"Error al cargar archivo de mensajes: {e}", 'error')
            messagebox.showerror("Error de Lectura", f"No se pudo leer el archivo:\n{e}", parent=self.root)

    def handle_fidelizado_access(self):
        """Manejador del botón de Fidelizado (acceso directo)."""
        # Si la lógica de contraseña sigue siendo necesaria, se puede añadir aquí.
        # Por ahora, simplemente muestra la vista.
        self.show_fidelizado_view()

    def handle_sms_mode_access(self):
        """Manejador del acceso directo al SMS."""
        self.show_sms_view()

    def setup_fidelizado_view(self, parent):
        """Construye la interfaz de la vista Fidelizado."""
        # Cargar mensajes predeterminados si es la primera vez
        if not self.manual_messages_numbers and not self.manual_messages_groups:
            self._load_default_messages()

        # Contenedor principal de la vista Fidelizado
        fidelizado_container = ctk.CTkFrame(parent, fg_color="transparent")
        fidelizado_container.pack(fill=tk.BOTH, expand=True)

        # Contenido principal de Fidelizado
        content = ctk.CTkFrame(
            fidelizado_container,
            fg_color=self.colors['bg_card'],
            corner_radius=32,
            border_width=1,
            border_color=self._section_border_color()
        )
        self.fidelizado_main_card = content
        content.pack(fill=tk.X, expand=False, padx=10, pady=(10, 0))
        content.grid_columnconfigure(0, weight=1)

        # --- Header (Volver y Título) ---
        header_frame = ctk.CTkFrame(content, fg_color="transparent")
        header_frame.pack(fill=tk.X, padx=30, pady=(15, 10))

        self.back_to_traditional_btn = ctk.CTkButton(header_frame, text="←",
                                      width=40,
                                      command=self.show_traditional_view,
                                      fg_color=self.colors['bg'],
                                      text_color=self.colors['text'],
                                      font=('Inter', 16, 'bold'),
                                      corner_radius=20,
                                      hover_color=darken_color(self.colors['bg'], 0.1))
        self.back_to_traditional_btn.pack(side=tk.LEFT, padx=(0, 15))

        ctk.CTkLabel(header_frame, text="Fidelizado", font=('Inter', 26, 'bold'), text_color=self.colors['text']).pack(side=tk.LEFT)

        # --- STEP 1: DISPOSITIVOS ---
        step1_frame = ctk.CTkFrame(content, fg_color="transparent")
        step1_frame.pack(fill=tk.X, padx=20, pady=(10, 0))

        # Header Step 1
        s1_header = ctk.CTkFrame(step1_frame, fg_color="transparent")
        s1_header.pack(fill=tk.X, pady=(0, 5))
        self._create_step_badge(s1_header, 1).pack(side=tk.LEFT, padx=(0, 10))
        ctk.CTkLabel(s1_header, text="Dispositivos", font=('Inter', 16, 'bold'), text_color=self.colors['text']).pack(side=tk.LEFT)

        # Content Step 1
        device_card = ctk.CTkFrame(step1_frame, fg_color=self.colors['bg'], corner_radius=15)
        device_card.pack(fill=tk.X, pady=(5, 10))

        device_inner = ctk.CTkFrame(device_card, fg_color="transparent")
        device_inner.pack(fill=tk.X, padx=15, pady=15)

        self.fidelizado_detect_btn = ctk.CTkButton(device_inner, text="🔍 Detectar Dispositivos",
                                                  command=self.detect_devices,
                                                  font=self.fonts['button'],
                                                  fg_color=self.colors['action_detect'],
                                                  hover_color=self.hover_colors['action_detect'],
                                                  height=40)
        self.fidelizado_detect_btn.pack(fill='x', pady=(0, 10))

        self.fidelizado_device_list_label = ctk.CTkLabel(device_inner, text="Teléfonos - 0",
                                                        font=self.fonts['setting_label'],
                                                        text_color=self.colors['text_light'],
                                                        wraplength=350,
                                                        justify='left')
        self.fidelizado_device_list_label.pack(anchor='w')

        # --- STEP 2: CONFIGURACIÓN ---
        step2_frame = ctk.CTkFrame(content, fg_color="transparent")
        step2_frame.pack(fill=tk.X, padx=20, pady=(10, 0))

        s2_header = ctk.CTkFrame(step2_frame, fg_color="transparent")
        s2_header.pack(fill=tk.X, pady=(0, 5))
        self._create_step_badge(s2_header, 2).pack(side=tk.LEFT, padx=(0, 10))
        ctk.CTkLabel(s2_header, text="Configuración", font=('Inter', 16, 'bold'), text_color=self.colors['text']).pack(side=tk.LEFT)

        config_card = ctk.CTkFrame(step2_frame, fg_color=self.colors['bg'], corner_radius=15)
        config_card.pack(fill=tk.X, pady=(5, 10))

        config_grid = ctk.CTkFrame(config_card, fg_color="transparent")
        config_grid.pack(fill=tk.X, padx=15, pady=15)

        # Envíos (Mode)
        mode_container = ctk.CTkFrame(config_grid, fg_color="transparent")
        mode_container.pack(fill=tk.X, pady=(0, 10))
        ctk.CTkLabel(mode_container, text="Envíos:", font=self.fonts['button'], text_color=self.colors['text']).pack(side=tk.LEFT, padx=(0, 10))

        fidelizado_modes = ["Números", "Grupos"] # Plural
        # Internal mapping will be handled in logic
        current_mode_ui = "Números" # Default
        self.fidelizado_mode_var = tk.StringVar(value=current_mode_ui)

        mode_menu = ctk.CTkOptionMenu(
            mode_container,
            variable=self.fidelizado_mode_var,
            values=fidelizado_modes,
            font=self.fonts['button'],
            dropdown_font=self.fonts['setting_label'],
            fg_color=self.colors['bg_card'],
            button_color=self.colors['blue'],
            button_hover_color=darken_color(self.colors['blue'], 0.15),
            text_color=self.colors['text'],
            height=30,
            width=120 # Resize width
        )
        mode_menu.pack(side=tk.LEFT)

        # Conversación Mode (Uno a uno / Uno a muchos)
        self.numeros_mode_container = ctk.CTkFrame(config_grid, fg_color="transparent")
        self.numeros_mode_container.pack(fill=tk.X, pady=(0, 10))
        ctk.CTkLabel(self.numeros_mode_container, text="Conversación:", font=self.fonts['button'], text_color=self.colors['text']).pack(side=tk.LEFT, padx=(0, 10))
        ctk.CTkRadioButton(self.numeros_mode_container, text="Uno a uno", variable=self.fidelizado_numeros_mode, value="Uno a uno", font=self.fonts['setting_label'], text_color=self.colors['text']).pack(side=tk.LEFT, padx=(0, 10))
        ctk.CTkRadioButton(self.numeros_mode_container, text="Uno a muchos", variable=self.fidelizado_numeros_mode, value="Uno a muchos", font=self.fonts['setting_label'], text_color=self.colors['text']).pack(side=tk.LEFT)

        # WhatsApp Select
        wa_container = ctk.CTkFrame(config_grid, fg_color="transparent")
        wa_container.pack(fill=tk.X, pady=(0, 10))
        ctk.CTkLabel(wa_container, text="WhatsApp a usar:", font=self.fonts['button'], text_color=self.colors['text']).pack(side=tk.LEFT, padx=(0, 10))
        self.fidelizado_whatsapp_menu = ctk.CTkSegmentedButton(wa_container, variable=self.whatsapp_mode, values=["Normal", "Business", "Ambas", "Todas"], font=self.fonts['button_small'])
        self.fidelizado_whatsapp_menu.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # Bucles y Ciclos
        loops_row = ctk.CTkFrame(config_grid, fg_color="transparent")
        loops_row.pack(fill=tk.X, pady=(0, 10))

        self.loops_container = ctk.CTkFrame(loops_row, fg_color="transparent")
        self.loops_container.pack(side=tk.LEFT, fill=tk.X, expand=True)
        ctk.CTkLabel(self.loops_container, text="Bucle:", font=self.fonts['button'], text_color=self.colors['text']).pack(side=tk.LEFT, padx=(0, 10))
        self.manual_loops_var = tk.IntVar(value=max(1, self.manual_loops))
        self.manual_loops_var.trace_add('write', self.update_per_whatsapp_stat)
        self._create_spinbox_widget(self.loops_container, self.manual_loops_var, min_val=1, max_val=100).pack(side=tk.LEFT)

        self.cycles_container = ctk.CTkFrame(loops_row, fg_color="transparent")
        self.cycles_container.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(20, 0))
        ctk.CTkLabel(self.cycles_container, text="Ciclos:", font=self.fonts['button'], text_color=self.colors['text']).pack(side=tk.LEFT, padx=(0, 10))
        self.manual_cycles_var = SafeIntVar(value=1)
        self._create_spinbox_widget(self.cycles_container, self.manual_cycles_var, min_val=1, max_val=100).pack(side=tk.LEFT)

        # Delays
        delay_container = ctk.CTkFrame(config_grid, fg_color="transparent")
        delay_container.pack(fill=tk.X, pady=(0, 10))
        ctk.CTkLabel(delay_container, text="Retardo Envíos (s):", font=self.fonts['button'], text_color=self.colors['text']).pack(side=tk.LEFT, padx=(0, 10))
        sp_frame = ctk.CTkFrame(delay_container, fg_color="transparent")
        sp_frame.pack(side=tk.LEFT)
        self._create_spinbox_widget(sp_frame, self.fidelizado_send_delay_min, 1, 300).pack(side=tk.LEFT)
        ctk.CTkLabel(sp_frame, text="-", font=self.fonts['setting_label'], fg_color="transparent").pack(side=tk.LEFT, padx=5)
        self._create_spinbox_widget(sp_frame, self.fidelizado_send_delay_max, 1, 300).pack(side=tk.LEFT)

        # --- Controles de Variante Mixto (inicialmente ocultos) ---
        self.mixto_variant_container = ctk.CTkFrame(config_grid, fg_color="transparent")
        ctk.CTkLabel(self.mixto_variant_container, text="Variante Modo Mixto:", font=self.fonts['setting_label'], text_color=self.colors['text']).pack(anchor='w', pady=(0, 8))
        mixto_radio_frame = ctk.CTkFrame(self.mixto_variant_container, fg_color="transparent")
        mixto_radio_frame.pack(anchor='w')
        ctk.CTkRadioButton(mixto_radio_frame, text="1G:1N", variable=self.mixto_variant, value=1, font=self.fonts['setting_label'], text_color=self.colors['text']).pack(side=tk.LEFT, padx=(0, 15))
        ctk.CTkRadioButton(mixto_radio_frame, text="2G:1N", variable=self.mixto_variant, value=2, font=self.fonts['setting_label'], text_color=self.colors['text']).pack(side=tk.LEFT, padx=(0, 15))
        ctk.CTkRadioButton(mixto_radio_frame, text="3G:1N", variable=self.mixto_variant, value=3, font=self.fonts['setting_label'], text_color=self.colors['text']).pack(side=tk.LEFT)

        # --- Controles de Envío Simultáneo (Grupos) ---
        self.simultaneous_mode_container = ctk.CTkFrame(config_grid, fg_color="transparent")
        # Se mostrará condicionalmente en _update_fidelizado_ui_mode
        self.simultaneous_switch = ctk.CTkSwitch(
            self.simultaneous_mode_container,
            text="🚀 Envío Simultáneo (Por tandas)",
            variable=self.simultaneous_mode,
            font=self.fonts['button'],
            text_color=self.colors['text']
        )
        self.simultaneous_switch.pack(anchor='w')

        # --- STEP 3: CARGA ---
        step3_frame = ctk.CTkFrame(content, fg_color="transparent")
        step3_frame.pack(fill=tk.X, padx=20, pady=(10, 0))

        # Header Step 3
        s3_header = ctk.CTkFrame(step3_frame, fg_color="transparent")
        s3_header.pack(fill=tk.X, pady=(0, 5))
        self._create_step_badge(s3_header, 3).pack(side=tk.LEFT, padx=(0, 10))
        ctk.CTkLabel(s3_header, text="Carga", font=('Inter', 16, 'bold'), text_color=self.colors['text']).pack(side=tk.LEFT)

        # Content Step 3
        self.fidelizado_carga_section = ctk.CTkFrame(step3_frame, fg_color=self.colors['bg'], corner_radius=15)
        self.fidelizado_carga_section.pack(fill=tk.X, pady=(5, 10))

        # File Loader
        self.fidelizado_messages_container = ctk.CTkFrame(self.fidelizado_carga_section, fg_color="transparent")
        self.fidelizado_messages_container.pack(fill=tk.X, padx=15, pady=15)

        load_msg_frame = ctk.CTkFrame(self.fidelizado_messages_container, fg_color="transparent")
        load_msg_frame.pack(fill=tk.X, pady=(0, 10))

        load_messages_btn = ctk.CTkButton(load_msg_frame, text="Cargar Archivo Mensajes",
                                          command=self._load_fidelizado_messages_from_file,
                                          font=self.fonts['button'],
                                          fg_color=self.colors['blue'],
                                          hover_color=darken_color(self.colors['blue'], 0.15),
                                          height=35)
        load_messages_btn.pack(side=tk.LEFT)

        self.fidelizado_message_count_label = ctk.CTkLabel(load_msg_frame, text="", font=self.fonts['setting_label'], text_color=self.colors['text'])
        self.fidelizado_message_count_label.pack(side=tk.LEFT, padx=15)

        # Manual Inputs Frame (Holds Numbers and Groups boxes)
        self.fidelizado_manual_inputs_frame = ctk.CTkFrame(self.fidelizado_messages_container, fg_color="transparent")
        self.fidelizado_manual_inputs_frame.pack(fill=tk.BOTH, expand=True)

        # Numbers Box
        self.fidelizado_numbers_frame = ctk.CTkFrame(self.fidelizado_manual_inputs_frame, fg_color="transparent")
        self.fidelizado_numbers_frame.pack(fill=tk.X, pady=(0, 5))
        ctk.CTkLabel(self.fidelizado_numbers_frame, text="Números", font=('Inter', 14, 'bold'), text_color=self.colors['text']).pack(anchor='w')
        self.fidelizado_numbers_text = ctk.CTkTextbox(self.fidelizado_numbers_frame, font=('Inter', 13), corner_radius=10, border_width=1, border_color="#cccccc", wrap=tk.WORD, height=80)
        self.fidelizado_numbers_text.pack(fill=tk.X, pady=(5,0))
        self.fidelizado_numbers_text.bind("<<Modified>>", self._on_fidelizado_numbers_changed)

        # Groups Box
        self.fidelizado_groups_frame = ctk.CTkFrame(self.fidelizado_manual_inputs_frame, fg_color="transparent")
        self.fidelizado_groups_frame.pack(fill=tk.X, pady=(5, 0))
        ctk.CTkLabel(self.fidelizado_groups_frame, text="Links / Grupos", font=('Inter', 14, 'bold'), text_color=self.colors['text']).pack(anchor='w')
        self.fidelizado_groups_text = ctk.CTkTextbox(self.fidelizado_groups_frame, font=('Inter', 13), corner_radius=10, border_width=1, border_color="#cccccc", wrap=tk.WORD, height=80)
        self.fidelizado_groups_text.pack(fill=tk.X, pady=(5,0))

        initial_message_count = len(self.manual_messages_numbers)
        if initial_message_count > 0:
            self.fidelizado_message_count_label.configure(text=f"✅ {initial_message_count} mensajes cargados")
        else:
            self.fidelizado_message_count_label.configure(text="⚠️ No hay mensajes cargados")

        # --- STEP 4: ACCIONES ---
        step4_frame = ctk.CTkFrame(content, fg_color="transparent")
        step4_frame.pack(fill=tk.X, padx=20, pady=(10, 20))

        s4_header = ctk.CTkFrame(step4_frame, fg_color="transparent")
        s4_header.pack(fill=tk.X, pady=(0, 5))
        self._create_step_badge(s4_header, 4).pack(side=tk.LEFT, padx=(0, 10))
        ctk.CTkLabel(s4_header, text="Iniciar", font=('Inter', 16, 'bold'), text_color=self.colors['text']).pack(side=tk.LEFT)

        self.actions_frame = ctk.CTkFrame(step4_frame, fg_color="transparent")
        self.actions_frame.pack(fill=tk.X, pady=(5, 0))

        self.fidelizado_btn_start = ctk.CTkButton(self.actions_frame, text="▶ INICIAR ENVÍO FIDELIZADO", command=self.start_fidelizado_sending, fg_color=self.colors['action_start'], hover_color=self.hover_colors['action_start'], text_color=self.colors['text_header_buttons'], font=self.fonts['button'], corner_radius=10, height=50)
        self.fidelizado_btn_start.pack(fill=tk.X, pady=5)

        self.unirse_grupos_btn = ctk.CTkButton(self.actions_frame, text="🔗 UNIRSE A GRUPOS", command=self.start_unirse_grupos, fg_color=self.colors['action_detect'], hover_color=self.hover_colors['action_detect'], text_color=self.colors['text_header_buttons'], font=self.fonts['button'], corner_radius=10, height=50)
        self.unirse_grupos_btn.pack(fill=tk.X, pady=5)

        self.control_buttons_frame = ctk.CTkFrame(step4_frame, fg_color="transparent")
        self.fidelizado_btn_pause = ctk.CTkButton(self.control_buttons_frame, text="⏸  PAUSAR", command=self.pause_sending, fg_color=self.colors['action_pause'], hover_color=self.hover_colors['action_pause'], text_color=self.colors['text_header_buttons'], font=self.fonts['button'], corner_radius=10, height=45)
        self.fidelizado_btn_pause.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))

        self.fidelizado_btn_stop = ctk.CTkButton(self.control_buttons_frame, text="⏹  CANCELAR", command=self.stop_sending, fg_color=self.colors['action_cancel'], hover_color=self.hover_colors['action_cancel'], text_color=self.colors['text_header_buttons'], font=self.fonts['button'], corner_radius=10, height=45)
        self.fidelizado_btn_stop.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 0))

        self.control_buttons_frame.pack(fill=tk.X, pady=(5, 0))
        self.control_buttons_frame.pack_forget() # Initially hidden

        # --- Lógica de la Vista ---
        self.fidelizado_mode_var.trace_add('write', self._update_fidelizado_ui_mode)
        self.fidelizado_numeros_mode.trace_add('write', self._update_fidelizado_ui_mode) # NUEVO: trace para los radio buttons
        self._update_fidelizado_ui_mode()
        self._populate_fidelizado_inputs()

    def _on_fidelizado_numbers_changed(self, event=None):
        """Se activa cuando el contenido de la caja de texto de números cambia."""
        self.manual_inputs_numbers = [line.strip() for line in self.fidelizado_numbers_text.get("1.0", tk.END).splitlines() if line.strip()]
        self._update_fidelizado_ui_mode()
        # Marcar el widget para que el evento no se dispare recursivamente
        self.fidelizado_numbers_text.edit_modified(False)

    def _toggle_fidelizado_carga_section(self):
        """Alterna la visibilidad del bloque de cargas dentro del panel de configuración."""
        if not hasattr(self, 'fidelizado_carga_section') or not hasattr(self, 'fidelizado_carga_toggle_btn'):
            return

        if self.fidelizado_carga_section.winfo_manager():
            self.fidelizado_carga_section.pack_forget()
            self.fidelizado_carga_toggle_btn.configure(text="📥 Carga ▶")
        else:
            self.fidelizado_carga_section.pack(fill=tk.X, padx=20, pady=(0, 20))
            self.fidelizado_carga_toggle_btn.configure(text="📥 Carga ▼")

    def _update_fidelizado_ui_mode(self, *args):
        """Muestra u oculta los widgets y reorganiza el layout según el modo Fidelizado seleccionado."""
        mode_ui = self.fidelizado_mode_var.get()
        numeros_submode = self.fidelizado_numeros_mode.get()

        # Map UI values to internal keys. Note the plurals.
        mode_map_from_ui = {"Números": "NUMEROS", "Grupos": "GRUPOS", "Mixto": "MIXTO"}
        self.fidelizado_mode = mode_map_from_ui.get(mode_ui, "NUMEROS")
        self.set_activity_table_enabled(self.fidelizado_mode == "NUMEROS")

        show_mixto_variant = self.fidelizado_mode == "MIXTO"
        show_cycles = self.fidelizado_mode == "NUMEROS" and numeros_submode == "Uno a uno"

        # --- 1. Visibilidad de Widgets de Input (Carga) ---
        if self.fidelizado_mode == "NUMEROS":
            self.fidelizado_numbers_frame.pack(fill=tk.X, pady=(0, 5))
            self.fidelizado_groups_frame.pack_forget()

            if hasattr(self, 'numeros_mode_container'):
                self.numeros_mode_container.pack(fill=tk.X, pady=(0, 10))

        elif self.fidelizado_mode == "GRUPOS":
            self.fidelizado_numbers_frame.pack_forget()
            self.fidelizado_groups_frame.pack(fill=tk.X, pady=(5, 0))

            if hasattr(self, 'numeros_mode_container'):
                self.numeros_mode_container.pack_forget()

        elif self.fidelizado_mode == "MIXTO":
            self.fidelizado_numbers_frame.pack(fill=tk.X, pady=(0, 5))
            self.fidelizado_groups_frame.pack(fill=tk.X, pady=(5, 0))

            if hasattr(self, 'numeros_mode_container'):
                self.numeros_mode_container.pack_forget()

        # --- 2. Visibilidad de Configuración Adicional ---
        if show_mixto_variant:
            if not self.mixto_variant_container.winfo_manager():
                self.mixto_variant_container.pack(fill=tk.X, pady=(0, 10))
        else:
            self.mixto_variant_container.pack_forget()

        if show_cycles:
            if not self.cycles_container.winfo_manager():
                self.cycles_container.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(20, 0))
        else:
            self.cycles_container.pack_forget()

        # --- Configuración Simultáneo (Solo Grupos) ---
        if self.fidelizado_mode == "GRUPOS":
            if not self.simultaneous_mode_container.winfo_manager():
                self.simultaneous_mode_container.pack(fill=tk.X, pady=(0, 10))
        else:
            self.simultaneous_mode_container.pack_forget()

        # --- 3. Visibilidad de Botones de Acción ---
        if self.fidelizado_mode == "GRUPOS":
            if not self.unirse_grupos_btn.winfo_manager():
                self.unirse_grupos_btn.pack(fill=tk.X, pady=5)
            # El botón de simultáneo se eliminó de aquí, ahora es un switch en configuración
            self.fidelizado_btn_start.configure(text="▶ INICIAR ENVÍO A GRUPOS")
        else:
            self.unirse_grupos_btn.pack_forget()
            self.fidelizado_btn_start.configure(text="▶ INICIAR ENVÍO FIDELIZADO")

        # --- 4. Actualización del Menú de WhatsApp ---
        if self.fidelizado_mode == "NUMEROS":
            self.fidelizado_whatsapp_menu.configure(values=["Normal", "Business", "Ambas"])
            if self.whatsapp_mode.get() == "Todas":
                self.whatsapp_mode.set("Ambas")
        else:
            if hasattr(self, 'fidelizado_whatsapp_menu'):
                self.fidelizado_whatsapp_menu.configure(values=["Normal", "Business", "Ambas", "Todas"])

        self.update_per_whatsapp_stat()

    def _populate_fidelizado_inputs(self):
        """Limpia y rellena los campos de texto con los datos guardados en las variables."""
        # Limpiar contenido existente
        self.fidelizado_groups_text.delete("1.0", tk.END)
        if hasattr(self, 'fidelizado_numbers_text'):
            self.fidelizado_numbers_text.delete("1.0", tk.END)

        # Rellenar con datos guardados
        if self.manual_inputs_groups:
            self.fidelizado_groups_text.insert("1.0", "\n".join(self.manual_inputs_groups))
        if self.manual_inputs_numbers:
             if hasattr(self, 'fidelizado_numbers_text'):
                self.fidelizado_numbers_text.insert("1.0", "\n".join(self.manual_inputs_numbers))

        # Si no hay mensajes de grupo pero sí de número (caso común), usarlos también para grupos
        if self.manual_messages_numbers and not self.manual_messages_groups:
             self.manual_messages_groups = self.manual_messages_numbers

        # --- NUEVO: Actualizar el contador de mensajes ---
        message_count = len(self.manual_messages_numbers)
        if hasattr(self, 'fidelizado_message_count_label'): # Asegurarse de que el widget exista
            if message_count > 0:
                self.fidelizado_message_count_label.configure(text=f"✅ {message_count} mensajes cargados")
            else:
                self.fidelizado_message_count_label.configure(text="⚠️ No hay mensajes cargados")

    def start_fidelizado_sending(self):
        """Envoltorio para iniciar Fidelizado con opción de programar."""
        # El flag self.simultaneous_mode se gestiona ahora via self.simultaneous_mode (BooleanVar) desde la UI
        # No hace falta setearlo aquí manualmente
        self._open_schedule_dialog(lambda confirm: self._start_fidelizado_sending_impl(confirm=confirm))

    def _start_fidelizado_sending_impl(self, confirm=True):
        """Función específica para validar y preparar el envío desde la vista Fidelizado (Lógica Interna)."""
        # 1. Guardar los datos de los TextBoxes en las variables de la clase
        self.manual_inputs_groups = [line.strip() for line in self.fidelizado_groups_text.get("1.0", tk.END).splitlines() if line.strip()]
        if hasattr(self, 'fidelizado_numbers_text'):
            self.manual_inputs_numbers = [line.strip() for line in self.fidelizado_numbers_text.get("1.0", tk.END).splitlines() if line.strip()]

        # 2. Validar los datos según el modo
        if self.fidelizado_mode == "NUMEROS":
            if not self.manual_messages_numbers:
                messagebox.showerror("Error", "Modo Números requiere mensajes cargados.", parent=self.root)
                return
            # Si hay números en la caja de texto, se usa el modo manual.
            if self.manual_inputs_numbers:
                self.manual_mode = True
                self.fidelizado_mode = "NUMEROS_MANUAL" # Un modo interno para distinguirlo
                self.links = []
                self.link_retry_map = {}
                self._start_sending_impl(confirm=confirm)
                return
            else:
                # Si no, se usa el modo automático.
                if self.detected_phone_lines:
                    self.log("Usando líneas detectadas previamente para iniciar el envío.", 'info')
                    if not self._prepare_fidelizado_numbers_data():
                        return
                    self.root.after(0, self._start_numeros_mode_after_review)
                else:
                    # Pasar confirmación (invertida para auto_confirm)
                    auto_confirm = not confirm
                    threading.Thread(target=self.start_fidelizado_numeros_thread, args=(auto_confirm,), daemon=True).start()
                return

        if self.fidelizado_mode == "GRUPOS" and not self.manual_inputs_groups:
            messagebox.showerror("Error", "El 'Modo Grupos' requiere al menos un link de grupo.", parent=self.root)
            return
        if self.fidelizado_mode == "MIXTO" and (not self.manual_inputs_groups):
            messagebox.showerror("Error", "El 'Modo Mixto' requiere grupos.", parent=self.root)
            return
        if not self.manual_messages_numbers:
            messagebox.showerror("Error", "Se requiere al menos un mensaje.", parent=self.root)
            return

        # 3. Marcar el modo manual y llamar a la función de envío principal (SOLO para Grupos y Mixto)
        self.manual_mode = True
        self.group_mode = self.fidelizado_mode == "GRUPOS" # Flag legacy
        self.links = [] # Limpiar links del modo tradicional
        self.link_retry_map = {}

        self._start_sending_impl(confirm=confirm) # Llamar a la lógica de envío compartida

    def start_fidelizado_numeros_thread(self, auto_confirm=False):
        """
        Hilo de trabajo para preparar e iniciar el envío del modo NÚMEROS automático.
        Detecta los números, calcula el total y espera la confirmación desde el editor.
        """
        if not self._detect_and_prepare_phone_lines(auto_confirm=auto_confirm):
            self.root.after(0, lambda: messagebox.showerror("Error", "No se detectaron suficientes líneas de WhatsApp para iniciar (se requieren al menos 2).", parent=self.root))
            return

        if not self.numbers_editor_start_requested:
            self.log("Inicio de envío cancelado tras revisar los números detectados.", 'info')
            return

        if not self._prepare_fidelizado_numbers_data(from_thread=True):
            return

        self.root.after(0, self._start_numeros_mode_after_review)

    def _prepare_fidelizado_numbers_data(self, *, from_thread=False):
        """Filtra y calcula los datos necesarios para iniciar el modo números."""
        if not self.detected_phone_lines:
            msg = "No hay líneas detectadas para iniciar el modo Números."
            self.log(msg, 'error')
            self.numbers_editor_start_requested = False
            handler = (lambda: messagebox.showerror("Error", msg, parent=self.root))
            if from_thread:
                self.root.after(0, handler)
            else:
                handler()
            return False

        whatsapp_mode = self.whatsapp_mode.get()
        original_line_count = len(self.detected_phone_lines)

        if whatsapp_mode == "Normal":
            filtered_lines = [line for line in self.detected_phone_lines if line['type'] == 'WhatsApp']
            self.log(f"Filtrando por 'Normal'. Líneas a usar: {len(filtered_lines)} de {original_line_count}", 'info')
        elif whatsapp_mode == "Business":
            filtered_lines = [line for line in self.detected_phone_lines if line['type'] == 'WhatsApp Business']
            self.log(f"Filtrando por 'Business'. Líneas a usar: {len(filtered_lines)} de {original_line_count}", 'info')
        else:
            filtered_lines = list(self.detected_phone_lines)
            modo = "Ambas" if whatsapp_mode == "Ambas" else ("Todas" if whatsapp_mode == "Todas" else whatsapp_mode)
            self.log(f"Modo '{modo}'. Usando {len(filtered_lines)} líneas detectadas.", 'info')

        if len(filtered_lines) < 2 and self.fidelizado_numeros_mode.get() != "Uno a muchos":
            msg = (
                f"Después de filtrar por '{whatsapp_mode}', solo quedan {len(filtered_lines)} "
                "líneas. Se requieren al menos 2 para este modo."
            )
            self.log(msg, 'error')
            self.numbers_editor_start_requested = False
            handler = (lambda: messagebox.showerror("Error", msg, parent=self.root))
            if from_thread:
                self.root.after(0, handler)
            else:
                handler()
            return False

        self.detected_phone_lines = filtered_lines

        num_lines = len(self.detected_phone_lines)
        num_bucles = self.manual_loops_var.get()
        num_ciclos = self.manual_cycles_var.get()

        if self.fidelizado_numeros_mode.get() == "Uno a uno":
            total = (num_lines // 2) * 2 * num_ciclos * num_bucles
        else:
            num_pairs = num_lines * (num_lines - 1) // 2
            total = num_pairs * 2 * num_bucles

        self.total_messages = total
        self.log(f"Cálculo para Modo '{self.fidelizado_numeros_mode.get()}': {self.total_messages} mensajes en total.", 'info')
        return True

    def _start_numeros_mode_after_review(self):
        """Inicia el envío del modo números después de revisar y confirmar desde el editor."""
        self.numbers_editor_start_requested = False
        self.reset_fidelizado_activity_records()
        self.show_activity_log()

        # Iniciar el envío
        self.manual_mode = True
        self.fidelizado_mode = "NUMEROS"
        self.set_activity_table_enabled(True)
        self.links = []
        self.link_retry_map = {}

        self._enter_task_mode()
        self.update_stats()

        threading.Thread(target=self.send_thread, daemon=True).start()

    def start_unirse_grupos(self):
        """Envoltorio para unirse a grupos con opción de programar."""
        self._open_schedule_dialog(lambda confirm: self._start_unirse_grupos_impl(confirm=confirm))

    def _start_unirse_grupos_impl(self, confirm=True):
        """Valida e inicia el hilo para unirse a grupos (Lógica Interna)."""
        if not self.devices:
            messagebox.showerror("Error", "Paso 1: Detecta al menos un dispositivo.", parent=self.root)
            return

        grupos = [line.strip() for line in self.fidelizado_groups_text.get("1.0", tk.END).splitlines() if line.strip()]
        if not grupos:
            messagebox.showerror("Error", "Ingresa al menos un link de grupo en la caja de texto.", parent=self.root)
            return

        if confirm:
            if not messagebox.askyesno("Confirmar Acción", f"¿Estás seguro de que deseas intentar unirte a {len(grupos)} grupo(s) en {len(self.devices)} dispositivo(s)?", parent=self.root):
                return

        # Iniciar el proceso en un hilo para no bloquear la UI
        threading.Thread(target=self.run_unirse_grupos, args=(grupos,), daemon=True).start()

    def validate_numbers(self, inputs_raw, parent_window):
        """Valida una lista de números. Devuelve lista limpia o None si hay error."""
        inputs_clean_nums = []
        for raw in inputs_raw:
            s = raw.strip()
            norm = ''.join(s.split())
            if not s: continue
            if norm.startswith('+549'): 
                messagebox.showerror("Error", "No incluyas el prefijo '+549' en los números.", parent=parent_window); return None
            if norm.startswith('+'): norm = norm[1:]
            if not norm.isdigit(): 
                messagebox.showerror("Error", f"Número inválido encontrado: {s}", parent=parent_window); return None
            inputs_clean_nums.append(norm)
        return inputs_clean_nums

    def validate_groups(self, inputs_raw, parent_window):
        """Valida una lista de links de grupo. Devuelve lista limpia o None si hay error."""
        inputs_clean_groups = []
        for raw in inputs_raw:
            s = raw.strip()
            if not s: continue
            if not (s.startswith("https://chat.whatsapp.com/") or s.startswith("http://chat.whatsapp.com/")):
                messagebox.showerror("Error", f"Link de grupo inválido encontrado:\n{s}", parent=parent_window); return None
            inputs_clean_groups.append(s)
        return inputs_clean_groups

    def generate_manual_links(self, numbers, messages, loops):
        """Genera la lista de enlaces para el modo Fidelizado (Números)."""
        if not numbers or not messages:
            return []
        
        # Lógica de "loops" para Modo Números:
        # Repite la *lista de mensajes* 'loops' veces.
        # Asigna números rotativamente a esta lista extendida de mensajes.
        
        total_messages_to_send = len(messages) * loops
        final_links = []
        
        for i in range(total_messages_to_send):
            msg = messages[i % len(messages)]
            num = numbers[i % len(numbers)] # Rota los números
            link = f"https://wa.me/549{num}?text={urllib.parse.quote(msg, safe='')}"
            final_links.append(link)
            
        return final_links

    def generate_manual_pairs(self, links_or_nums, messages, loops):
        """Genera pares (link_o_numero, mensaje) para el modo Fidelizado (Grupos)."""
        if not links_or_nums or not messages:
            return []
        
        # Lógica de "loops" para Modo Grupos:
        # Repite la *lista de mensajes* 'loops' veces.
        # Asigna grupos rotativamente a esta lista extendida de mensajes.
        
        total_messages_to_send = len(messages) * loops
        final_pairs = []

        for i in range(total_messages_to_send):
            msg = messages[i % len(messages)]
            target = links_or_nums[i % len(links_or_nums)] # Rota los grupos/números
            final_pairs.append((target, msg))
            
        return final_pairs

    # --- Lógica del Procesador de Excel ---

    def _calculate_sms_stats(self, text):
        """Calcula estadísticas de SMS según reglas GSM/Unicode definidas por el usuario."""
        if not text:
            return "0 chars | GSM | 0 SMS"

        # Detección de codificación
        # Si hay algún caracter que NO esté en GSM_BASIC_CHARS ni en GSM_EXTENDED_CHARS, es Unicode.
        # (Esto incluye ñ, acentos, emojis, etc.)
        use_unicode = False
        for char in text:
            if char not in GSM_BASIC_CHARS and char not in GSM_EXTENDED_CHARS:
                use_unicode = True
                break

        count = 0
        if use_unicode:
            mode = "Unicode"
            # Unicode: límite 70 (1 segmento) o 67 (multi)
            count = len(text) # Python len es correcto para chars unicode
            limit_single = 70
            limit_multi = 67
        else:
            mode = "GSM"
            # GSM: límite 160 (1 segmento) o 153 (multi)
            # Caracteres extendidos cuentan doble
            for char in text:
                if char in GSM_EXTENDED_CHARS:
                    count += 2
                else:
                    count += 1
            limit_single = 160
            limit_multi = 153

        if count <= limit_single:
            sms_count = 1
        else:
            sms_count = math.ceil(count / limit_multi)

        return f"{count} chars | {mode} | {sms_count} SMS"

    def open_processor_window(self, original_file):
        """Abre la ventana para configurar la plantilla de mensajes."""
        proc_window = ctk.CTkToplevel(self.root)
        proc_window.title("Procesar Excel")
        proc_window.transient(self.root)

        width, height = 900, 750
        # Centrar en la pantalla
        screen_width = proc_window.winfo_screenwidth()
        screen_height = proc_window.winfo_screenheight()
        x = (screen_width // 2) - (width // 2)
        y = (screen_height // 2) - (height // 2)
        proc_window.geometry(f"{width}x{height}+{x}+{y}"); proc_window.after(100, proc_window.focus_force)

        main_cont = ctk.CTkFrame(proc_window, fg_color=self.colors['bg'], corner_radius=0)
        main_cont.pack(fill=tk.BOTH, expand=True)

        # Header (Minimalista)
        header_frame = ctk.CTkFrame(main_cont, fg_color="transparent")
        header_frame.pack(fill=tk.X, padx=30, pady=(25, 10))

        ctk.CTkLabel(header_frame, text="Procesar Excel", font=('Inter', 26, 'bold'), text_color=self.colors['text']).pack(side=tk.LEFT)

        # Contenido scrollable
        scroll_f = ctk.CTkScrollableFrame(main_cont, fg_color="transparent", corner_radius=0)
        scroll_f.pack(fill=tk.BOTH, expand=True, padx=20)

        # Tarjeta 1: Info Archivo (Simplificada)
        step1_card = ctk.CTkFrame(scroll_f, fg_color=self.colors['bg_card'], corner_radius=15)
        step1_card.pack(fill=tk.X, padx=10, pady=(15, 15))

        file_name_short = os.path.basename(original_file)
        ctk.CTkLabel(step1_card, text=f"Información del Archivo ({file_name_short})", font=self.fonts['card_title'], text_color=self.colors['text']).pack(anchor='w', padx=20, pady=20)

        # Crear acordeón
        # Nota: Usamos verde para todos los pasos para mejor visibilidad en dark mode
        green_color = self.colors['green']
        steps_data = [
            {"title": "Columnas de Teléfono", "color": green_color, "id": "step2"},
            {"title": "Columnas para Mensaje", "color": green_color, "id": "step3"},
            {"title": "Plantilla de Mensaje", "color": green_color, "id": "step4"}
        ]
        toggles = {}

        for i, data in enumerate(steps_data):
            step_card = ctk.CTkFrame(scroll_f, fg_color=self.colors['bg_card'], corner_radius=15)
            step_card.pack(fill=tk.X, padx=10, pady=15)

            tb = ctk.CTkFrame(step_card, fg_color="transparent", cursor='hand2')
            tb.pack(fill=tk.X, padx=20, pady=20)

            hi = ctk.CTkFrame(tb, fg_color="transparent")
            hi.pack(fill=tk.X)

            # Números de paso en verde
            ctk.CTkLabel(hi, text=str(i+1), font=('Inter', 18, 'bold'), fg_color="transparent", text_color=data["color"]).pack(side=tk.LEFT, padx=(0, 12))
            ctk.CTkLabel(hi, text=data["title"], font=self.fonts['card_title'], text_color=self.colors['text']).pack(side=tk.LEFT)
            # Flecha en verde
            al = ctk.CTkLabel(hi, text="▼", font=('Inter', 16, 'bold'), text_color=data["color"])
            al.pack(side=tk.RIGHT, padx=10)

            cf = ctk.CTkFrame(step_card, fg_color="transparent")
            cf.pack_forget() # Oculto por defecto

            toggles[data["id"]] = {"bar": tb, "header": hi, "arrow": al, "content": cf}

            # Función de toggle
            def create_tf(content_frame, arrow_label):
                def toggle_func(event=None):
                    if content_frame.winfo_ismapped():
                        content_frame.pack_forget()
                        arrow_label.configure(text="▼")
                    else:
                        content_frame.pack(fill=tk.X, pady=(0, 20), padx=20)
                        arrow_label.configure(text="▲")
                return toggle_func

            tf = create_tf(cf, al)
            tb.bind('<Button-1>', tf)
            for w in hi.winfo_children():
                w.bind('<Button-1>', tf)

        # Rellenar Step 2 (Teléfonos)
        step2_c = toggles["step2"]["content"]
        ctk.CTkLabel(step2_c, text="Selecciona qué columnas contienen números de teléfono:", font=self.fonts['setting_label'], text_color=self.colors['text_light']).pack(anchor='w', pady=(0, 10))
        self.phone_vars = {}
        pb_frame = ctk.CTkFrame(step2_c, fg_color="transparent")
        pb_frame.pack(fill=tk.X)
        for i, pc in enumerate(self.phone_columns):
            var = tk.BooleanVar(value=(i==0)) # Marcar solo la primera por defecto
            self.phone_vars[pc] = var
            ctk.CTkCheckBox(pb_frame, text=pc, variable=var, font=self.fonts['setting_label'], text_color=self.colors['text'], border_color=self.colors['text_light'], hover_color=self.colors['bg'], fg_color=self.colors['blue']).pack(anchor='w', pady=4)

        # Rellenar Step 3 (Columnas Mensaje)
        step3_c = toggles["step3"]["content"]
        ctk.CTkLabel(step3_c, text="Selecciona las columnas que usarás en el mensaje:", font=self.fonts['setting_label'], text_color=self.colors['text_light']).pack(anchor='w', pady=(0, 10))
        self.column_vars = {}
        cg = ctk.CTkFrame(step3_c, fg_color="transparent")
        cg.pack(fill=tk.X)
        col_c, row_c = 0, 0
        for col in self.columns:
            if col and col not in self.phone_columns:
                var = tk.BooleanVar(value=False)
                self.column_vars[col] = var
                ctk.CTkCheckBox(cg, text=col, variable=var, font=self.fonts['setting_label'], text_color=self.colors['text'], border_color=self.colors['text_light'], hover_color=self.colors['bg'], fg_color=self.colors['blue']).grid(row=row_c, column=col_c, sticky='w', padx=10, pady=4)
                col_c += 1
                if col_c >= 3: # 3 columnas de checkboxes
                    col_c = 0
                    row_c += 1

        # Rellenar Step 4 (Plantilla)
        step4_c = toggles["step4"]["content"]
        ctk.CTkLabel(step4_c, text="Insertar columna:", font=self.fonts['setting_label'], text_color=self.colors['text_light']).pack(anchor='w')
        bc = ctk.CTkFrame(step4_c, fg_color="transparent") # Contenedor de botones de columna
        bc.pack(fill=tk.X, pady=(5, 10))

        ctk.CTkLabel(step4_c, text="Plantilla de Mensaje (usa {Columna}):", font=self.fonts['setting_label'], text_color=self.colors['text_light']).pack(anchor='w', pady=(10, 5))
        mt = ctk.CTkTextbox(step4_c, height=100, font=self.fonts['setting_label'], corner_radius=10, border_width=1, border_color="#cccccc", wrap=tk.WORD)
        mt.pack(fill=tk.BOTH, expand=True)

        # Previsualización Mejorada
        preview_container = ctk.CTkFrame(step4_c, fg_color="transparent")
        preview_container.pack(fill=tk.X, pady=(15, 0))

        # Estado de visibilidad de la preview
        self.preview_visible = False

        # Frame de la preview (oculto inicialmente)
        pf = ctk.CTkFrame(step4_c, fg_color=self.colors['bg'], corner_radius=10, border_width=1, border_color=self.colors["text_light"])

        # Botón Ojo
        def toggle_preview():
            if self.preview_visible:
                pf.pack_forget()
                self.preview_visible = False
                # btn_preview.configure(text="👁️ Mostrar Vista Previa") # Opcional: cambiar texto
            else:
                pf.pack(fill=tk.BOTH, expand=True, pady=(10, 0))
                self.preview_visible = True
                update_preview()
                # btn_preview.configure(text="👁️ Ocultar Vista Previa")

        # Frame contenedor para el botón y el contador
        preview_controls_row = ctk.CTkFrame(preview_container, fg_color="transparent")
        preview_controls_row.pack(anchor='w', fill=tk.X)

        btn_preview = ctk.CTkButton(preview_controls_row, text="👁️",
                                    command=toggle_preview,
                                    fg_color="transparent",
                                    text_color=self.colors['text'],
                                    hover_color=self.colors['bg_card'],
                                    font=('Inter', 24),
                                    width=50, height=40)
        btn_preview.pack(side=tk.LEFT)

        # Etiqueta de estadísticas SMS (oculta inicialmente o vacía)
        self.lbl_sms_stats = ctk.CTkLabel(preview_controls_row, text="", font=('Inter', 12, 'bold'), text_color=self.colors['text'])
        self.lbl_sms_stats.pack(side=tk.LEFT, padx=(15, 0))

        ctk.CTkLabel(preview_container, text="Mostrar Vista Previa", font=('Inter', 12), text_color=self.colors['text_light']).pack(anchor='w', padx=10)

        # Configuración del Textbox de Preview (Más grande y legible)
        ctk.CTkLabel(pf, text="Vista Previa (Fila 1):", font=('Inter', 12, 'bold'), text_color=self.colors['text_light']).pack(anchor='w', padx=15, pady=(10, 5))
        pt = ctk.CTkTextbox(pf, height=150, font=('Inter', 13), fg_color=self.colors['bg_card'], text_color=self.colors['text'], corner_radius=8, wrap=tk.WORD, border_width=0)
        pt.pack(fill=tk.BOTH, expand=True, padx=15, pady=(0, 15))
        pt.configure(state=tk.DISABLED)

        def update_preview(*a):
            try:
                cm = mt.get('1.0', tk.END).strip()

                # Calcular preview (incluso si no es visible, para el contador)
                pm = cm
                if cm and self.raw_data:
                    er = self.raw_data[0] # Primera fila de datos
                    for c in self.columns:
                        pl = f"{{{c}}}"
                        if pl in pm:
                             v = er.get(c, '')
                             v = '' if v is None else str(v)
                             # Formato especial para valores monetarios
                             if '$ Hist.' in c or '$ Asig.' in c:
                                 v = format_currency_value(v)
                             pm = pm.replace(pl, v)

                # Actualizar estadísticas siempre (si hay texto procesado o plantilla)
                # Si hay raw_data, pm es el procesado. Si no, pm es la plantilla (cm).
                # Pero si no hay raw_data, preferimos mostrar stats de la plantilla tal cual.
                stats_text = self._calculate_sms_stats(pm)
                self.lbl_sms_stats.configure(text=stats_text)

                # Solo actualizar el widget de texto si es visible
                if self.preview_visible:
                    pt.configure(state=tk.NORMAL)
                    pt.delete('1.0', tk.END)
                    if not cm:
                        pt.insert('1.0', '(Escribe un mensaje para previsualizar)')
                    elif self.raw_data:
                        pt.insert('1.0', pm)
                    else:
                        pt.insert('1.0', '(No hay datos para previsualizar)')
                    pt.configure(state=tk.DISABLED)

            except Exception:
                pass # Evitar errores durante la escritura

        mt.bind('<KeyRelease>', update_preview)
        mt.bind('<ButtonRelease>', update_preview)
        # update_preview() # No llamar al inicio ya que está oculto

        def update_buttons(*a):
            """Actualiza los botones de inserción rápida."""
            [w.destroy() for w in bc.winfo_children()] # Limpiar botones anteriores
            sel = [c for c, v in self.column_vars.items() if v.get()]
            if not sel:
                ctk.CTkLabel(bc, text="(Selecciona columnas en el Paso 3 para insertarlas)", font=('Inter',10,'italic'), text_color=self.colors['text_light']).pack(anchor='w')
                return

            def ins(fn):
                mt.insert(tk.INSERT, f"{{{fn}}}")
                mt.focus()
                update_preview()

            col, row = 0, 0
            for c in sel:
                ctk.CTkButton(bc, text=c, command=lambda x=c: ins(x),
                              fg_color=self.colors['blue'], hover_color=darken_color(self.colors['blue'],0.18),
                              text_color=self.colors['text_header'], font=('Inter',9,'bold'),
                              height=30, corner_radius=10).grid(row=row, column=col, padx=3, pady=3, sticky='ew')
                col += 1
                if col >= 4: # 4 botones por fila
                    col = 0
                    row += 1

        [v.trace('w', update_buttons) for v in self.column_vars.values()]
        update_buttons()

        # Barra de botones inferior
        button_bar = ctk.CTkFrame(main_cont, fg_color="transparent", corner_radius=0, border_width=0)
        button_bar.pack(fill=tk.X, side=tk.BOTTOM, pady=0)
        btn_inner_frame = ctk.CTkFrame(button_bar, fg_color="transparent")
        btn_inner_frame.pack(fill=tk.X, padx=30, pady=20)

        def process_config():
            sp = [c for c, v in self.phone_vars.items() if v.get()]
            if not sp:
                messagebox.showwarning("Aviso", "Selecciona al menos una columna de Teléfono (Paso 2)", parent=proc_window)
                return

            sc = [c for c, v in self.column_vars.items() if v.get()]
            mtmpl = mt.get("1.0", tk.END).strip()
            if not mtmpl:
                messagebox.showwarning("Aviso", "Escribe una plantilla de Mensaje (Paso 4)", parent=proc_window)
                return

            self.log("Procesando...", 'info')
            # Cerrar ventana primero
            proc_window.destroy()
            self.root.focus_force()
            # Luego procesar datos y mostrar mensaje
            self.process_excel_data(sc, mtmpl, sp)

        def cancel_config():
            proc_window.destroy()
            self.root.focus_force()

        proc_window.protocol("WM_DELETE_WINDOW", cancel_config)
        ctk.CTkButton(btn_inner_frame, text="Cancelar", command=cancel_config, fg_color=self.colors['action_cancel'], hover_color=self.hover_colors['action_cancel'], font=self.fonts['button'], corner_radius=10, height=40).pack(side=tk.RIGHT, padx=(10, 0))
        ctk.CTkButton(btn_inner_frame, text="Procesar y Generar", command=process_config, fg_color=self.colors['action_start'], hover_color=self.hover_colors['action_start'], font=self.fonts['button'], corner_radius=10, height=40).pack(side=tk.RIGHT)

        # Abrir todos los acordeones por defecto
        self.root.update_idletasks()
        for i in range(2, 5):
            toggles[f"step{i}"]["bar"].event_generate("<Button-1>")

    def process_excel_data(self, selected_columns, message_template, selected_phones):
        """Genera la lista de URLs de WhatsApp a partir de los datos y la plantilla."""
        processed_rows = []
        retry_map = {}
        for row in self.raw_data:
            # Obtener todos los números de las columnas de teléfono seleccionadas
            phone_nums = []
            for ph_col in selected_phones:
                ph_val = str(row.get(ph_col, '')) if row.get(ph_col) else ''
                # Soportar números separados por guión (ej. "111-222")
                phone_nums.extend([n.strip() for n in ph_val.split('-') if n.strip()])

            if not phone_nums:
                continue # Sin número en esta fila

            links_for_row = []
            msg = message_template
            # Rellenar plantilla una sola vez por fila
            for col in selected_columns:
                pl = f"{{{col}}}"
                if pl in msg:
                    val = row.get(col, '')
                    val = '' if val is None else str(val)
                    # Formato especial para valores monetarios
                    if '$ Hist.' in col or '$ Asig.' in col:
                        val = format_currency_value(val)
                    msg = msg.replace(pl, val)

            for phone in phone_nums:
                if not phone or not phone.strip():
                    continue

                ph_digits = re.sub(r'\D', '', phone)
                if not ph_digits:
                    continue

                enc_msg = urllib.parse.quote(msg, safe='')
                if self.sms_mode_active:
                    link = f"sms:{ph_digits}?body={enc_msg}"
                else:
                    link = f"https://wa.me/549{ph_digits}?text={enc_msg}"
                links_for_row.append(link)

            if not links_for_row:
                continue

            # Agregar todos los números encontrados como tareas independientes
            processed_rows.extend(links_for_row)

        self.links = processed_rows
        self.link_retry_map = {} # Ya no se usan reintentos para múltiples números, son tareas individuales
        self.total_messages = len(self.links)
        self.update_stats()
        self.log(f"{len(self.links)} URLs generados", 'success')
        self.update_per_whatsapp_stat()

        if not self.manual_mode:
            self.save_processed_excel() # Ofrecer guardar solo si no es modo Fidelizado

    def save_processed_excel(self):
        """Ofrece guardar un nuevo archivo Excel solo con las URLs generadas."""
        try:
            self.root.attributes('-topmost', True) # Asegurar que el diálogo esté al frente
            out_path = filedialog.asksaveasfilename(
                parent=self.root,
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
                title="Guardar Excel Procesado con URLs"
            )
            self.root.attributes('-topmost', False); self.root.focus_force() # Devolver foco

            if out_path:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "URLs"
                ws['A1'] = 'URL'
                for i, url in enumerate(self.links, 2):
                    ws[f'A{i}'] = url
                wb.save(out_path)
                self.log(f"Excel guardado: {os.path.basename(out_path)}", 'success')
                messagebox.showinfo("Éxito", f"Archivo Excel guardado con éxito.\nSe generaron {len(self.links)} URLs listos para enviar.", parent=self.root)
        except Exception as e:
            self.log(f"Error al guardar Excel: {e}", 'error')
            messagebox.showerror("Error", f"Error al guardar el archivo:\n{e}", parent=self.root)
            self.root.attributes('-topmost', False); self.root.focus_force()

    def _get_phone_from_link(self, link):
        """Extrae un número de teléfono de un link wa.me o sms:."""
        if not isinstance(link, str):
            return None
        try:
            clean_link = link.strip()
            if "wa.me/" in clean_link:
                return clean_link.split("wa.me/")[1].split("?")[0]
            elif clean_link.lower().startswith("sms:"):
                return clean_link.split(":")[1].split("?")[0]
        except (IndexError, AttributeError):
            pass
        return None

    # --- Lógica de Envío (Threading y ADB) ---

    # --- INICIO MODIFICACIÓN: start_sending (Modo Bucles Blast V2) ---
    def start_sending(self):
        """Envoltorio para iniciar envío con opción de programar."""
        self._open_schedule_dialog(lambda confirm: self._start_sending_impl(confirm=confirm))

    def _start_sending_impl(self, confirm=True):
        """Valida e inicia el hilo de envío de mensajes (Lógica Interna)."""
        if not self.adb_path.get() or not os.path.exists(self.adb_path.get()):
            messagebox.showerror("Error", "ADB no encontrado.\nVe a la carpeta 'scrcpy' o ejecuta INSTALAR.bat.", parent=self.root); return
        if not self.devices:
            messagebox.showerror("Error", "Paso 1: Detecta al menos un dispositivo.", parent=self.root); return
        
        # --- Validación de Tareas ---
        # Modos que NO requieren self.links (se procesan directamente en el hilo)
        modos_sin_links = ["NUMEROS", "GRUPOS", "MIXTO"]
        
        if not self.links and self.fidelizado_mode not in modos_sin_links:
            messagebox.showerror("Error", "Paso 2 o Fidelizado: Carga datos o genera enlaces.", parent=self.root)
            return
        
        # Validaciones específicas por modo
        if self.fidelizado_mode == "NUMEROS":
            if not self.manual_inputs_numbers:
                messagebox.showerror("Error", "Modo Números requiere números cargados.", parent=self.root); return
            if not self.manual_messages_numbers:
                messagebox.showerror("Error", "Modo Números requiere mensajes cargados.", parent=self.root); return

            # Calcular total_messages para Modo Números según WhatsApp seleccionado
            num_dev = len(self._get_filtered_devices(silent=True))
            if num_dev == 0:
                messagebox.showerror("Error", "No hay dispositivos compatibles con el filtro de WhatsApp seleccionado.", parent=self.root); return
            num_numeros = len(self.manual_inputs_numbers)
            whatsapp_multiplier = 3 if self.whatsapp_mode.get() == "Todas" else (2 if self.whatsapp_mode.get() == "Ambas" else 1)
            self.total_messages = self.manual_loops * num_numeros * num_dev * whatsapp_multiplier
            wa_mode_str = self.whatsapp_mode.get()
            self.log(f"Modo Números ({wa_mode_str}): {self.total_messages} envíos totales ({self.manual_loops} ciclos x {num_numeros} números x {num_dev} disp. x {whatsapp_multiplier} app(s))", 'info')
        
        elif self.fidelizado_mode == "GRUPOS":
            if not self.manual_inputs_groups:
                messagebox.showerror("Error", "Modo Grupos requiere grupos cargados.", parent=self.root); return
            if not self.manual_messages_groups:
                messagebox.showerror("Error", "Modo Grupos requiere mensajes cargados.", parent=self.root); return

            # --- CÁLCULO CORREGIDO ---
            num_dev = len(self._get_filtered_devices(silent=True))
            if num_dev == 0:
                messagebox.showerror("Error", "No hay dispositivos compatibles con el filtro de WhatsApp seleccionado.", parent=self.root); return
            num_grupos = len(self.manual_inputs_groups)
            num_bucles = self.manual_loops_var.get()
            whatsapp_multiplier = 3 if self.whatsapp_mode.get() == "Todas" else (2 if self.whatsapp_mode.get() == "Ambas" else 1)
            self.total_messages = num_bucles * num_grupos
            wa_mode_str = self.whatsapp_mode.get()
            self.log(f"Modo Grupos ({wa_mode_str}): {self.total_messages} envíos totales ({num_bucles} bucles x {num_grupos} grupos)", 'info')
        
        elif self.fidelizado_mode == "MIXTO":
            if not self.manual_inputs_groups or not self.manual_inputs_numbers:
                messagebox.showerror("Error", "Modo Mixto requiere Grupos Y Números cargados.", parent=self.root); return
            if not self.manual_messages_numbers:
                messagebox.showerror("Error", "Modo Mixto requiere mensajes cargados.", parent=self.root); return

            # --- CÁLCULO CORREGIDO ---
            num_dev = len(self._get_filtered_devices(silent=True))
            if num_dev == 0:
                messagebox.showerror("Error", "No hay dispositivos compatibles con el filtro de WhatsApp seleccionado.", parent=self.root); return
            num_grupos = len(self.manual_inputs_groups)
            num_numeros = len(self.manual_inputs_numbers)
            num_bucles = self.manual_loops_var.get()
            whatsapp_multiplier = 3 if self.whatsapp_mode.get() == "Todas" else (2 if self.whatsapp_mode.get() == "Ambas" else 1)

            tasks_per_bucle = (num_grupos + num_numeros) * num_dev * whatsapp_multiplier
            self.total_messages = num_bucles * tasks_per_bucle
            wa_mode_str = self.whatsapp_mode.get()
            self.log(f"Modo Mixto ({wa_mode_str}): {self.total_messages} envíos totales ({num_bucles} bucles x ({num_grupos} grupos + {num_numeros} nums) x {num_dev} disp. x {whatsapp_multiplier} app(s))", 'info')

        elif self.sms_mode_active:
            self.total_messages = len(self.links)
            self.log(f"SMS: {self.total_messages} envíos totales", 'info')

        # (total_messages para otros modos ya está calculado)
        # --- Fin Validación ---

        if self.is_running:
            return

        if confirm:
            if not messagebox.askyesno("Confirmar Envío", f"¿Estás seguro de que deseas iniciar el envío de {self.total_messages} mensajes?", parent=self.root):
                return

        self.report_data = [] # Limpiar datos para el nuevo reporte
        
        # Calcular total_messages para modo tradicional según Simple/Doble/Triple
        if not self.manual_mode and not self.sms_mode_active:
            mode = self.traditional_send_mode.get()
            base_links = len(self.links)
            if mode == "Simple":
                self.total_messages = base_links
            elif mode == "Doble":
                self.total_messages = base_links * 2
            elif mode == "Triple":
                self.total_messages = base_links * 3
            self.log(f"Modo Tradicional ({mode}): {self.total_messages} envíos totales", 'info')
        
        # Limpieza de flags
        if not self.manual_mode:
            # Modo tradicional (Excel/CSV)
            self.group_mode = False
            self.fidelizado_mode = None  # No usar modo fidelizado
            self.manual_paired_messages = []

        self._enter_task_mode()
        self.update_stats() # Actualizar UI con el total

        # Reiniciar el contador de audios antes de comenzar un nuevo envío
        self.audio_message_counter = 0
        self.audio_next_trigger_count = None
        self.audio_pending_for_next_sender = False
        self.audio_pending_trigger_combo = None
        # self.audio_known_combinations.clear() # Removed as per user request to remove audio logic

        # Iniciar hilo
        threading.Thread(target=self.send_thread, daemon=True).start()

    def pause_sending(self):
        """Pausa o reanuda el envío."""
        with self.pause_lock:
            if self.is_paused:
                self.is_paused = False
                if hasattr(self, 'btn_pause'):
                    self.btn_pause.configure(text="⏸  PAUSAR")
                if hasattr(self, 'fidelizado_btn_pause'):
                    self.fidelizado_btn_pause.configure(text="⏸  PAUSAR")
                if hasattr(self, 'sms_btn_pause'):
                    self.sms_btn_pause.configure(text="⏸  PAUSAR")
                if hasattr(self, 'calls_btn_pause'):
                    self.calls_btn_pause.configure(text="⏸  PAUSAR")
                self.log("Reanudado", 'success')
            else:
                self.is_paused = True
                if hasattr(self, 'btn_pause'):
                    self.btn_pause.configure(text="▶  REANUDAR")
                if hasattr(self, 'fidelizado_btn_pause'):
                    self.fidelizado_btn_pause.configure(text="▶  REANUDAR")
                if hasattr(self, 'sms_btn_pause'):
                    self.sms_btn_pause.configure(text="▶  REANUDAR")
                if hasattr(self, 'calls_btn_pause'):
                    self.calls_btn_pause.configure(text="▶  REANUDAR")
                self.log("Pausado", 'warning')

    def stop_sending(self):
        """Solicita la detención del hilo de envío."""
        if messagebox.askyesno("Confirmar Cancelación", "¿Estás seguro de que deseas cancelar el envío actual?", parent=self.root):
            self.should_stop = True
            self.log("Cancelando...", 'warning')

    def _generate_report(self, parent_window=None):
        """Genera y guarda un reporte Excel con los resultados del envío."""
        # Asignar a self.root como padre por defecto si no se especifica
        parent = parent_window if parent_window is not None else self.root

        # --- OPTIMIZACIÓN: Calcular "No procesado" solo cuando se solicita el reporte ---
        if self.should_stop:
            self.log("Generando datos de reporte para envío cancelado...", 'info')
            processed_numbers = {d['number'] for d in self.report_data}

            if not self.manual_mode and self.links:
                for link in self.links:
                    phone_number = self._get_phone_from_link(link)
                    if phone_number and phone_number not in processed_numbers:
                        self.report_data.append({'number': phone_number, 'status': 'No procesado'})

        if not self.report_data:
            messagebox.showinfo("Sin Datos", "No hay datos de envío para generar un reporte.", parent=parent)
            return

        try:
            # Sugerir nombre de archivo
            today_str = datetime.now().strftime("%Y-%m-%d")
            default_filename = f"Reporte Hermes - {today_str}.xlsx"

            filepath = filedialog.asksaveasfilename(
                title="Guardar Reporte de Envío",
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx")],
                initialfile=default_filename,
                parent=parent
            )

            if not filepath:
                self.log("Guardado de reporte cancelado por el usuario.", 'info')
                return

            # Crear el libro de trabajo y la hoja
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Resultados de Envío"

            # Escribir encabezados
            ws['A1'] = "Numeros"
            ws['B1'] = "Estado"

            # Escribir datos
            for index, record in enumerate(self.report_data, start=2):
                ws[f'A{index}'] = record.get('number', 'N/A')
                ws[f'B{index}'] = record.get('status', 'Desconocido')

            # Auto-ajustar ancho de columnas
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            # Guardar el archivo
            wb.save(filepath)
            self.log(f"Reporte de envío guardado exitosamente en: {filepath}", 'success')
            messagebox.showinfo("Éxito", "El reporte se ha guardado correctamente.", parent=parent)

        except Exception as e:
            self.log(f"Error al generar el reporte: {e}", 'error')
            messagebox.showerror("Error", f"Ocurrió un error al generar el reporte:\n{e}", parent=parent)

    def _show_completion_dialog(self, title="Envío Completado", message="Hermes entregó tus mensajes correctamente."):
        """Muestra la ventana personalizada de finalización (MOD 28)."""
        dialog = ctk.CTkToplevel(self.root)
        dialog.title(title)
        dialog.transient(self.root); dialog.grab_set(); dialog.attributes('-topmost', True)
        dialog.resizable(False, False)

        width, height = 450, 220
        self.root.update_idletasks()
        root_x, root_y = self.root.winfo_x(), self.root.winfo_y()
        root_w, root_h = self.root.winfo_width(), self.root.winfo_height()
        x, y = root_x + (root_w // 2) - (width // 2), root_y + (root_h // 2) - (height // 2)
        dialog.geometry(f"{width}x{height}+{x}+{y}")
        dialog.after(100, dialog.focus_force)

        main_frame = ctk.CTkFrame(dialog, fg_color=self.colors['bg_card'])
        main_frame.pack(expand=True, fill=tk.BOTH, padx=20, pady=20)
        main_frame.grid_columnconfigure(0, weight=1)
        main_frame.grid_rowconfigure(0, weight=1)
        main_frame.grid_rowconfigure(1, weight=0)

        content_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        content_frame.grid(row=0, column=0, pady=(10, 20))

        try:
            logo_p = os.path.join(BASE_DIR, 'logo_left.png')
            logo_i = Image.open(logo_p).resize((60, 60), Image.Resampling.LANCZOS)
            logo_img = ctk.CTkImage(light_image=logo_i, dark_image=logo_i, size=(60, 60))
            ctk.CTkLabel(content_frame, image=logo_img, text="").pack(pady=(0, 10))
        except Exception as e:
            print(f"Error cargando logo para diálogo: {e}")

        ctk.CTkLabel(content_frame,
                     text=message,
                     font=self.fonts['dialog_text'],
                     text_color=self.colors['text'],
                     wraplength=400).pack()

        def close_dialog(e=None):
            dialog.grab_release()
            dialog.destroy()
            self.root.focus_force()

        def generate_report_and_close():
            # Pasar el diálogo como padre para que la ventana de guardar aparezca encima
            self._generate_report(parent_window=dialog)
            close_dialog()

        buttons_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        buttons_frame.grid(row=1, column=0, pady=(0, 10))
        buttons_frame.grid_columnconfigure((0, 1), weight=1, uniform="dialog_buttons")

        report_button = ctk.CTkButton(buttons_frame, text="Reporte", command=generate_report_and_close,
                                   font=self.fonts['button'],
                                   fg_color=self.colors['blue'],
                                   hover_color=darken_color(self.colors['blue'], 0.18),
                                   width=140)
        report_button.grid(row=0, column=0, sticky="e", padx=(0, 8))

        ok_button = ctk.CTkButton(buttons_frame, text="OK", command=close_dialog,
                                  font=self.fonts['button'],
                                  fg_color='transparent',
                                  text_color=self.colors['text'],
                                  border_width=1,
                                  border_color=self.colors['text_light'],
                                  hover_color=self.colors['bg'],
                                  width=140)
        ok_button.grid(row=0, column=1, sticky="w", padx=(8, 0))

        dialog.bind('<Return>', close_dialog)
        dialog.protocol("WM_DELETE_WINDOW", close_dialog)
        self.root.wait_window(dialog)

    # --- INICIO MODIFICACIÓN: send_thread (Refactorizado) ---
    def send_thread(self):
        """Hilo de trabajo que envía los mensajes uno por uno."""
        try:
            self.log("INICIANDO ENVÍO", 'success')

            # Limpieza inicial
            for dev in self.devices:
                if self.should_stop: break
                self.close_all_apps(dev)

            if self.should_stop: self.log("Cancelado", 'warning'); return
            self.log("Pausa inicial de 3s...", 'info'); time.sleep(3)
            if self.should_stop: self.log("Cancelado", 'warning'); return

            # --- Lógica de envío (depende del modo) ---
            # La variable `self.sms_mode_active` es la que determina si se usa la
            # lógica de SMS o la de WhatsApp.
            # Los modos Fidelizado (GRUPOS, NUMEROS, etc.) son sub-modos de WhatsApp.
            if self.fidelizado_mode == "GRUPOS":
                if self.simultaneous_mode.get():
                    self.run_grupos_simultaneous_thread()
                else:
                    self.run_grupos_dual_whatsapp_thread() # <-- WHATSAPP
            elif self.fidelizado_mode == "NUMEROS_MANUAL":
                self.run_numeros_manual_thread() # <-- WHATSAPP
            elif self.fidelizado_mode == "NUMEROS":
                if self.fidelizado_numeros_mode.get() == "Uno a uno":
                    self.run_uno_a_uno_thread() # <-- WHATSAPP
                else: # Uno a muchos
                    self.run_uno_a_muchos_thread() # <-- WHATSAPP
            elif self.fidelizado_mode == "MIXTO":
                self.run_mixto_dual_whatsapp_thread() # <-- WHATSAPP
            elif self.sms_mode_active:
                if self.sms_simultaneous_mode.get():
                    self.run_sms_parallel_thread() # <-- SMS SIMULTÁNEO
                else:
                    self.run_sms_thread() # <-- SMS
            else:
                self.run_default_thread() # <-- WHATSAPP (Tradicional)
            # --- Fin Lógica de envío ---

            # Finalización
            if self.should_stop:
                self.log("Cancelado. Mostrando diálogo de finalización.", 'info')
                title = "Envío Cancelado"
                message = "El envío fue cancelado. Puedes generar un reporte de los mensajes procesados."
                self.root.after(100, lambda: self._show_completion_dialog(title=title, message=message))
            else:
                self.log("ENVÍO FINALIZADO", 'success')
                self.log(f"Resumen: Enviados: {self.sent_count} | Fallidos: {self.failed_count}", 'info')
                self.root.after(100, self._show_completion_dialog)

        except Exception as e:
            self.log(f"Error CRÍTICO en el hilo de envío: {e}", 'error')
            import traceback
            traceback_str = traceback.format_exc()
            print(f"ERROR THREAD ENVIO:\n{traceback_str}")
            self.root.after(100, lambda: messagebox.showerror("Error Crítico", f"Ocurrió un error inesperado durante el envío:\n{e}\n\nRevise el log para más detalles.", parent=self.root))
        finally:
            # Siempre reestablecer la UI
            self.root.after(100, self._finalize_sending)
    
    def _maybe_pause_sending(self, task_index):
        """Checks if a pause is needed based on the task index and configured settings."""
        pause_count = self.pause_sends_count.get()
        if pause_count <= 0:
            return

        # Check if the current task index is a multiple of the pause count
        if task_index > 0 and task_index % pause_count == 0:
            delay_min = self.pause_sends_delay_min.get()
            delay_max = self.pause_sends_delay_max.get()
            pause_duration = random.uniform(delay_min, delay_max)

            self.log(f"Pausa programada alcanzada. Esperando {pause_duration:.1f} segundos...", 'info')
            self._controlled_sleep(pause_duration)

    def run_single_task(self, device, link, message_to_send, task_index, whatsapp_package="com.whatsapp.w4b", link_index=None, activity_info=None, skip_delay=False):
        """
        Ejecuta una única tarea de envío (abrir link, enviar, esperar), gestionando la conexión de uiautomator2.
        Args:
            skip_delay (bool): Si es True, omite el retardo posterior al envío (útil para modo simultáneo).
        """
        ui_device = None
        # --- MODIFICACIÓN: Siempre intentar conectar uiautomator2 ---
        try:
            ui_device = u2.connect(device)
            ui_device.unlock()
        except Exception as e:
            self.log(f"No se pudo conectar uiautomator2 a {device}: {e}", "warning")
            # En este nuevo enfoque, un fallo aquí debería ser crítico.
            return False

        # Bucle de pausa
        while self.is_paused and not self.should_stop:
            time.sleep(0.1)
        if self.should_stop: return False

        self.current_index = task_index
        self.root.after(0, self.update_stats)

        # En modo fidelizado ya se realiza una limpieza general antes de arrancar el hilo.
        # Evitamos cerrar las apps en cada mensaje para reducir el tiempo extra entre envíos
        # y que el intervalo se acerque más al configurado por el usuario.
        if not self.manual_mode:
            self.close_all_apps(device)

        if self.should_stop:
            return False

        send_start = time.time()
        # Enviar mensaje, pasando el objeto ui_device
        success = self.send_msg(
            device,
            link,
            task_index,
            self.total_messages,
            message_to_send,
            whatsapp_package,
            ui_device,
            primary_index=link_index,
        )
        
        # --- Importante: Actualizar contadores y reporte DESPUÉS de send_msg ---
        phone_number = self._get_phone_from_link(link)
        if phone_number:
            status = "Enviado" if success else "Fallo"
            self.report_data.append({'number': phone_number, 'status': status})

        if success:
            self.sent_count += 1

            # --- Lógica de Llamada (Solo Modo SMS) ---
            if self.sms_mode_active and self.sms_enable_call.get():
                try:
                    # Extraer número de sms:NUMBER?body=...
                    # target_link viene como "sms:12345?body=..."
                    number_part = link.split(':')[1]
                    if '?' in number_part:
                        number = number_part.split('?')[0]
                    else:
                        number = number_part

                    if number:
                        self.log(f"[{device}] Iniciando llamada a {number}...", 'info')
                        self._perform_call(device, number, self.sms_call_duration.get())
                except Exception as e:
                    self.log(f"[{device}] Error al intentar realizar llamada: {e}", 'error')

            if activity_info and self.fidelizado_mode == "NUMEROS":
                sender = activity_info.get('from')
                receiver = activity_info.get('to')
                self.record_fidelizado_activity(sender, receiver)
        else:
            self.failed_count += 1

        # Actualizar UI (contadores y barra de progreso)
        self.root.after(0, self.update_stats)
        # --- Fin actualización contadores ---

        # Espera entre mensajes (solo si no es la última tarea y no se omite)
        if task_index < self.total_messages and not self.should_stop and not skip_delay:
            send_elapsed = time.time() - send_start
            if self.manual_mode:
                # FIX: Usar las nuevas variables de retardo de envío
                target_delay = random.uniform(self.fidelizado_send_delay_min.get(), self.fidelizado_send_delay_max.get())
            elif self.sms_mode_active:
                target_delay = random.uniform(self.sms_delay_min.get(), self.sms_delay_max.get())
            else:
                target_delay = random.uniform(self.delay_min.get(), self.delay_max.get())

            remaining_delay = max(0, target_delay - send_elapsed)
            self.log(
                f"Esperando {remaining_delay:.1f}s (objetivo {target_delay:.1f}s, envío tomó {send_elapsed:.1f}s)"
                f"... (Post-tarea {task_index})",
                'info'
            )
            elapsed = 0
            while elapsed < remaining_delay and not self.should_stop:
                while self.is_paused and not self.should_stop: time.sleep(0.1)
                if self.should_stop: break
                time.sleep(0.1); elapsed += 0.1
        
        return success

    def run_default_thread(self):
        """
        Lógica de envío tradicional (Excel/CSV) con soporte para Simple/Doble/Triple.
        """
        if not self.links:
            self.log("Error: No hay links para enviar (modo tradicional)", 'error')
            return
        
        mode = self.traditional_send_mode.get()
        self.log(f"Modo de envío: {mode}", 'info')
        
        if mode == "Business":
            self._run_simple_mode()
        elif mode == "Normal":
            # Reutiliza _run_simple_mode pero cambiando el paquete de WA
            self._run_simple_mode(whatsapp_package="com.whatsapp")
        elif mode == "Business/Normal":
            self._run_doble_mode()
        elif mode == "Business/Normal 1/Normal 2":
            self._run_triple_mode()
    
    def _run_simple_mode(self, whatsapp_package="com.whatsapp.w4b"):
        """Modo Simple: 1 URL por teléfono, usando el paquete de WhatsApp especificado."""
        log_msg = "Ejecutando Modo Business..." if whatsapp_package == "com.whatsapp.w4b" else "Ejecutando Modo Normal..."
        self.log(log_msg, 'info')
        idx = 0  # Índice del dispositivo a usar
        
        for i, link in enumerate(self.links):
            if self.should_stop:
                self.log("Cancelado en bucle", 'warning')
                break

            device = self.devices[idx]
            idx = (idx + 1) % len(self.devices)

            # Check for pause before running the task
            self._maybe_pause_sending(i)

            # Ejecutar tarea con el paquete de WA especificado
            self.run_single_task(device, link, None, i + 1, whatsapp_package=whatsapp_package, link_index=i)

    def _run_doble_mode(self):
        """Modo Doble: Rota secuencialmente entre dispositivos y cuentas Business/Normal."""
        self.log("Ejecutando Modo Doble (Rotación Correcta)...", 'info')

        # 1. Crear la lista de todas las combinaciones de envío
        envio_combinations = []
        for device in self.devices:
            envio_combinations.append({"device": device, "wa_name": "Business", "wa_package": "com.whatsapp.w4b"})
            envio_combinations.append({"device": device, "wa_name": "Normal", "wa_package": "com.whatsapp"})

        num_combinations = len(envio_combinations)
        if num_combinations == 0:
            self.log("Error: No hay dispositivos para el modo Doble.", "error")
            return

        # 2. Iterar una vez sobre los links, rotando las combinaciones
        for i, link in enumerate(self.links):
            if self.should_stop:
                self.log("Cancelado en bucle", 'warning')
                break
            
            self.last_task_time = time.time()

            # Check for pause before running the task
            self._maybe_pause_sending(i)

            # Seleccionar la combinación de envío (dispositivo + cuenta)
            combination = envio_combinations[i % num_combinations]
            device = combination["device"]
            wa_name = combination["wa_name"]
            wa_package = combination["wa_package"]

            self.log(f"[{device}] Enviando con {wa_name}", 'info')
            self.run_single_task(device, link, None, i + 1, whatsapp_package=wa_package, link_index=i)

    def _run_triple_mode(self):
        """Modo Triple: Rota secuencialmente entre dispositivos y las 3 cuentas."""
        self.log("Ejecutando Modo Triple (Rotación Correcta)...", 'info')
        is_normal_account_2 = {dev_id: False for dev_id in self.devices}

        # 1. Crear la lista de todas las combinaciones de envío
        envio_combinations = []
        for device in self.devices:
            envio_combinations.append({"device": device, "wa_name": "Business", "wa_package": "com.whatsapp.w4b", "needs_switch": False})
            envio_combinations.append({"device": device, "wa_name": "Normal (Cuenta 1)", "wa_package": "com.whatsapp", "needs_switch": False})
            envio_combinations.append({"device": device, "wa_name": "Normal (Cuenta 2)", "wa_package": "com.whatsapp", "needs_switch": True})

        num_combinations = len(envio_combinations)
        if num_combinations == 0:
            self.log("Error: No hay dispositivos para el modo Triple.", "error")
            return

        # 2. Iterar una vez sobre los links, rotando las combinaciones
        for i, link in enumerate(self.links):
            if self.should_stop:
                self.log("Cancelado en bucle", 'warning')
                break

            self.last_task_time = time.time()

            # Check for pause before running the task
            self._maybe_pause_sending(i)

            # Seleccionar la combinación de envío (dispositivo + cuenta)
            combination = envio_combinations[i % num_combinations]
            device = combination["device"]
            wa_name = combination["wa_name"]
            wa_package = combination["wa_package"]
            needs_switch_to_acc2 = combination["needs_switch"]
            
            # Gestionar cambio de cuenta si es necesario
            if "Normal" in wa_name:
                currently_is_acc2 = is_normal_account_2.get(device, False)
                if needs_switch_to_acc2 and not currently_is_acc2:
                    self.log(f"Cambiando a Cuenta 2 en {device}...", 'info')
                    self._switch_whatsapp_account(device)
                    time.sleep(4)
                    is_normal_account_2[device] = True
                elif not needs_switch_to_acc2 and currently_is_acc2:
                    self.log(f"Restaurando a Cuenta 1 en {device}...", 'info')
                    self._switch_whatsapp_account(device)
                    time.sleep(4)
                    is_normal_account_2[device] = False

            if self.should_stop: break

            self.log(f"[{device}] Enviando con {wa_name}", 'info')
            self.run_single_task(device, link, None, i + 1, whatsapp_package=wa_package, link_index=i)

        # Dejar todas las cuentas Normal en el estado inicial (Cuenta 1)
        self.log("Finalizando y restaurando cuentas a estado inicial...", 'info')
        for dev, is_acc2 in is_normal_account_2.items():
            if is_acc2:
                self.log(f"Restaurando a Cuenta 1 en {dev}...", 'info')
                self._switch_whatsapp_account(dev)
                time.sleep(4)

    def run_sms_thread(self):
        """Lógica de envío para el SMS."""
        if not self.links:
            self.log("Error: No hay enlaces SMS para enviar.", 'error')
            return

        self.log("Ejecutando SMS...", 'info')
        idx = 0

        for i, link in enumerate(self.links):
            if self.should_stop:
                self.log("Cancelado en SMS", 'warning')
                break

            device = self.devices[idx]
            idx = (idx + 1) % len(self.devices)

            self.run_single_task(device, link, None, i + 1, whatsapp_package=None, link_index=i)

    def start_calling(self):
        """Envoltorio para iniciar llamadas con opción de programar."""
        self._open_schedule_dialog(lambda confirm: self._start_calling_impl(confirm=confirm))

    def _start_calling_impl(self, confirm=True):
        """Inicia el proceso de llamadas masivas (Lógica Interna)."""
        if not self.raw_data:
            messagebox.showwarning("Sin datos", "Carga un archivo Excel primero.")
            return

        # Validar selección de columnas
        # self.calls_selected_columns ya fue seteado por el popup _on_calls_columns_confirmed
        if not self.calls_selected_columns:
            messagebox.showwarning("Sin columnas", "Selecciona al menos una columna de teléfono (Carga el Excel y usa el popup).")
            return

        if not self.devices:
            self.detect_devices()
            if not self.devices:
                return

        if confirm:
            if not messagebox.askyesno("Confirmar", "¿Iniciar proceso de llamadas masivas?", parent=self.root):
                return

        self._enter_task_mode()
        threading.Thread(target=self.run_calls_thread, daemon=True).start()

    def run_calls_thread(self):
        """Hilo principal para la lógica de llamadas."""
        try:
            self.log("🚀 INICIANDO LLAMADAS MASIVAS", 'info')

            # Generar lista de tareas
            tasks = []
            for row in self.raw_data:
                for col in self.calls_selected_columns:
                    val = str(row.get(col, '')).strip()
                    if val:
                        # Usar regex para dividir por guiones (todos los tipos), comas, punto y coma, saltos de línea
                        nums = [n.strip() for n in re.split(r'[-–—,\n;]', val) if n.strip()]
                        for n in nums:
                            # Clean number
                            clean_n = ''.join(filter(str.isdigit, n))
                            if clean_n:
                                tasks.append(clean_n)

            self.total_messages = len(tasks)
            self.root.after(0, self.update_stats)
            self.log(f"Total números a llamar: {len(tasks)}", 'info')

            wa_mode = self.calls_whatsapp_mode.get() # Business, Normal, Ambos
            simultaneous = self.calls_simultaneous_mode.get()
            delay_min = self.calls_delay_min.get()
            delay_max = self.calls_delay_max.get()
            duration = self.calls_duration.get()

            # Definir estrategia de apps
            # Si es "Ambos", alternamos. Si no, fijo.
            # Alternancia: Business -> Normal -> Business ...

            def get_app_sequence():
                if wa_mode == "Business": return [("com.whatsapp.w4b", "Business")]
                if wa_mode == "Normal": return [("com.whatsapp", "Normal")]
                if wa_mode == "Ambos": return [("com.whatsapp.w4b", "Business"), ("com.whatsapp", "Normal")]
                return [("com.whatsapp.w4b", "Business")]

            app_seq = get_app_sequence()

            # Estado de rotación por dispositivo
            device_app_idx = {d: 0 for d in self.devices}

            def process_call(device_id, phone_number, task_idx):
                if self.should_stop: return False

                # 0. Cerrar ambos WhatsApps para asegurar estado limpio
                # El usuario solicitó: "cada vez que se envie y al empezar el primer, que cierres ambos whatsapp"
                self.log(f"[{device_id}] Cerrando WhatsApps antes de la tarea...", 'info')
                self._run_adb_command(['-s', device_id, 'shell', 'am', 'force-stop', 'com.whatsapp'])
                self._run_adb_command(['-s', device_id, 'shell', 'am', 'force-stop', 'com.whatsapp.w4b'])
                time.sleep(1)

                # Seleccionar App
                seq_idx = device_app_idx[device_id]
                pkg, app_name = app_seq[seq_idx % len(app_seq)]
                device_app_idx[device_id] += 1 # Rotar para la próxima

                self.current_index = task_idx
                self.root.after(0, self.update_stats)

                log_prefix = f"({task_idx}/{len(tasks)}) 📞 {phone_number} [{device_id}] ({app_name})"
                self.log(f"Llamando a {phone_number}...", 'info')

                # 1. Abrir Home de WhatsApp (No abrir chat directo con link, sino la app)
                # Usamos el launch normal
                open_args = ['-s', device_id, 'shell', 'monkey', '-p', pkg, '-c', 'android.intent.category.LAUNCHER', '1']
                if not self._run_adb_command(open_args, timeout=10):
                    self.log(f"{log_prefix} ✗ Fallo al abrir WhatsApp", 'error')
                    self.failed_count += 1
                    return

                # 2. Esperar Home
                time.sleep(3)

                # 3. Ejecutar flujo de llamada (Nuevo)
                if self._perform_whatsapp_call_action(device_id, duration, phone_number=phone_number, package_name=pkg):
                    self.log(f"{log_prefix} ✓ Llamada finalizada", 'success')
                    self.sent_count += 1
                else:
                    self.log(f"{log_prefix} ✗ No se pudo realizar la llamada (¿Sin WhatsApp?)", 'error')
                    self.failed_count += 1

                # 4. Delay entre llamadas
                wait = random.uniform(delay_min, delay_max)
                self.log(f"Pausa de {wait:.1f}s...", 'info')
                self._controlled_sleep(wait)

            if simultaneous and len(self.devices) > 1:
                # Modo paralelo
                import concurrent.futures

                # Dividir tareas para round-robin
                num_devs = len(self.devices)

                # Ejecutar en lotes del tamaño de devices
                for i in range(0, len(tasks), num_devs):
                    if self.should_stop: break
                    while self.is_paused and not self.should_stop: time.sleep(0.1)

                    batch = tasks[i : i + num_devs]

                    with concurrent.futures.ThreadPoolExecutor(max_workers=num_devs) as executor:
                        futures = []
                        for b_idx, phone in enumerate(batch):
                            dev = self.devices[b_idx % num_devs]
                            t_idx = i + b_idx + 1
                            futures.append(executor.submit(process_call, dev, phone, t_idx))

                        # Esperar lote
                        concurrent.futures.wait(futures)
            else:
                # Modo secuencial (1 dispositivo o 1 por 1)
                dev_idx = 0
                for i, phone in enumerate(tasks):
                    if self.should_stop: break
                    while self.is_paused and not self.should_stop: time.sleep(0.1)

                    dev = self.devices[dev_idx]
                    dev_idx = (dev_idx + 1) % len(self.devices) # Round robin simple

                    process_call(dev, phone, i + 1)

        except Exception as e:
            self.log(f"Error en hilo de llamadas: {e}", 'error')
        finally:
            self.log("Finalizando proceso de llamadas...", 'info')
            self._finalize_sending()

    def _perform_whatsapp_call_action(self, device, duration, phone_number=None, package_name="com.whatsapp"):
        """
        NUEVA LÓGICA DE LLAMADA:
        1. Click en el 1er chat (fijado).
        2. Escribir número y enviar.
        3. Click en el mensaje enviado.
        4. Esperar popup 'Llamar en Whatsapp'.
        5. Click en esa opción.
        6. Esperar y colgar.
        """
        try:
            d = u2.connect(device)

            # 1. Click en el primer chat de la lista (asumimos que ya estamos en la Home de WA)
            # Intentar encontrar el primer item de la lista de conversaciones
            self.log(f"[{device}] Buscando primer chat (fijado)...", 'info')

            # Estrategia 1: Por ID de la fila de conversación
            row_selector = d(resourceId=f"{package_name}:id/conversations_row_contact_name")

            chat_clicked = False
            if row_selector.exists:
                # Click en el primero (índice 0)
                try:
                    row_selector[0].click()
                    chat_clicked = True
                except:
                    pass

            if not chat_clicked:
                # Estrategia 2: Click por coordenadas relativas (arriba de la lista)
                # Asumiendo una resolución estándar, el primer chat suele estar en el top 20-30%
                self.log(f"[{device}] Usando click relativo para primer chat...", 'warning')
                w, h = d.window_size()
                d.click(w // 2, int(h * 0.25)) # Click en el cuarto superior central
                chat_clicked = True

            time.sleep(2) # Esperar a que abra el chat

            # 2. Escribir número y enviar
            self.log(f"[{device}] Enviando número {phone_number}...", 'info')

            # Buscar campo de texto (usando helper existente)
            chat_field = self._locate_chat_field(d, wait_timeout=5)

            if not chat_field:
                self.log(f"[{device}] No se encontró campo de texto.", 'error')
                return False

            try:
                # 1. Escribir primero para que aparezca el botón de enviar
                chat_field.click()
                time.sleep(0.5)
                chat_field.set_text(phone_number)
                time.sleep(1.0) # Esperar a que el UI actualice y muestre el botón enviar

                # 2. Buscar botón de enviar AHORA (estrictamente el avión, evitando 'Reenviar')
                send_selectors = [
                    dict(resourceId=f"{package_name}:id/send"),
                    dict(contentDescription="Enviar"),
                    dict(description="Enviar"),
                    dict(resourceId=f"{package_name}:id/send_button")
                ]

                send_btn = None
                for sel in send_selectors:
                    btn = d(**sel)
                    if btn.exists:
                        send_btn = btn
                        break

                if send_btn:
                    send_btn.click()
                else:
                    self.log(f"[{device}] Botón enviar no encontrado, intentando Enter...", 'warning')
                    self._run_adb_command(['-s', device, 'shell', 'input', 'keyevent', 'KEYCODE_ENTER'])

                time.sleep(1.5) # Esperar a que se envíe y aparezca la burbuja
            except Exception as e:
                self.log(f"[{device}] Error al enviar número: {e}", 'error')
                return False

            # 3. Click en el mensaje enviado (el último)
            self.log(f"[{device}] Buscando mensaje enviado...", 'info')

            # Buscar el último mensaje que contenga el número
            # Los mensajes suelen tener id 'message_text'
            msg_selector = d(resourceId=f"{package_name}:id/message_text", text=phone_number)

            if not msg_selector.exists:
                # Intentar buscar por texto contiene (por si hay formato)
                msg_selector = d(resourceId=f"{package_name}:id/message_text", textContains=phone_number[-4:]) # Últimos 4 dígitos

            if msg_selector.exists:
                # Hacer click en la última instancia encontrada (la más reciente)
                try:
                    count = msg_selector.count
                    if count > 0:
                        msg_selector[count - 1].click()
                    else:
                        msg_selector.click()
                except Exception as e:
                    self.log(f"[{device}] Error al clickear mensaje: {e}", 'error')
                    return False
            else:
                self.log(f"[{device}] No se encontró la burbuja del mensaje con el número.", 'error')
                return False

            time.sleep(1.5) # Esperar al popup

            # 4. Esperar popup 'Llamar en Whatsapp' y 5. Clickear
            self.log(f"[{device}] Buscando opción 'Llamar en Whatsapp'...", 'info')

            popup_options = [
                "Llamar en WhatsApp",
                "Llamar en Whatsapp",
                "Call on WhatsApp",
                "Llamada de voz"
            ]

            call_option_clicked = False
            for opt in popup_options:
                btn = d(text=opt)
                if btn.exists:
                    btn.click()
                    call_option_clicked = True
                    self.log(f"[{device}] Opción '{opt}' encontrada y clickeada.", 'success')
                    break

            if not call_option_clicked:
                # Intentar búsqueda laxa
                btn = d(textContains="Llamar")
                if btn.exists:
                     # Verificar que no sea "Llamar" de llamada normal (GSM) si se distingue
                     # Pero el usuario dijo "Llamar en Whatsapp", así que priorizamos eso.
                     # Si sale solo "Llamar", podría ser GSM.
                     # Busquemos especificamente el texto que dijo el usuario.
                     pass

                self.log(f"[{device}] No se encontró la opción de llamada de WhatsApp. ¿Tiene WhatsApp este número?", 'warning')
                # Cerrar el popup (click fuera o back)
                d.press("back")
                return False

            # 6. Esperar y colgar
            self.log(f"[{device}] Llamada iniciada. Esperando {duration}s...", 'info')

            # Esperar duración
            elapsed = 0
            while elapsed < duration:
                if self.should_stop: break
                time.sleep(1)
                elapsed += 1

            # Colgar (Botón rojo)
            self.log(f"[{device}] Colgando...", 'info')

            # Buscar botón de colgar (rojo)
            # IDs comunes: footer_end_call_btn, end_call_btn
            end_btn_selectors = [
                dict(resourceId=f"{package_name}:id/footer_end_call_btn"),
                dict(resourceId="com.whatsapp:id/footer_end_call_btn"),
                dict(resourceId="com.whatsapp.w4b:id/footer_end_call_btn"),
                dict(descriptionMatches="(?i)fin.*llamada|cortar|end call|colgar"),
                dict(resourceIdMatches="(?i).*end_call.*")
            ]

            end_clicked = False
            for sel in end_btn_selectors:
                btn = d(**sel)
                if btn.exists:
                    btn.click()
                    end_clicked = True
                    break

            if not end_clicked:
                # Fallback: ADB keycode endcall
                self.log(f"[{device}] Usando KEYCODE_ENDCALL para colgar...", 'info')
                self._run_adb_command(['-s', device, 'shell', 'input', 'keyevent', 'KEYCODE_ENDCALL'])

            # Volver al inicio (Home de la app)
            time.sleep(1)
            # Presionar back un par de veces para salir del chat y volver a la lista
            d.press("back")
            time.sleep(0.5)
            d.press("back")

            return True

        except Exception as e:
            self.log(f"[{device}] Error en _perform_whatsapp_call_action: {e}", 'error')
            return False


    def run_sms_parallel_thread(self):
        """Lógica de envío SIMULTÁNEO para el SMS."""
        if not self.links:
            self.log("Error: No hay enlaces SMS para enviar.", 'error')
            return

        self.log("Ejecutando SMS en modo SIMULTÁNEO...", 'info')

        num_devices = len(self.devices)
        total_links = len(self.links)

        # Process in batches
        for i in range(0, total_links, num_devices):
            # Check for pause
            while self.is_paused and not self.should_stop:
                time.sleep(0.1)

            if self.should_stop:
                self.log("Cancelado en SMS Simultáneo", 'warning')
                break

            batch_links = self.links[i : i + num_devices]
            threads = []

            self.log(f"--- Iniciando lote simultáneo ({len(batch_links)} envíos) ---", 'info')

            for j, link in enumerate(batch_links):
                # Ensure we have enough devices (cycle if fewer devices than batch size - unlikely here as batch size is len(devices))
                device = self.devices[j % num_devices]
                # Task index logic: global index
                task_idx = i + j + 1

                # Create thread
                t = threading.Thread(
                    target=self.run_single_task,
                    args=(device, link, None, task_idx),
                    kwargs={
                        'whatsapp_package': None,
                        'link_index': i+j,
                        'skip_delay': True # NEW ARGUMENT: Skip individual delay
                    }
                )
                threads.append(t)
                t.start()

            # Wait for all threads in this batch to finish
            for t in threads:
                t.join()

            if self.should_stop: break

            # Global delay between batches (if not finished)
            if i + num_devices < total_links:
                delay = random.uniform(self.sms_delay_min.get(), self.sms_delay_max.get())
                self.log(f"⏳ Esperando {delay:.1f}s antes del siguiente lote simultáneo...", 'info')
                self._controlled_sleep(delay)

    def run_numeros_manual_thread(self):
        """Lógica de envío para MODO NÚMEROS - MANUAL."""
        self.log("Iniciando MODO NÚMEROS (Manual)...", 'info')

        active_devices = self._get_filtered_devices()
        if not active_devices:
            self.log("No hay dispositivos que cumplan con el filtro.", "error")
            return

        num_devices = len(active_devices)
        num_bucles = self.manual_loops_var.get()
        task_counter = 0
        mensaje_index = self.mensaje_start_index
        total_mensajes_lib = len(self.manual_messages_numbers)
        whatsapp_apps = self._get_whatsapp_apps_to_use()

        for bucle_num in range(num_bucles):
            if self.should_stop: break
            self.log(f"\n--- BUCLE {bucle_num + 1}/{num_bucles} ---", 'info')

            for num_idx, numero in enumerate(self.manual_inputs_numbers):
                if self.should_stop: break

                device = active_devices[num_idx % num_devices]

                for wa_idx, (wa_name, wa_package) in enumerate(whatsapp_apps):
                    if self.should_stop: break
                    task_counter += 1
                    mensaje = self.manual_messages_numbers[mensaje_index % total_mensajes_lib]; mensaje_index += 1
                    link = f"https://wa.me/549{numero}?text={urllib.parse.quote(mensaje, safe='')}"

                    self.run_single_task(device, link, None, task_counter, whatsapp_package=wa_package)
    
    def _get_filtered_devices(self, silent=False):
        """
        Filtra self.devices según el tipo de WhatsApp seleccionado.
        Devuelve una lista de series de dispositivos a utilizar.
        """
        whatsapp_mode = self.whatsapp_mode.get()

        if whatsapp_mode == "Ambas":
            if not silent:
                self.log("Modo 'Ambas' seleccionado, usando todos los dispositivos.", 'info')
            return self.devices

        def _log(msg, level='info'):
            if not silent:
                self.log(msg, level)

        if not self.detected_phone_lines:
            inferred_lines = self._infer_lines_from_installed_apps(log_errors=not silent)
            if inferred_lines:
                allowed_serials = {line['device'] for line in inferred_lines if line['type'] == ("WhatsApp" if whatsapp_mode == "Normal" else "WhatsApp Business")}
                filtered_devices = [device for device in self.devices if device in allowed_serials]
                if filtered_devices:
                    _log("No se detectaron números, se usarán las apps instaladas para filtrar dispositivos.", 'info')
                    return filtered_devices

            _log("No se han detectado líneas de WhatsApp. No se puede filtrar. Usando todos los dispositivos.", 'warning')
            return self.devices

        target_type = "WhatsApp" if whatsapp_mode == "Normal" else "WhatsApp Business"

        allowed_serials = {line['device'] for line in self.detected_phone_lines if line['type'] == target_type}

        filtered_devices = [device for device in self.devices if device in allowed_serials]

        _log(f"Filtrando por '{whatsapp_mode}'. Dispositivos a usar: {len(filtered_devices)} de {len(self.devices)}", 'info')

        return filtered_devices

    def _get_whatsapp_apps_to_use(self):
        """
        Retorna una lista de tuplas (nombre, package) según la selección del usuario.
        Opciones: Business, Normal, Ambas (Business + Normal), Todas (Business + Normal + Normal con cambio de cuenta)
        """
        wa_mode = self.whatsapp_mode.get()
        
        if wa_mode == "Normal":
            return [("Normal", "com.whatsapp")]
        elif wa_mode == "Business":
            return [("Business", "com.whatsapp.w4b")]
        elif wa_mode == "Ambas":
            return [("Business", "com.whatsapp.w4b"), ("Normal", "com.whatsapp")]
        else:  # "Todas"
            # Business + Normal (cuenta 1) + Normal (cuenta 2 después de cambio automático)
            return [("Business", "com.whatsapp.w4b"), ("Normal", "com.whatsapp"), ("Normal", "com.whatsapp")]
    
    
    def _send_to_target_with_whatsapp(self, device, target_link, wa_name, wa_package, mensaje, task_counter):
        """
        Envía un mensaje a un target, gestionando la conexión de uiautomator2 y usando el método de envío unificado.
        """
        # Conectar con uiautomator2 al inicio de la tarea
        ui_device = None
        try:
            ui_device = u2.connect(device)
            ui_device.unlock()
        except Exception as e:
            self.log(f"No se pudo conectar uiautomator2 a {device}: {e}", "warning")
            ui_device = None

        # Llamar a send_msg, que ahora contiene toda la lógica de envío y fallback
        success = self.send_msg(device, target_link, task_counter, self.total_messages, message_to_send=mensaje, whatsapp_package=wa_package, ui_device=ui_device)
        
        if success:
            self.sent_count += 1
        else:
            self.failed_count += 1
        self.root.after(0, self.update_stats)
        
        # Si es Normal y el modo es "Todas", ejecutar el cambio de cuenta DESPUÉS de enviar
        # IMPORTANTE: Si se envió audio, restablecer el contador para que la siguiente cuenta no lo envíe
        # a menos que le toque aleatoriamente.
        if wa_name == "Normal" and self.whatsapp_mode.get() == "Todas":
            self.log(f"[{device}] Cerrando WhatsApp Normal después de enviar...", 'info')
            close_cmd = ['-s', device, 'shell', 'am', 'force-stop', 'com.whatsapp']
            self._run_adb_command(close_cmd, timeout=5)
            time.sleep(1)
            
            self.log(f"[{device}] Reabriendo WhatsApp Normal...", 'info')
            open_cmd = ['-s', device, 'shell', 'am', 'start', '-n', 'com.whatsapp/.Main']
            self._run_adb_command(open_cmd, timeout=5)
            time.sleep(3)  # Esperar 3 segundos para que WhatsApp se abra completamente
            
            self.log(f"[{device}] Cambiando de cuenta...", 'info')
            self._switch_account_for_device(device, delay=0.2)
            time.sleep(4)
            
            self.log(f"[{device}] Cerrando WhatsApp Normal después de cambiar cuenta...", 'info')
            close_cmd = ['-s', device, 'shell', 'am', 'force-stop', 'com.whatsapp']
            self._run_adb_command(close_cmd, timeout=5)
            time.sleep(1)
            
            self.log(f"[{device}] Reabriendo WhatsApp Normal con nueva cuenta...", 'info')
            open_cmd = ['-s', device, 'shell', 'am', 'start', '-n', 'com.whatsapp/.Main']
            self._run_adb_command(open_cmd, timeout=5)
            time.sleep(2)

            # Resetear pending audio para la nueva cuenta si no está habilitado globalmente
            # o si la lógica de 'una cuenta a la vez' estaba causando repeticiones
        
        return True
    
    def run_mixto_dual_whatsapp_thread(self):
        """
        Lógica de envío para MODO MIXTO con 3 variantes.
        Variante 1: G1→N1→G2→N2 (alternar 1 a 1)
        Variante 2: G1→G2→N1→G3→G4→N2 (2 grupos por número)
        Variante 3: G1→G2→G3→N1→G4→G5→G6→N2 (3 grupos por número)
        
        Los grupos y números se repiten en bucle si hay menos de los necesarios.
        Todas las líneas (dispositivos) siguen la misma secuencia.
        """
        # --- FILTRADO DE DISPOSITIVOS ---
        active_devices = self._get_filtered_devices()
        if not active_devices:
            self.log("No hay dispositivos que cumplan con el filtro seleccionado.", "error")
            self.root.after(0, lambda: messagebox.showerror("Error", "No se encontraron dispositivos que cumplan con el filtro de WhatsApp seleccionado.", parent=self.root))
            return
        # --- FIN FILTRADO ---

        num_devices = len(active_devices)
        num_grupos = len(self.manual_inputs_groups)
        num_numeros = len(self.manual_inputs_numbers)
        num_bucles = self.manual_loops
        variant = self.mixto_variant.get()
        
        if len(self.manual_messages_numbers) < 1:
            self.log("Error: Modo Mixto requiere al menos 1 mensaje cargado.", "error")
            messagebox.showerror("Error", "Debes cargar al menos 1 archivo de mensajes para el modo mixto.", parent=self.root)
            return
        
        # Usar índice de inicio aleatorio
        mensaje_index = self.mensaje_start_index
        total_mensajes = len(self.manual_messages_numbers)
        task_counter = 0
        whatsapp_apps = self._get_whatsapp_apps_to_use()
        
        variant_names = {1: "1:1", 2: "2:1", 3: "3:1"}
        self.log(f"Modo Mixto (Variante {variant} - {variant_names[variant]}): {num_bucles} ciclo(s), {num_grupos} grupo(s), {num_numeros} número(s), {num_devices} dispositivo(s)", 'info')
        self.log(f"WhatsApp: {self.whatsapp_mode.get()}", 'info')
        self.log(f"Total de envíos: {self.total_messages}", 'info')
        
        for ciclo in range(num_bucles):
            if self.should_stop: break
            self.log(f"\n--- CICLO {ciclo + 1}/{num_bucles} ---", 'info')
            
            # Crear lista de targets según la variante
            targets = []
            
            if variant == 1:
                # Variante 1: G1→N1→G2→N2 (alternar 1 a 1)
                max_len = max(num_grupos, num_numeros)
                for i in range(max_len):
                    grupo_idx = i % num_grupos
                    numero_idx = i % num_numeros
                    targets.append(('grupo', grupo_idx, self.manual_inputs_groups[grupo_idx]))
                    targets.append(('numero', numero_idx, self.manual_inputs_numbers[numero_idx]))
                    
            elif variant == 2:
                # Variante 2: G1→G2→N1→G3→G4→N2 (2 grupos por número)
                grupo_idx = 0
                for num_idx in range(num_numeros):
                    # Añadir 2 grupos
                    for _ in range(2):
                        targets.append(('grupo', grupo_idx % num_grupos, self.manual_inputs_groups[grupo_idx % num_grupos]))
                        grupo_idx += 1
                    # Añadir 1 número
                    targets.append(('numero', num_idx, self.manual_inputs_numbers[num_idx]))
                # Si sobran grupos, continuar añadiendo en bucle
                if grupo_idx < num_grupos:
                    remaining = num_grupos - grupo_idx
                    for _ in range(remaining):
                        targets.append(('grupo', grupo_idx % num_grupos, self.manual_inputs_groups[grupo_idx % num_grupos]))
                        grupo_idx += 1
                        
            elif variant == 3:
                # Variante 3: G1→G2→G3→N1→G4→G5→G6→N2 (3 grupos por número)
                grupo_idx = 0
                for num_idx in range(num_numeros):
                    # Añadir 3 grupos
                    for _ in range(3):
                        targets.append(('grupo', grupo_idx % num_grupos, self.manual_inputs_groups[grupo_idx % num_grupos]))
                        grupo_idx += 1
                    # Añadir 1 número
                    targets.append(('numero', num_idx, self.manual_inputs_numbers[num_idx]))
                # Si sobran grupos, continuar añadiendo en bucle
                if grupo_idx < num_grupos:
                    remaining = num_grupos - grupo_idx
                    for _ in range(remaining):
                        targets.append(('grupo', grupo_idx % num_grupos, self.manual_inputs_groups[grupo_idx % num_grupos]))
                        grupo_idx += 1
            
            # Procesar cada target en la secuencia
            for target_type, target_idx, target_value in targets:
                if self.should_stop: break
                
                tipo_str = "Grupo" if target_type == 'grupo' else "Número"
                if target_type == 'grupo':
                    self.log(f"\n=== GRUPO {target_idx + 1}/{num_grupos}: {target_value[:50]}... ===", 'info')
                else:
                    self.log(f"\n=== NÚMERO {target_idx + 1}/{num_numeros}: +549{target_value} ===", 'info')
                
                # Por cada dispositivo (todas las líneas procesan la misma secuencia)
                for device in self.devices:
                    if self.should_stop: break
                    
                    # Por cada WhatsApp (Normal, Business, o Ambos)
                    for wa_idx, (wa_name, wa_package) in enumerate(whatsapp_apps):
                        if self.should_stop: break
                        
                        task_counter += 1
                        self.current_index = task_counter
                        self.root.after(0, self.update_stats)
                        
                        # Obtener mensaje rotativo
                        mensaje = self.manual_messages_numbers[mensaje_index % total_mensajes]
                        mensaje_index += 1
                        
                        # Construir link según tipo
                        if target_type == 'grupo':
                            target_link = target_value
                        else:
                            target_link = f"https://wa.me/549{target_value}"
                        
                        # Enviar usando la función auxiliar
                        success = self._send_to_target_with_whatsapp(
                            device, target_link, wa_name, wa_package, mensaje, task_counter
                        )
                        
                        # Pausa entre WhatsApps si hay más de uno y es Business (primero)
                        if success and len(whatsapp_apps) > 1 and wa_idx == 0:
                            wait_between = self.wait_between_messages.get()
                            if wait_between > 0:
                                self.log(f"Esperando {wait_between}s antes del siguiente WhatsApp...", 'info')
                                elapsed = 0
                                while elapsed < wait_between and not self.should_stop:
                                    while self.is_paused and not self.should_stop: time.sleep(0.1)
                                    if self.should_stop: break
                                    time.sleep(0.1)
                                    elapsed += 0.1
                        
                        time.sleep(0.5)  # Pequeña pausa entre envíos
                
                if self.should_stop: break
                self.log(f"\n=== {tipo_str} {target_idx + 1} completado ===", 'success')
            
            if self.should_stop: break
            self.log(f"\n--- CICLO {ciclo + 1} completado ---", 'success')
        
        self.log(f"\nModo Mixto (Variante {variant}) finalizado", 'success')

    def run_uno_a_uno_thread(self):
        """
        Lógica de envío para MODO NÚMEROS - UNO A UNO.
        Las líneas detectadas se emparejan aleatoriamente y conversan entre sí.
        Se repite según la configuración de Bucles y Ciclos.
        """
        self.log("Iniciando MODO NÚMEROS (Uno a uno)...", 'info')
        if len(self.detected_phone_lines) < 2:
            self.log("Se necesitan al menos 2 líneas de WhatsApp para este modo.", 'error')
            messagebox.showerror("Error", "Se necesitan al menos 2 líneas de WhatsApp detectadas para el modo 'Uno a uno'.", parent=self.root)
            return

        lines = self.detected_phone_lines.copy()
        num_lines = len(lines)
        num_bucles = self.manual_loops_var.get()
        num_ciclos = self.manual_cycles_var.get() # NUEVO
        
        task_counter = 0
        mensaje_index = self.mensaje_start_index
        total_mensajes_lib = len(self.manual_messages_numbers)

        # Bucle externo para los Bucles
        for bucle_num in range(num_bucles):
            if self.should_stop: break
            self.log(f"\n--- INICIANDO BUCLE {bucle_num + 1}/{num_bucles} ---", 'success')

            # Bucle interno para los Ciclos
            for ciclo_num in range(num_ciclos):
                if self.should_stop: break
                self.log(f"\n--- Ciclo {ciclo_num + 1}/{num_ciclos} (dentro del Bucle {bucle_num + 1}) ---", 'info')

                random.shuffle(lines)

                # Manejar número impar de líneas
                excluded = None
                current_lines = lines.copy()
                if num_lines % 2 != 0:
                    excluded = current_lines.pop()
                    self.log(f"Número impar de líneas. {excluded['number']} quedará fuera de este ciclo.", 'warning')

                # Crear parejas
                pairs = []
                for i in range(0, len(current_lines), 2):
                    pairs.append((current_lines[i], current_lines[i+1]))

                self.log(f"Se formaron {len(pairs)} parejas para este ciclo.", 'info')

                for i, (line_a, line_b) in enumerate(pairs):
                    if self.should_stop: break
                    self.log(f"\n=== Pareja {i+1}/{len(pairs)}: {line_a['number']} <-> {line_b['number']} ===", 'info')

                    # Conversación de ida y vuelta
                    # A -> B
                    task_counter += 1
                    mensaje_ida = self.manual_messages_numbers[mensaje_index % total_mensajes_lib]; mensaje_index += 1
                    link_ida = f"https://wa.me/{line_b['number'].replace('+', '')}?text={urllib.parse.quote(mensaje_ida, safe='')}"
                    self.run_single_task(
                        line_a['device'],
                        link_ida,
                        None,
                        task_counter,
                        whatsapp_package=line_a['package'],
                        activity_info={'from': line_a['number'], 'to': line_b['number']}
                    )
                    if self.should_stop: break

                    # B -> A
                    task_counter += 1
                    mensaje_vuelta = self.manual_messages_numbers[mensaje_index % total_mensajes_lib]; mensaje_index += 1
                    link_vuelta = f"https://wa.me/{line_a['number'].replace('+', '')}?text={urllib.parse.quote(mensaje_vuelta, safe='')}"
                    self.run_single_task(
                        line_b['device'],
                        link_vuelta,
                        None,
                        task_counter,
                        whatsapp_package=line_b['package'],
                        activity_info={'from': line_b['number'], 'to': line_a['number']}
                    )

            if self.should_stop: break
            self.log(f"\n--- FIN BUCLE {bucle_num + 1}/{num_bucles} ---", 'success')

    def run_uno_a_muchos_thread(self):
        """
        Lógica de envío para MODO NÚMEROS - UNO A MUCHOS.
        Cada línea detectada conversa con todas las demás, una sola vez por par.
        """
        self.log("Iniciando MODO NÚMEROS (Uno a muchos)...", 'info')
        if len(self.detected_phone_lines) < 2:
            self.log("Se necesitan al menos 2 líneas de WhatsApp para este modo.", 'error')
            messagebox.showerror("Error", "Se necesitan al menos 2 líneas de WhatsApp detectadas para el modo 'Uno a muchos'.", parent=self.root)
            return

        lines = self.detected_phone_lines.copy()
        num_lines = len(lines)
        num_bucles = self.manual_loops_var.get()

        task_counter = 0
        mensaje_index = self.mensaje_start_index
        total_mensajes_lib = len(self.manual_messages_numbers)

        for bucle_num in range(num_bucles):
            if self.should_stop: break
            self.log(f"\n--- BUCLE {bucle_num + 1}/{num_bucles} ---", 'info')

            # Crear todas las parejas únicas
            pairs = []
            for i in range(num_lines):
                for j in range(i + 1, num_lines):
                    pairs.append((lines[i], lines[j]))

            self.log(f"Se formaron {len(pairs)} parejas únicas para este bucle.", 'info')

            for i, (line_a, line_b) in enumerate(pairs):
                if self.should_stop: break
                self.log(f"\n=== Conversación {i+1}/{len(pairs)}: {line_a['number']} <-> {line_b['number']} ===", 'info')

                # Conversación de ida y vuelta
                # A -> B
                task_counter += 1
                mensaje_ida = self.manual_messages_numbers[mensaje_index % total_mensajes_lib]; mensaje_index += 1
                link_ida = f"https://wa.me/{line_b['number'].replace('+', '')}?text={urllib.parse.quote(mensaje_ida, safe='')}"
                self.run_single_task(
                    line_a['device'],
                    link_ida,
                    None,
                    task_counter,
                    whatsapp_package=line_a['package'],
                    activity_info={'from': line_a['number'], 'to': line_b['number']}
                )
                if self.should_stop: break

                # B -> A
                task_counter += 1
                mensaje_vuelta = self.manual_messages_numbers[mensaje_index % total_mensajes_lib]; mensaje_index += 1
                link_vuelta = f"https://wa.me/{line_a['number'].replace('+', '')}?text={urllib.parse.quote(mensaje_vuelta, safe='')}"
                self.run_single_task(
                    line_b['device'],
                    link_vuelta,
                    None,
                    task_counter,
                    whatsapp_package=line_b['package'],
                    activity_info={'from': line_b['number'], 'to': line_a['number']}
                )
    
    def _get_ordered_workers(self, active_devices):
        """
        Genera la lista de 'workers' (Device, Name, Package) en el orden solicitado:
        1. Todos los dispositivos en Business (si aplica)
        2. Todos los dispositivos en Normal (si aplica)
        """
        workers = []
        mode = self.whatsapp_mode.get()

        # Primero Business
        if mode in ["Business", "Ambas", "Todas"]:
            for device in active_devices:
                workers.append((device, "WhatsApp Business", "com.whatsapp.w4b"))

        # Luego Normal
        if mode in ["Normal", "Ambas", "Todas"]:
            for device in active_devices:
                workers.append((device, "WhatsApp Normal", "com.whatsapp"))

        return workers

    def run_grupos_simultaneous_thread(self):
        """
        Lógica de envío SIMULTÁNEO para MODO GRUPOS.
        Asigna grupos a workers y ejecuta en tandas paralelas respetando los dispositivos físicos.
        """
        # --- FILTRADO DE DISPOSITIVOS ---
        active_devices = self._get_filtered_devices()
        if not active_devices:
            self.log("No hay dispositivos que cumplan con el filtro seleccionado.", "error")
            self.root.after(0, lambda: messagebox.showerror("Error", "No se encontraron dispositivos que cumplan con el filtro de WhatsApp seleccionado.", parent=self.root))
            return

        num_devices_phys = len(active_devices)
        grupos = self.manual_inputs_groups
        num_grupos = len(grupos)
        num_bucles = self.manual_loops_var.get()

        if len(self.manual_messages_groups) < 1:
            self.log("Error: Modo Grupos requiere al menos 1 mensaje cargado.", "error")
            messagebox.showerror("Error", "Debes cargar al menos 1 archivo de mensajes.", parent=self.root)
            return

        # Construir lista de workers
        workers = self._get_ordered_workers(active_devices)
        num_workers = len(workers)

        self.log(f"Modo Grupos Simultáneo: {num_bucles} bucle(s), {num_grupos} grupos, {num_workers} workers", 'info')

        # Indices
        mensaje_index = self.mensaje_start_index
        total_mensajes_lib = len(self.manual_messages_groups)
        task_counter = 0

        # --- Bucle principal ---
        for bucle_num in range(num_bucles):
            if self.should_stop: break
            self.log(f"\n--- INICIANDO BUCLE {bucle_num + 1}/{num_bucles} ---", 'success')

            # Crear lista de tareas (Grupo, Worker)
            tasks = []
            for i, grupo in enumerate(grupos):
                # Aplicar rotación basada en el número de bucle para evitar repetición
                worker_idx = (i + bucle_num) % num_workers
                worker = workers[worker_idx]
                tasks.append((grupo, worker))

            # Procesar en tandas de tamaño = num_devices_phys
            # Como workers alterna dispositivos, tomar N tareas debería usar N dispositivos distintos (idealmente)
            batch_size = num_devices_phys

            for i in range(0, len(tasks), batch_size):
                if self.should_stop: break

                batch_tasks = tasks[i : i + batch_size]
                self.log(f"\n>>> Iniciando Tanda {i // batch_size + 1} ({len(batch_tasks)} tareas) <<<", 'info')

                threads = []
                for grupo_link, (device, wa_name, wa_package) in batch_tasks:
                    if self.should_stop: break
                    task_counter += 1

                    # Mensaje rotativo
                    mensaje = self.manual_messages_groups[mensaje_index % total_mensajes_lib]
                    mensaje_index += 1

                    def worker_func(dev, link, msg, t_idx, pkg, w_name):
                        self.log(f"[{dev}] -> Grupo {link.split('/')[-1][:10]} ({w_name})", 'info')
                        self.run_single_task(dev, link, msg, t_idx, whatsapp_package=pkg, skip_delay=True)

                        # Lógica cambio cuenta "Todas"
                        if w_name == "WhatsApp Normal" and self.whatsapp_mode.get() == "Todas":
                             self._switch_account_for_device(dev, delay=0.2)

                    t = threading.Thread(target=worker_func, args=(device, grupo_link, mensaje, task_counter, wa_package, wa_name))
                    threads.append(t)
                    t.start()

                for t in threads:
                    t.join()

                if self.should_stop: break
                self.log(f">>> Tanda finalizada", 'success')

                # Delay entre tandas
                if i + batch_size < len(tasks):
                    delay = random.uniform(self.fidelizado_send_delay_min.get(), self.fidelizado_send_delay_max.get())
                    self.log(f"⏳ Esperando {delay:.1f}s...", 'info')
                    self._controlled_sleep(delay)

            if self.should_stop: break
            self.log(f"\n--- FIN BUCLE {bucle_num + 1}/{num_bucles} ---", 'success')

        self.log(f"\nModo Grupos Simultáneo finalizado", 'success')

    def run_grupos_dual_whatsapp_thread(self):
        """
        Lógica de envío para MODO GRUPOS (Secuencial).
        Asigna grupos 1-a-1 a workers (Dispositivo + App) en orden round-robin.
        """
        # --- FILTRADO DE DISPOSITIVOS ---
        active_devices = self._get_filtered_devices()
        if not active_devices:
            self.log("No hay dispositivos que cumplan con el filtro seleccionado.", "error")
            self.root.after(0, lambda: messagebox.showerror("Error", "No se encontraron dispositivos que cumplan con el filtro de WhatsApp seleccionado.", parent=self.root))
            return
        # --- FIN FILTRADO ---

        num_devices = len(active_devices)
        grupos = self.manual_inputs_groups
        num_grupos = len(grupos)
        num_bucles = self.manual_loops_var.get()

        if len(self.manual_messages_groups) < 1:
            self.log("Error: Modo Grupos requiere al menos 1 mensaje cargado.", "error")
            messagebox.showerror("Error", "Debes cargar al menos 1 archivo de mensajes.", parent=self.root)
            return

        # Construir lista de workers
        workers = self._get_ordered_workers(active_devices)
        num_workers = len(workers)

        self.log(f"Modo Grupos Secuencial: {num_bucles} bucle(s), {num_grupos} grupos, {num_workers} workers ({num_devices} dispositivos)", 'info')
        self.log(f"Orden de asignación: {[f'{d}-{n}' for d,n,p in workers]}", 'info')

        # Indices
        mensaje_index = self.mensaje_start_index
        total_mensajes_lib = len(self.manual_messages_groups)
        task_counter = 0

        # --- Bucle principal ---
        for bucle_num in range(num_bucles):
            if self.should_stop: break
            self.log(f"\n--- INICIANDO BUCLE {bucle_num + 1}/{num_bucles} ---", 'success')

            # Iterar grupos y asignar round-robin
            for idx_grupo, grupo_link in enumerate(grupos):
                if self.should_stop: break

                # Determinar worker con rotación basada en bucle
                worker_idx = (idx_grupo + bucle_num) % num_workers
                worker = workers[worker_idx]
                device, wa_name, wa_package = worker

                grupo_display = grupo_link.split('chat.whatsapp.com/')[-1][:10] if 'chat.whatsapp.com' in grupo_link else str(idx_grupo+1)
                self.log(f"\n=== GRUPO {idx_grupo + 1}/{num_grupos} -> {device} ({wa_name}) ===", 'info')

                task_counter += 1

                # Obtener mensaje rotativo
                mensaje = self.manual_messages_groups[mensaje_index % total_mensajes_lib]
                mensaje_index += 1

                # Ejecutar envío
                success = self.run_single_task(
                    device, grupo_link, mensaje, task_counter, whatsapp_package=wa_package
                )

                # Lógica de cambio de cuenta para el modo "TODAS" (Solo si es Normal)
                if wa_name == "WhatsApp Normal" and self.whatsapp_mode.get() == "Todas":
                    self.log(f"[{device}] Cerrando WhatsApp Normal después de enviar...", 'info')
                    self._run_adb_command(['-s', device, 'shell', 'am', 'force-stop', 'com.whatsapp'], timeout=5)
                    time.sleep(1)

                    self.log(f"[{device}] Reabriendo WhatsApp Normal...", 'info')
                    self._run_adb_command(['-s', device, 'shell', 'am', 'start', '-n', 'com.whatsapp/.Main'], timeout=5)
                    time.sleep(3)

                    self.log(f"[{device}] Cambiando de cuenta...", 'info')
                    self._switch_account_for_device(device, delay=0.2)
                    time.sleep(4)

                    self.log(f"[{device}] Cerrando WhatsApp Normal después de cambiar cuenta...", 'info')
                    self._run_adb_command(['-s', device, 'shell', 'am', 'force-stop', 'com.whatsapp'], timeout=5)
                    time.sleep(1)

                    self.log(f"[{device}] Reabriendo WhatsApp Normal con nueva cuenta...", 'info')
                    self._run_adb_command(['-s', device, 'shell', 'am', 'start', '-n', 'com.whatsapp/.Main'], timeout=5)
                    time.sleep(2)

            if self.should_stop: break
            self.log(f"\n--- FIN BUCLE {bucle_num + 1}/{num_bucles} ---", 'success')

        self.log(f"\nModo Grupos finalizado", 'success')

    def run_unirse_grupos(self, grupos):
        """
        Función para unirse automáticamente a grupos.
        NUEVA LÓGICA CON THREADING (EJECUCIÓN PARALELA):
        Por cada grupo:
          - TODOS los dispositivos se unen SIMULTÁNEAMENTE según la selección de WhatsApp
        Proceso:
          - Presiona DPAD_DOWN 3 veces (con pausas de 2s)
          - Presiona ENTER
          - Presiona BACK para salir
        """
        try:
            self._enter_task_mode()
            num_devices = len(self.devices)
            num_grupos = len(grupos)

            # Obtener qué WhatsApp usar
            wa_mode = self.whatsapp_mode.get()

            # Determinar cuántas uniones totales habrá
            if wa_mode == "Todas":
                total_uniones = num_grupos * num_devices * 3
            elif wa_mode == "Ambas":
                total_uniones = num_grupos * num_devices * 2
            else:
                total_uniones = num_grupos * num_devices

            self.log(f"\n=== UNIRSE A GRUPOS (MODO PARALELO) ===", 'info')
            self.log(f"Grupos: {num_grupos}", 'info')
            self.log(f"Dispositivos: {num_devices}", 'info')
            self.log(f"WhatsApp: {wa_mode}", 'info')
            self.log(f"Total de uniones: {total_uniones}", 'info')

            total = num_grupos * num_devices * 2

            # Función auxiliar para unirse a un grupo en un dispositivo
            def unirse_a_grupo_device(device, grupo_link, whatsapp_package, whatsapp_name):
                """Ejecuta el proceso completo de unión para un dispositivo."""
                try:
                    if self.should_stop:
                        return False

                    # Verificar pausa
                    while self.is_paused and not self.should_stop:
                        time.sleep(0.1)
                    if self.should_stop:
                        return False

                    self.log(f"[{device}] Uniéndose por {whatsapp_name}...", 'info')

                    # Abrir grupo
                    open_args = ['-s', device, 'shell', 'am', 'start', '-a', 'android.intent.action.VIEW',
                                '-d', grupo_link, '-p', whatsapp_package]

                    if not self._run_adb_command(open_args, timeout=20):
                        self.log(f"[{device}] Fallo al abrir grupo en {whatsapp_name}", "error")
                        return False

                    # Esperar 1 segundo (reducido de 2 para acelerar)
                    time.sleep(1)

                    if self.should_stop:
                        return False

                    # Presionar DPAD_DOWN 3 veces
                    for i in range(3):
                        if self.should_stop:
                            return False
                        down_args = ['-s', device, 'shell', 'input', 'keyevent', 'KEYCODE_DPAD_DOWN']
                        self._run_adb_command(down_args, timeout=5)
                        time.sleep(1) # Reducido de 2s

                    if self.should_stop:
                        return False

                    # Presionar ENTER
                    enter_args = ['-s', device, 'shell', 'input', 'keyevent', 'KEYCODE_ENTER']
                    self._run_adb_command(enter_args, timeout=10)

                    # Esperar 1 segundo (reducido de 2s)
                    time.sleep(1)

                    # Presionar BACK para salir del grupo
                    back_args = ['-s', device, 'shell', 'input', 'keyevent', 'KEYCODE_BACK']
                    self._run_adb_command(back_args, timeout=10)
                    self.log(f"[{device}] Presionando BACK para salir...", 'info')

                    # Esperar 0.5 segundos final (reducido de 1s)
                    time.sleep(0.5)

                    self.log(f"[{device}] Unido a grupo por {whatsapp_name}", 'success')
                    return True

                except Exception as e:
                    self.log(f"[{device}] Error en unión: {e}", 'error')
                    return False
            
            # Por cada grupo
            for idx_grupo, grupo_link in enumerate(grupos):
                if self.should_stop:
                    break
                
                grupo_display = grupo_link[:50] + "..." if len(grupo_link) > 50 else grupo_link
                self.log(f"\n--- GRUPO {idx_grupo + 1}/{num_grupos}: {grupo_display} ---", 'info')
                
                # ===== FASE 1: WHATSAPP BUSINESS (si corresponde) =====
                if wa_mode == "Business" or wa_mode == "Ambas" or wa_mode == "Todas":
                    fase_num = 1 if (wa_mode == "Ambas" or wa_mode == "Todas") else 0
                    if fase_num == 1:
                        self.log(f"\n>>> FASE 1: Todos los dispositivos uniéndose SIMULTÁNEAMENTE por WhatsApp Business...", 'info')
                    else:
                        self.log(f"\n>>> Todos los dispositivos uniéndose SIMULTÁNEAMENTE por WhatsApp Business...", 'info')
                    
                    threads_business = []
                    for device in self.devices:
                        if self.should_stop:
                            break
                        thread = threading.Thread(
                            target=unirse_a_grupo_device,
                            args=(device, grupo_link, 'com.whatsapp.w4b', 'WhatsApp Business'),
                            daemon=True
                        )
                        threads_business.append(thread)
                        thread.start()

                    # Esperar a que TODOS los threads de Business terminen
                    for thread in threads_business:
                        thread.join()

                    if fase_num == 1:
                        self.log(f"\n>>> FASE 1 completada: Todos unidos por WhatsApp Business", 'success')
                    else:
                        self.log(f"\n>>> Completado: Todos unidos por WhatsApp Business", 'success')
                    
                    if self.should_stop:
                        break
                
                # ===== FASE 2: WHATSAPP NORMAL (si corresponde) =====
                if wa_mode == "Normal" or wa_mode == "Ambas" or wa_mode == "Todas":
                    fase_num = 2 if (wa_mode == "Ambas" or wa_mode == "Todas") else 0
                    if fase_num == 2:
                        self.log(f"\n>>> FASE 2: Todos los dispositivos uniéndose SIMULTÁNEAMENTE por WhatsApp Normal...", 'info')
                    else:
                        self.log(f"\n>>> Todos los dispositivos uniéndose SIMULTÁNEAMENTE por WhatsApp Normal...", 'info')

                    threads_normal = []
                    for device in self.devices:
                        if self.should_stop:
                            break
                        thread = threading.Thread(
                            target=unirse_a_grupo_device,
                            args=(device, grupo_link, 'com.whatsapp', 'WhatsApp Normal'),
                            daemon=True
                        )
                        threads_normal.append(thread)
                        thread.start()

                    # Esperar a que TODOS los threads de Normal terminen
                    for thread in threads_normal:
                        thread.join()

                    if fase_num == 2:
                        self.log(f"\n>>> FASE 2 completada: Todos unidos por WhatsApp Normal", 'success')
                    else:
                        self.log(f"\n>>> Completado: Todos unidos por WhatsApp Normal", 'success')

                    # Si el modo es "Todas", cambiar de cuenta DESPUÉS de unirse con Normal
                    if wa_mode == "Todas":
                        self.log(f"\n>>> Cambiando de cuenta en todos los dispositivos...", 'info')

                        for device in self.devices:
                            if self.should_stop:
                                break

                            self.log(f"[{device}] Cerrando WhatsApp Normal...", 'info')
                            close_cmd = ['-s', device, 'shell', 'am', 'force-stop', 'com.whatsapp']
                            self._run_adb_command(close_cmd, timeout=5)
                            time.sleep(1)

                            self.log(f"[{device}] Reabriendo WhatsApp Normal...", 'info')
                            open_cmd = ['-s', device, 'shell', 'am', 'start', '-n', 'com.whatsapp/.Main']
                            self._run_adb_command(open_cmd, timeout=5)
                            time.sleep(3)  # Esperar 3 segundos para que WhatsApp se abra completamente

                            self.log(f"[{device}] Cambiando de cuenta...", 'info')
                            self._switch_account_for_device(device)
                            time.sleep(1)

                            self.log(f"[{device}] Cerrando WhatsApp Normal después de cambiar cuenta...", 'info')
                            close_cmd = ['-s', device, 'shell', 'am', 'force-stop', 'com.whatsapp']
                            self._run_adb_command(close_cmd, timeout=5)
                            time.sleep(1)

                            self.log(f"[{device}] Reabriendo WhatsApp Normal con nueva cuenta...", 'info')
                            open_cmd = ['-s', device, 'shell', 'am', 'start', '-n', 'com.whatsapp/.Main']
                            self._run_adb_command(open_cmd, timeout=5)
                            time.sleep(2)

                        if self.should_stop:
                            break
                
                # ===== FASE 3: WHATSAPP NORMAL 2 (si corresponde) =====
                if wa_mode == "Todas":
                    self.log(f"\n>>> FASE 3: Todos los dispositivos uniéndose SIMULTÁNEAMENTE por WhatsApp Normal (cuenta 2)...", 'info')

                    threads_normal2 = []
                    for device in self.devices:
                        if self.should_stop:
                            break
                        thread = threading.Thread(
                            target=unirse_a_grupo_device,
                            args=(device, grupo_link, 'com.whatsapp', 'WhatsApp Normal (cuenta 2)'),
                            daemon=True
                        )
                        threads_normal2.append(thread)
                        thread.start()

                    # Esperar a que TODOS los threads de Normal 2 terminen
                    for thread in threads_normal2:
                        thread.join()

                    self.log(f"\n>>> FASE 3 completada: Todos unidos por WhatsApp Normal (cuenta 2)", 'success')
                
                self.log(f"\n=== GRUPO {idx_grupo + 1} completado ===", 'success')
            
            self.log(f"\n=== PROCESO DE UNIÓN A GRUPOS FINALIZADO ===", 'success')
            messagebox.showinfo("Éxito", f"Proceso completado.\n\nSe unieron a {num_grupos} grupo(s) con {num_devices} dispositivo(s).", parent=self.root)
        finally:
            self._finalize_sending()
    
    def _write_message_with_keyevents(self, device, message):
        """
        Escribe un mensaje carácter por carácter usando input text de ADB.
        Retorna True si tuvo éxito, False si falló.
        """
        try:
            if self.should_stop:
                return False
            
            while self.is_paused and not self.should_stop:
                time.sleep(0.1)
            if self.should_stop:
                return False

            # Escribir carácter por carácter
            for char in message:
                if self.should_stop:
                    return False
                
                while self.is_paused and not self.should_stop:
                    time.sleep(0.1)
                if self.should_stop:
                    return False
                
                # Escapar el carácter individual
                if char == ' ':
                    char_escaped = '%s'
                elif char in ['\\', '"', "'", '$', '`', '!', '&', '|', ';', '<', '>', 
                             '(', ')', '[', ']', '{', '}', '*', '?', '#', '~']:
                    char_escaped = f'\\{char}'
                else:
                    char_escaped = char
                
                # Enviar el carácter
                text_args = ['-s', device, 'shell', 'input', 'text', char_escaped]

                if not self._run_adb_command(text_args, timeout=5):
                    # Si falla un carácter, intentar continuar
                    self.log(f"Advertencia: fallo al escribir '{char}'", "warning")

            return True

        except Exception as e:
            self.log(f"Error al escribir mensaje: {e}", 'error')
            return False
    
    def _perform_call(self, device, number, duration):
        """Realiza una llamada telefónica normal, espera y cuelga."""
        if self.should_stop: return

        try:
            # 1. Iniciar llamada
            cmd_call = ['-s', device, 'shell', 'am', 'start', '-a', 'android.intent.action.CALL', '-d', f'tel:{number}']
            self._run_adb_command(cmd_call, timeout=10)

            # 2. Esperar duración configurada
            self.log(f"[{device}] Llamada en curso ({duration}s)...", 'info')
            self._controlled_sleep(duration)

            if self.should_stop: return

            # 3. Colgar
            self.log(f"[{device}] Colgando...", 'info')
            self._run_adb_command(['-s', device, 'shell', 'input', 'keyevent', 'KEYCODE_ENDCALL'], timeout=5)

            # 4. Cerrar ventana / Ir a inicio (para limpiar)
            time.sleep(1)
            self._run_adb_command(['-s', device, 'shell', 'input', 'keyevent', 'KEYCODE_HOME'], timeout=5)

        except Exception as e:
            self.log(f"[{device}] Excepción en _perform_call: {e}", 'error')

    def run_bucle_blast_thread_V2(self):
        """
        Lógica de envío NUEVA (Modo Bucles G/N Blast V2).
        Definición de 1 Bucle: Recorrer TODA la lista de G y N.
        Repetir esto N veces.
        """
        num_devices = len(self.devices)
        num_bucles = self.manual_loops
        
        # Crear copias de las listas de tareas
        group_targets = list(self.manual_inputs_groups)
        number_targets = list(self.manual_inputs_numbers)
        group_messages = list(self.manual_messages_groups)
        number_messages = list(self.manual_messages_numbers)

        # Índices para rotación de mensajes
        g_msg_idx = 0
        n_msg_idx = 0
        
        # Contador global de tareas (1-based)
        task_counter = 0

        if not group_targets or not number_targets or not group_messages or not number_messages:
            self.log("Error: Modo Bucles Blast requiere Grupos, Números y sus respectivos Mensajes.", "error")
            return
            
        # Longitud del "sub-bucle" (la lista de targets más larga)
        max_len = max(len(group_targets), len(number_targets))

        # --- Funciones helper para rotar mensajes ---
        def get_next_g_msg():
            nonlocal g_msg_idx
            msg = group_messages[g_msg_idx % len(group_messages)]
            g_msg_idx += 1
            return msg

        def get_next_n_msg():
            nonlocal n_msg_idx
            msg = number_messages[n_msg_idx % len(number_messages)]
            n_msg_idx += 1
            return msg
        # --- Fin helpers ---

        # --- Bucle Principal (N Repeticiones) ---
        for b in range(num_bucles):
            if self.should_stop: break
            rep_num = b + 1
            self.log(f"--- Iniciando REPETICIÓN {rep_num} / {num_bucles} ---", 'info')

            # --- Bucle Interno (Recorrer todos los targets) ---
            for i in range(max_len):
                if self.should_stop: break
                
                # --- Etapa 1: Blast de Grupo ---
                # Target rotativo
                current_group_target = group_targets[i % len(group_targets)]
                self.log(f"Repetición {rep_num}, Etapa {i+1}.1: Todos a GRUPO {current_group_target[:40]}...", 'info')
                
                for device in self.devices:
                    if self.should_stop: break
                    task_counter += 1
                    msg_g = get_next_g_msg() # Mensaje rotativo
                    self.run_single_task(device, current_group_target, msg_g, task_counter)
                
                if self.should_stop: break # Check entre etapas

                # --- Etapa 2: Blast de Número ---
                # Target rotativo
                current_number_target = number_targets[i % len(number_targets)]
                self.log(f"Repetición {rep_num}, Etapa {i+1}.2: Todos a NÚMERO {current_number_target}", 'info')
                
                for device in self.devices:
                    if self.should_stop: break
                    task_counter += 1
                    msg_n = get_next_n_msg() # Mensaje rotativo
                    wa_link = f"https://wa.me/549{current_number_target}?text={urllib.parse.quote(msg_n, safe='')}"
                    self.run_single_task(device, wa_link, None, task_counter) # None pork el msg va en el link
            
            if self.should_stop: break # Check al final del bucle interno

            self.log(f"--- Fin REPETICIÓN {rep_num} ---", 'info')
    

    def _finalize_sending(self):
        """Reestablece la UI al finalizar o cancelar el envío."""
        self._unblock_ui_from_task()  # Desbloquear UI
        self.is_running = False

        if self.sms_mode_active:
            # -- Vista SMS --
            if hasattr(self, 'sms_btn_start'):
                self.sms_btn_start.configure(state=tk.NORMAL)
            if hasattr(self, 'sms_btn_load'):
                self.sms_btn_load.configure(state=tk.NORMAL)
            if hasattr(self, 'sms_btn_pause'):
                self.sms_btn_pause.configure(state=tk.DISABLED, text="⏸  PAUSAR")
            if hasattr(self, 'sms_btn_stop'):
                self.sms_btn_stop.configure(state=tk.DISABLED)
            if hasattr(self, 'sms_back_btn'):
                self.sms_back_btn.configure(state=tk.NORMAL)
        else:
            # -- Vista Tradicional --
            if hasattr(self, 'btn_start'):
                self.btn_start.configure(state=tk.NORMAL)
            if hasattr(self, 'btn_load'):
                self.btn_load.configure(state=tk.NORMAL)
            if hasattr(self, 'fidelizado_unlock_btn') and self.fidelizado_unlock_btn:
                self.fidelizado_unlock_btn.configure(state=tk.NORMAL)
            if hasattr(self, 'btn_pause'):
                self.btn_pause.configure(state=tk.DISABLED, text="⏸  PAUSAR")
            if hasattr(self, 'btn_stop'):
                self.btn_stop.configure(state=tk.DISABLED)

            # -- Vista Llamadas --
            if hasattr(self, 'calls_btn_start'):
                self.calls_btn_start.configure(state=tk.NORMAL)
                if hasattr(self, 'calls_btn_load'):
                    self.calls_btn_load.configure(state=tk.NORMAL)
                self.calls_btn_pause.configure(state=tk.DISABLED, text="⏸  PAUSAR")
                self.calls_btn_stop.configure(state=tk.DISABLED)
                if hasattr(self, 'calls_back_btn'):
                    self.calls_back_btn.configure(state=tk.NORMAL)

            # -- Vista Fidelizado --
            if hasattr(self, 'fidelizado_btn_start'):
                # Ocultar botones de control, mostrar los de inicio
                self.control_buttons_frame.pack_forget()
                self.actions_frame.pack(fill=tk.X, pady=(5, 0))

                # Configurar estado de botones
                self.fidelizado_btn_start.configure(state=tk.NORMAL)
                self.unirse_grupos_btn.configure(state=tk.NORMAL)
                self.fidelizado_btn_pause.configure(state=tk.DISABLED, text="⏸  PAUSAR")
                self.fidelizado_btn_stop.configure(state=tk.DISABLED)
                if hasattr(self, 'back_to_traditional_btn'):
                    self.back_to_traditional_btn.configure(state=tk.NORMAL)

    def _enter_task_mode(self):
        """Configura la UI para un estado de 'tarea en ejecución'."""
        self.is_running = True
        self.is_paused = False
        self.should_stop = False
        self.sent_count = 0
        self.failed_count = 0
        self.current_index = 0
        self.start_time = datetime.now()

        # Actualizar UI
        if self.sms_mode_active:
            if hasattr(self, 'sms_btn_start'):
                self.sms_btn_start.configure(state=tk.DISABLED)
            if hasattr(self, 'sms_btn_load'):
                self.sms_btn_load.configure(state=tk.DISABLED)
            if hasattr(self, 'sms_btn_pause'):
                self.sms_btn_pause.configure(state=tk.NORMAL)
            if hasattr(self, 'sms_btn_stop'):
                self.sms_btn_stop.configure(state=tk.NORMAL)
            if hasattr(self, 'sms_back_btn'):
                self.sms_back_btn.configure(state=tk.DISABLED)
        else:
            # -- Vista Tradicional --
            if hasattr(self, 'btn_start'):
                self.btn_start.configure(state=tk.DISABLED)
            if hasattr(self, 'btn_load'):
                self.btn_load.configure(state=tk.DISABLED)
            if hasattr(self, 'fidelizado_unlock_btn') and self.fidelizado_unlock_btn:
                self.fidelizado_unlock_btn.configure(state=tk.DISABLED)
            if hasattr(self, 'btn_pause'):
                self.btn_pause.configure(state=tk.NORMAL)
            if hasattr(self, 'btn_stop'):
                self.btn_stop.configure(state=tk.NORMAL)

            # -- Vista Llamadas --
            if hasattr(self, 'calls_btn_start'):
                self.calls_btn_start.configure(state=tk.DISABLED)
                if hasattr(self, 'calls_btn_load'):
                    self.calls_btn_load.configure(state=tk.DISABLED)
                self.calls_btn_pause.configure(state=tk.NORMAL)
                self.calls_btn_stop.configure(state=tk.NORMAL)
                if hasattr(self, 'calls_back_btn'):
                    self.calls_back_btn.configure(state=tk.DISABLED)

            # -- Vista Fidelizado --
            if hasattr(self, 'fidelizado_btn_start'):
                # Ocultar botones de inicio, mostrar los de control
                self.actions_frame.pack_forget()
                self.control_buttons_frame.pack(fill=tk.X, pady=(5, 0))

                # Configurar estado de botones
                self.fidelizado_btn_start.configure(state=tk.DISABLED)
                self.unirse_grupos_btn.configure(state=tk.DISABLED)
                self.fidelizado_btn_pause.configure(state=tk.NORMAL)
                self.fidelizado_btn_stop.configure(state=tk.NORMAL)
                if hasattr(self, 'back_to_traditional_btn'):
                    self.back_to_traditional_btn.configure(state=tk.DISABLED)

        self._block_ui_for_task()  # Bloquear interacciones

    def _recursive_set_blocking(self, widget, block=True, skip_widgets=None):
        """
        Recorre recursivamente los widgets y deshabilita la interacción
        preservando la apariencia visual original.
        """
        if skip_widgets and widget in skip_widgets:
            return

        # Procesar el widget actual
        if block:
            # Verificar si el widget soporta 'state'
            try:
                # Solo bloquear si no está ya deshabilitado (o si ya lo procesamos)
                # Usamos el ID del widget como clave única
                w_id = id(widget)

                if w_id not in self._blocked_widgets_state:
                    # Guardar estado original si tiene 'state'
                    current_config = {}
                    has_state = False

                    try:
                        current_state = widget.cget("state")
                        has_state = True
                        current_config["state"] = current_state
                    except:
                        pass

                    if has_state:
                        # Guardar colores para mantener apariencia
                        for attr in ["fg_color", "text_color", "border_color", "button_color",
                                     "text_color_disabled", "fg_color_disabled", "border_color_disabled",
                                     "button_color_disabled", "segment_color", "unselected_color",
                                     "selected_color", "unselected_hover_color", "selected_hover_color",
                                     "progress_color"]:
                            try:
                                val = widget.cget(attr)
                                current_config[attr] = val
                            except:
                                pass

                        self._blocked_widgets_state[w_id] = current_config

                        # Configurar para deshabilitar pero manteniendo colores
                        new_config = {"state": "disabled"}

                        # Mapeo de colores normales a disabled para CustomTkinter
                        if "fg_color" in current_config:
                            new_config["fg_color_disabled"] = current_config["fg_color"]
                        if "text_color" in current_config:
                            new_config["text_color_disabled"] = current_config["text_color"]
                        if "border_color" in current_config:
                            new_config["border_color_disabled"] = current_config["border_color"]
                        if "button_color" in current_config: # Checkbox, Switch, OptionMenu
                            new_config["button_color_disabled"] = current_config["button_color"]

                        # Aplicar cambios
                        widget.configure(**new_config)
            except Exception as e:
                # Algunos widgets pueden no soportar cget/configure de la manera estándar
                pass
        else:
            # Restaurar estado
            w_id = id(widget)
            if w_id in self._blocked_widgets_state:
                original_config = self._blocked_widgets_state.pop(w_id)
                try:
                    # Restaurar colores disabled a sus valores originales (si existían) o default
                    # Y restaurar el estado
                    restore_config = {}
                    if "state" in original_config:
                        restore_config["state"] = original_config["state"]

                    for attr in ["text_color_disabled", "fg_color_disabled", "border_color_disabled", "button_color_disabled"]:
                        if attr in original_config:
                            restore_config[attr] = original_config[attr]
                        # Si no estaba guardado, no lo tocamos (se queda con el valor modificado o default)
                        # Para ser precisos, CustomTkinter no tiene "default" fácil de recuperar sin reiniciar,
                        # pero al restaurar el estado a "normal", los colores disabled dejan de usarse.

                    widget.configure(**restore_config)
                except:
                    pass

        # Recorrer hijos
        for child in widget.winfo_children():
            self._recursive_set_blocking(child, block, skip_widgets)

    def _block_ui_for_task(self):
        """Bloquea la interacción en la interfaz excepto controles permitidos."""
        skip = []

        # Controles globales a mantener
        if hasattr(self, 'btn_pause'): skip.append(self.btn_pause)
        if hasattr(self, 'btn_stop'): skip.append(self.btn_stop)
        if hasattr(self, 'header_logo_left'): skip.append(self.header_logo_left)
        if hasattr(self, 'header_logo_right'): skip.append(self.header_logo_right)
        if hasattr(self, 'header_title_label'): skip.append(self.header_title_label)
        if hasattr(self, 'fidelizado_btn_pause'): skip.append(self.fidelizado_btn_pause)
        if hasattr(self, 'fidelizado_btn_stop'): skip.append(self.fidelizado_btn_stop)
        if hasattr(self, 'control_buttons_frame'): skip.append(self.control_buttons_frame)
        if hasattr(self, 'sms_btn_pause'): skip.append(self.sms_btn_pause)
        if hasattr(self, 'sms_btn_stop'): skip.append(self.sms_btn_stop)
        if hasattr(self, 'calls_btn_pause'): skip.append(self.calls_btn_pause)
        if hasattr(self, 'calls_btn_stop'): skip.append(self.calls_btn_stop)

        # El panel derecho (Log/Tabla) NO se bloquea

        # Áreas a bloquear
        areas_to_block = []

        # 1. Header (Botón atrás, etc.)
        if hasattr(self, 'header_frame'):
            areas_to_block.append(self.header_frame)

        # 2. Vistas (Settings, Inputs)
        # Tradicional
        if hasattr(self, 'traditional_view_frame') and self.traditional_view_frame and self.traditional_view_frame.winfo_ismapped():
            areas_to_block.append(self.traditional_view_frame)
            # Excluir explícitamente controles si están dentro (ya están en skip, pero por si acaso)

        # Fidelizado
        if hasattr(self, 'fidelizado_view_frame') and self.fidelizado_view_frame and self.fidelizado_view_frame.winfo_ismapped():
            areas_to_block.append(self.fidelizado_view_frame)

        # SMS
        if hasattr(self, 'sms_view_frame') and self.sms_view_frame and self.sms_view_frame.winfo_ismapped():
            areas_to_block.append(self.sms_view_frame)

        # Llamadas
        if hasattr(self, 'calls_view_frame') and self.calls_view_frame and self.calls_view_frame.winfo_ismapped():
            areas_to_block.append(self.calls_view_frame)

        for area in areas_to_block:
            self._recursive_set_blocking(area, block=True, skip_widgets=skip)

    def _unblock_ui_from_task(self):
        """Restaura la interacción en la interfaz."""
        areas_to_unblock = []
        if hasattr(self, 'header_frame'): areas_to_unblock.append(self.header_frame)
        if hasattr(self, 'traditional_view_frame') and self.traditional_view_frame is not None:
            areas_to_unblock.append(self.traditional_view_frame)
        if hasattr(self, 'fidelizado_view_frame') and self.fidelizado_view_frame is not None:
            areas_to_unblock.append(self.fidelizado_view_frame)
        if hasattr(self, 'sms_view_frame') and self.sms_view_frame is not None:
            areas_to_unblock.append(self.sms_view_frame)
        if hasattr(self, 'calls_view_frame') and self.calls_view_frame is not None:
            areas_to_unblock.append(self.calls_view_frame)

        for area in areas_to_unblock:
            self._recursive_set_blocking(area, block=False)

        self._blocked_widgets_state.clear()

    def _update_device_labels(self):
        """Actualiza todas las etiquetas de la UI que muestran la lista de dispositivos."""
        if hasattr(self, 'fidelizado_device_list_label'):
            if self.devices:
                count = len(self.devices)
                device_text = f"Teléfonos - {count}"
                self.fidelizado_device_list_label.configure(text=device_text)
            else:
                self.fidelizado_device_list_label.configure(text="Teléfonos - 0")

    # --- Programación de Envíos ---
    def _open_schedule_dialog(self, callback):
        """Abre un diálogo para elegir entre enviar ahora o programar."""
        dialog = ctk.CTkToplevel(self.root)
        dialog.title("Iniciar Envío")
        dialog.geometry("400x200")
        dialog.transient(self.root)
        dialog.grab_set()

        self._center_toplevel(dialog, 400, 200)

        label = ctk.CTkLabel(dialog, text="¿Cómo deseas iniciar el proceso?", font=self.fonts['card_title'], text_color=self.colors['text'])
        label.pack(pady=20)

        btn_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        btn_frame.pack(fill=tk.X, padx=20, pady=10)

        def send_now():
            dialog.destroy()
            callback(True) # Confirm=True (Comportamiento normal)

        def open_scheduler():
            dialog.destroy()
            self._show_scheduler_window(callback)

        btn_now = ctk.CTkButton(btn_frame, text="Enviar Ahora", command=send_now, fg_color=self.colors['action_start'], width=150)
        btn_now.pack(side=tk.LEFT, padx=10)

        btn_schedule = ctk.CTkButton(btn_frame, text="Programar Envío", command=open_scheduler, fg_color=self.colors['blue'], width=150)
        btn_schedule.pack(side=tk.RIGHT, padx=10)

    def _show_scheduler_window(self, callback):
        """Muestra la ventana para elegir fecha y hora."""
        window = ctk.CTkToplevel(self.root)
        window.title("Programar Envío")
        window.geometry("350x300")
        window.transient(self.root)
        window.grab_set()
        self._center_toplevel(window, 350, 300)

        now = datetime.now()
        default_time = now + timedelta(minutes=5)

        ctk.CTkLabel(window, text="Configurar Fecha y Hora", font=self.fonts['card_title']).pack(pady=15)

        # Date
        date_frame = ctk.CTkFrame(window, fg_color="transparent")
        date_frame.pack(pady=5)
        ctk.CTkLabel(date_frame, text="Fecha (DD/MM/AAAA):", width=150).pack(side=tk.LEFT)
        date_entry = ctk.CTkEntry(date_frame, width=120)
        date_entry.pack(side=tk.LEFT)
        date_entry.insert(0, default_time.strftime("%d/%m/%Y"))

        # Time
        time_frame = ctk.CTkFrame(window, fg_color="transparent")
        time_frame.pack(pady=5)
        ctk.CTkLabel(time_frame, text="Hora (HH:MM):", width=150).pack(side=tk.LEFT)
        time_entry = ctk.CTkEntry(time_frame, width=120)
        time_entry.pack(side=tk.LEFT)
        time_entry.insert(0, default_time.strftime("%H:%M"))

        def confirm_schedule():
            date_str = date_entry.get().strip()
            time_str = time_entry.get().strip()
            try:
                dt = datetime.strptime(f"{date_str} {time_str}", "%d/%m/%Y %H:%M")
                delay = (dt - datetime.now()).total_seconds()

                if delay < 0:
                    messagebox.showerror("Error", "La fecha/hora debe ser futura.", parent=window)
                    return

                window.destroy()
                self._start_scheduled_wait(dt, callback)

            except ValueError:
                messagebox.showerror("Error", "Formato inválido. Use DD/MM/AAAA y HH:MM", parent=window)

        ctk.CTkButton(window, text="Confirmar Programación", command=confirm_schedule, fg_color=self.colors['action_start']).pack(pady=20)

    def _start_scheduled_wait(self, target_time, callback):
        """Bloquea la UI y espera a la hora programada."""
        wait_window = ctk.CTkToplevel(self.root)
        wait_window.title("Esperando...")
        wait_window.geometry("400x250")
        wait_window.transient(self.root)
        wait_window.grab_set()
        # Prevent closing with X
        wait_window.protocol("WM_DELETE_WINDOW", lambda: None)
        self._center_toplevel(wait_window, 400, 250)

        ctk.CTkLabel(wait_window, text="⏳ Envío Programado", font=self.fonts['header']).pack(pady=20)
        lbl_target = ctk.CTkLabel(wait_window, text=f"Se iniciará el: {target_time.strftime('%d/%m/%Y %H:%M')}", font=self.fonts['card_title'])
        lbl_target.pack(pady=5)

        lbl_countdown = ctk.CTkLabel(wait_window, text="Calculando...", font=('Inter', 20, 'bold'), text_color=self.colors['orange'])
        lbl_countdown.pack(pady=20)

        cancel_flag = [False]

        def cancel():
            cancel_flag[0] = True
            wait_window.destroy()
            self.log("Programación cancelada por el usuario.", 'warning')

        ctk.CTkButton(wait_window, text="Cancelar Programación", command=cancel, fg_color=self.colors['action_cancel']).pack(pady=10)

        def update_countdown():
            if cancel_flag[0]: return

            now = datetime.now()
            remaining = target_time - now

            if remaining.total_seconds() <= 0:
                wait_window.destroy()
                self._execute_scheduled_task(callback)
            else:
                # Format HH:MM:SS
                s = int(remaining.total_seconds())
                h, r = divmod(s, 3600)
                m, sec = divmod(r, 60)
                lbl_countdown.configure(text=f"Faltan: {h:02d}:{m:02d}:{sec:02d}")
                self.root.after(1000, update_countdown)

        update_countdown()

    def _execute_scheduled_task(self, callback):
        """Ejecuta la tarea: Detectar dispositivos -> callback(confirm=False)."""
        self.log("⏰ Hora programada alcanzada. Iniciando...", 'info')

        # 1. Detectar Dispositivos (Force Wakeup)
        self.log("Detectando y despertando dispositivos...", 'info')
        self.detect_devices(silent=True) # Modificado para aceptar silent

        if not self.devices:
            self.log("Error: No se detectaron dispositivos al momento de iniciar.", 'error')
            messagebox.showerror("Error Programado", "Fallo al iniciar: No se detectaron dispositivos.")
            return

        # 2. Ejecutar (sin confirmación)
        callback(False)

    # --- ################################################################## ---
    # --- send_msg (MODIFICADO para loguear device)
    # --- ################################################################## ---
    
    def _switch_whatsapp_account(self, device):
        """
        Cambia de cuenta en WhatsApp Normal.
        
        Args:
            device: ID del dispositivo
        """
        self.log(f"[{device}] Cambiando de cuenta en WhatsApp Normal...", 'info')
        
        # 1) Cerrar todo
        close_commands = [
            ['-s', device, 'shell', 'am', 'force-stop', 'com.android.settings'],
            ['-s', device, 'shell', 'am', 'force-stop', 'com.whatsapp'],
            ['-s', device, 'shell', 'am', 'force-stop', 'com.whatsapp.w4b']
        ]
        for cmd in close_commands:
            self._run_adb_command(cmd, timeout=3)
        
        # 2) Abrir WhatsApp y cambiar de cuenta
        self._run_adb_command(['-s', device, 'shell', 'am', 'start', '-n', 'com.whatsapp/.Main'], timeout=10)
        time.sleep(3)  # Esperar a que abra
        
        # Navegar al menú de cambio de cuenta
        for _ in range(2):
            self._run_adb_command(['-s', device, 'shell', 'input', 'keyevent', 'KEYCODE_DPAD_UP'], timeout=3)
            time.sleep(0.2)
        
        self._run_adb_command(['-s', device, 'shell', 'input', 'keyevent', 'KEYCODE_DPAD_RIGHT'], timeout=3)
        time.sleep(0.2)
        self._run_adb_command(['-s', device, 'shell', 'input', 'keyevent', 'KEYCODE_ENTER'], timeout=3)
        time.sleep(0.2)
        
        for _ in range(7):
            self._run_adb_command(['-s', device, 'shell', 'input', 'keyevent', 'KEYCODE_TAB'], timeout=3)
            time.sleep(0.05)  # Más rápido: 0.05s entre TABs
        
        self._run_adb_command(['-s', device, 'shell', 'input', 'keyevent', 'KEYCODE_ENTER'], timeout=3)
        
        # Esperar 3 segundos con WhatsApp abierto para que carguen los mensajes
        self.log(f"[{device}] Esperando 3s para que carguen los mensajes...", 'info')
        time.sleep(3)
        
        # 3) Cerrar todo nuevamente
        for cmd in close_commands:
            self._run_adb_command(cmd, timeout=3)
        
        self.log(f"[{device}] Cambio de cuenta completado", 'success')
    
    def _run_adb_command(self, args, timeout=10):
        """Ejecuta un comando ADB y maneja errores comunes."""
        # --- FIX: Añadir bucle de pausa ---
        while self.is_paused and not self.should_stop:
            time.sleep(0.1)
        if self.should_stop:
            return False # Indicar que la tarea fue cancelada
        # --- FIN FIX ---

        adb = self.adb_path.get()
        full_args = [adb] + args # Construye la lista completa de argumentos

        # Ocultar ventana de consola de ADB
        startupinfo = subprocess.STARTUPINFO()
        startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
        startupinfo.wShowWindow = subprocess.SW_HIDE

        try:
            # Ejecutar SIEMPRE como lista, NUNCA con shell=True si hay rutas
            result = subprocess.run(full_args, capture_output=True, text=True, timeout=timeout, startupinfo=startupinfo, check=False, encoding='utf-8', errors='ignore')
            if result.returncode != 0 and result.stderr:
                # Limpiar errores comunes
                stderr_clean = result.stderr.strip()
                if "Usage: input" in stderr_clean:
                    stderr_clean = "Error Uso Input (argumentos inválidos)"
                elif "NullPointerException" in stderr_clean:
                    stderr_clean = "NullPointerException (procesando texto)"
                elif "Killed" in stderr_clean:
                    stderr_clean = "Proceso Killed (mensaje largo?)"
                elif "unknown command" in stderr_clean:
                     stderr_clean = "Comando ADB desconocido"
                elif "device unauthorized" in stderr_clean:
                     stderr_clean = "Dispositivo no autorizado (revisa el teléfono)"
                elif "device not found" in stderr_clean:
                     stderr_clean = "Dispositivo no encontrado (desconectado?)"

                self.log(f"Error ADB: {stderr_clean}", 'error')
                return False # Indicar fallo
            elif result.returncode != 0:
                 self.log(f"Error ADB (código {result.returncode}, sin stderr)", 'error')
                 return False # Indicar fallo
            return True # Indicar éxito
        except subprocess.TimeoutExpired:
            self.log("Timeout en comando ADB", 'error')
            return False
        except Exception as e:
            self.log(f"Error inesperado ejecutando ADB: {e}", 'error')
            return False

    def _controlled_sleep(self, duration):
        """Realiza una espera respetando pausas y cancelaciones."""
        try:
            duration = max(0.0, float(duration))
        except (TypeError, ValueError):
            duration = 0.0
        elapsed = 0.0
        while elapsed < duration and not self.should_stop:
            while self.is_paused and not self.should_stop:
                time.sleep(0.1)
            if self.should_stop:
                break
            time.sleep(0.1)
            elapsed += 0.1

    def _locate_chat_field(self, ui_device, wait_timeout=0.1):
        """Busca el campo de texto de la conversación actual de forma eficiente."""
        if ui_device is None:
            return None

        selectors = [
            # 1. Por resourceId (el más fiable)
            dict(resourceId="com.whatsapp:id/entry"),
            dict(resourceId="com.whatsapp.w4b:id/entry"),
            # 2. Por className (debe ser un EditText)
            dict(className="android.widget.EditText"),
            # 3. Por descripción de accesibilidad (fallback)
            dict(description="Mensaje"),
            dict(description="Message"),
            dict(descriptionMatches="(?i).*mensaje.*"),
            dict(descriptionMatches="(?i).*message.*"),
        ]

        for selector in selectors:
            try:
                candidate = ui_device(**selector)
                if candidate.wait(timeout=wait_timeout):
                    info = candidate.info or {}
                    bounds = info.get('bounds') or {}
                    if not bounds or bounds.get('left') == bounds.get('right'):
                        continue

                    self.log(f"✓ Campo de texto encontrado con selector: {selector}", "info")
                    return candidate
            except Exception:
                continue

        return None

    def _locate_message_send_button(self, ui_device, is_sms=False, wait_timeout=2):
        """Busca el botón de enviar dentro de la conversación actual."""
        if ui_device is None:
            return None

        selectors = [
            dict(description="Enviar"),
            dict(description="Enviar mensaje"),
            dict(description="Send"),
            dict(descriptionMatches="(?i).*enviar.*"),
            dict(descriptionMatches="(?i).*send.*"),
            dict(text="Enviar"),
            dict(text="Enviar mensaje"),
            dict(text="Send"),
            dict(textMatches="(?i).*enviar.*"),
            dict(textMatches="(?i).*send.*"),
            dict(textMatches=r"^→$"),
            dict(resourceId="com.whatsapp:id/send"),
            dict(resourceId="com.whatsapp.w4b:id/send"),
            dict(resourceId="com.whatsapp:id/send_button"),
            dict(resourceId="com.whatsapp.w4b:id/send_button"),
            dict(resourceIdMatches=r"(?i)com\.whatsapp(\.w4b)?\:id/(send(_button)?)")
        ]

        if is_sms:
            selectors.extend([
                dict(resourceIdMatches="(?i).*send.*"),
                dict(descriptionMatches="(?i).*sms.*enviar.*"),
                dict(textMatches="(?i).*sms.*enviar.*"),
            ])

        for attempt in range(2):
            for selector in selectors:
                # Verificar cancelación/pausa entre cada intento de selector
                if self.should_stop:
                    return None

                while self.is_paused and not self.should_stop:
                    time.sleep(0.1)

                try:
                    candidate = ui_device(**selector)

                    # Algunos nodos fallan en el método "wait" con errores del JSON-RPC.
                    # Para evitar que el flujo completo se interrumpa, comprobamos de forma
                    # segura la existencia y pasamos al siguiente selector si ocurre un error.
                    try:
                        candidate_exists = candidate.exists
                    except Exception:
                        candidate_exists = False

                    if not candidate_exists:
                        try:
                            if not candidate.wait_exists(timeout=wait_timeout):
                                continue
                        except Exception:
                            continue

                    info = candidate.info or {}
                    class_name = (info.get('className') or '').lower()
                    if 'edittext' in class_name:
                        continue

                    bounds = info.get('bounds') or {}
                    if not bounds or bounds.get('left') == bounds.get('right'):
                        continue

                    return candidate
                except Exception:
                    continue

            if attempt == 0 and not self.should_stop:
                time.sleep(0.1)

        return None

    def _locate_join_group_button(self, ui_device, wait_timeout=0.2):
        """Busca el botón de unirse/aceptar en pantallas de invitación a grupos."""
        if ui_device is None:
            return None

        selectors = [
            dict(textMatches="(?i)unirme"),
            dict(textMatches="(?i)unirse"),
            dict(textMatches="(?i)unir(me)?(\\s+al)?\\s+grupo"),
            dict(textMatches="(?i)join\\s*(group)?"),
            dict(textMatches="(?i)acept(ar|ar\\s+invitaci[oó]n)?"),
            dict(textMatches="(?i)accept"),
            dict(descriptionMatches="(?i)unirme"),
            dict(descriptionMatches="(?i)unirse"),
            dict(descriptionMatches="(?i)join\\s*(group)?"),
            dict(descriptionMatches="(?i)acept(ar|ar\\s+invitaci[oó]n)?"),
            dict(descriptionMatches="(?i)accept"),
        ]

        for selector in selectors:
            try:
                candidate = ui_device(**selector)
                if candidate.wait(timeout=wait_timeout):
                    info = candidate.info or {}
                    bounds = info.get('bounds') or {}
                    if bounds and bounds.get('left') != bounds.get('right'):
                        return candidate
            except Exception:
                continue

        return None

    def _wait_for_chat_ready(self, ui_device, wait_time, expected_package=None, is_sms=False, allow_group_join_check=False):
        """
        Espera a que WhatsApp/SMS esté listo para escribir y devuelve los objetos de la UI.
        Retorna (chat_field, send_button) si tiene éxito, o (None, None) si falla.
        """
        if ui_device is None:
            return None, None

        deadline = time.time() + max(0, wait_time)
        poll_interval = 0.2  # Intervalo de sondeo rápido
        chat_field = None
        send_button = None
        deadline_extended = False  # Flag para asegurar que solo extendemos el tiempo una vez
        join_attempted = False
        join_required_detected = False

        while not self.should_stop:
            # Check for pause
            while self.is_paused and not self.should_stop:
                time.sleep(0.1)

            if self.should_stop:
                return None, None

            # Si ya encontramos ambos elementos, el trabajo está hecho.
            if chat_field and send_button:
                break

            # Detectar pantallas de invitación a grupos antes de buscar el chat listo
            if allow_group_join_check and not is_sms:
                join_button = self._locate_join_group_button(ui_device, wait_timeout=poll_interval)
                if join_button:
                    join_required_detected = True
                    if not self.auto_join_groups.get():
                        self.log("✗ Se detectó pantalla de unión a grupo. Auto-unirse desactivado.", 'error')
                        return None, None

                    if not join_attempted:
                        try:
                            join_button.click()
                            join_attempted = True
                            self.log("✓ Pantalla de unión detectada. Haciendo clic para unirse...", 'info')
                            deadline += 6  # Dar tiempo extra tras unirse
                            self._controlled_sleep(0.6)
                            continue
                        except Exception as e:
                            self.log(f"✗ No se pudo hacer clic en 'Unirme/Aceptar': {e}", 'error')
                            return None, None

            # Buscar el campo de texto si aún no lo tenemos
            if not chat_field:
                chat_field = self._locate_chat_field(ui_device, wait_timeout=poll_interval)

            # Buscar el botón de envío si aún no lo tenemos
            if not send_button:
                button = self._locate_message_send_button(
                    ui_device, is_sms=is_sms, wait_timeout=poll_interval
                )
                if button:
                    send_button = button

            # Comprobar si hemos superado el tiempo de espera
            if time.time() > deadline:
                # LÓGICA DE ESPERA INTELIGENTE:
                # Si hemos encontrado solo uno de los dos elementos y aún no hemos extendido
                # el tiempo, lo extendemos unos segundos más para darle una oportunidad.
                if (chat_field or send_button) and not deadline_extended:
                    self.log("✓ Se encontró un elemento del chat, extendiendo la espera 4s...", 'info')
                    deadline += 4  # Extender por 4 segundos
                    deadline_extended = True
                    continue  # Saltar al siguiente ciclo con el nuevo tiempo límite

                # Si no se encontró nada o si ya extendimos el tiempo, nos rendimos.
                break

            # Si aún no hemos encontrado ambos, esperamos un poco antes de volver a intentar
            if not (chat_field and send_button):
                self._controlled_sleep(poll_interval)

        # Evaluar el resultado final (sin cambios aquí)
        if chat_field and send_button:
            self.log("✓ Conversación lista (campo y botón encontrados).", 'success')
            return chat_field, send_button

        # Si falla, registrar el motivo específico
        if not chat_field and not send_button:
            if join_required_detected:
                self.log("✗ Chat no listo: Se detectó pantalla de unión a grupo pero no se completó.", 'error')
            else:
                self.log("✗ Chat no listo: No se encontró ni el campo de texto ni el botón de envío.", 'error')
        elif not chat_field:
            self.log("✗ Chat no listo: Se encontró el botón de envío, pero no el campo de texto.", 'error')
        elif not send_button:
            self.log("✗ Chat no listo: Se encontró el campo de texto, pero no el botón de envío.", 'error')

        return None, None


    def send_msg(
        self,
        device,
        link,
        i,
        total,
        message_to_send=None,
        whatsapp_package="com.whatsapp.w4b",
        ui_device=None,
        primary_index=None,
    ):
        """Ejecuta los comandos para enviar un único mensaje."""
        is_group = bool(message_to_send)
        try:
            if not ui_device:
                self.log("✗ Error crítico: la conexión de uiautomator2 no está disponible.", "error")
                return False

            attempt_links = [link]
            if primary_index is not None:
                alternates = self.link_retry_map.get(primary_index, [])
                if alternates:
                    attempt_links.extend(alternates)

            total_attempts = len(attempt_links)

            def describe_link(current_link, is_group_flag):
                lower_link = current_link.lower()
                if is_group_flag and not lower_link.startswith("https://wa.me/"):
                    return f"Grupo ({current_link[:40]}...)"
                if lower_link.startswith("sms:"):
                    return current_link.split("sms:", 1)[1].split('?')[0]
                if 'wa.me/' in current_link:
                    return current_link.split('wa.me/')[1].split('?')[0]
                return current_link[:50]

            def attempt_send(current_link, attempt_idx, is_group_flag):
                local_is_sms = current_link.lower().startswith("sms:")
                num_display = describe_link(current_link, is_group_flag)
                intento = "Principal" if attempt_idx == 0 else f"Alternativo {attempt_idx}"
                log_prefix = f"({i}/{total}) → {num_display} [en {device}] ({intento})"

                # --- MODO SMS SIMPLIFICADO ---
                if local_is_sms:
                    # Asegurar pantalla encendida
                    self._run_adb_command(['-s', device, 'shell', 'input', 'keyevent', 'KEYCODE_WAKEUP'], timeout=5)
                    self._run_adb_command(['-s', device, 'shell', 'input', 'keyevent', 'KEYCODE_MENU'], timeout=5)

                    # 1. Inyectar URL
                    open_args = ['-s', device, 'shell', 'am', 'start', '-a', 'android.intent.action.VIEW', '-d', f'"{current_link}"']
                    if not self._run_adb_command(open_args, timeout=20):
                        self.log(log_prefix, 'error')
                        self.log(f"  └─ Motivo: No se pudo abrir la app de SMS.", 'error')
                        return False, False

                    # 2. Esperar tiempo "Esperar a enviar"
                    try:
                        wait_send = max(0, int(self.wait_after_first_enter.get()))
                    except:
                        wait_send = 2

                    # self.log(f"Esperando {wait_send}s para enviar...", 'info') # Verbose opcional
                    self._controlled_sleep(wait_send)

                    if self.should_stop: return False, False

                    # 3. Buscar y Clickear Botón Enviar
                    send_btn = self._locate_message_send_button(ui_device, is_sms=True, wait_timeout=2)
                    if send_btn:
                        try:
                            send_btn.click()
                            self.log(log_prefix, 'success')
                            return True, False
                        except Exception as e:
                            self.log(log_prefix, 'error')
                            self.log(f"  └─ Motivo: Error al hacer clic en el botón enviar: {e}", 'error')
                            return False, False
                    else:
                        self.log(log_prefix, 'error')
                        self.log(f"  └─ Motivo: No se encontró el botón de enviar SMS.", 'error')
                        return False, False

                # --- MODO WHATSAPP (Lógica original) ---
                open_args = ['-s', device, 'shell', 'am', 'start', '-a', 'android.intent.action.VIEW', '-d', f'"{current_link}"']
                if whatsapp_package:
                    open_args.extend(['-p', whatsapp_package])
                if not self._run_adb_command(open_args, timeout=20):
                    self.log(log_prefix, 'error')
                    self.log(f"  └─ Motivo: No se pudo abrir el chat.", 'error')
                    return False, False

                active_package = whatsapp_package

                wait_time = max(0, int(self.wait_after_open.get()))
                if is_group_flag and "chat.whatsapp.com" in current_link.lower():
                    wait_time += 5
                chat_field, send_button = self._wait_for_chat_ready(
                    ui_device,
                    wait_time,
                    expected_package=active_package,
                    is_sms=False,
                    allow_group_join_check=is_group,
                )

                if not chat_field or not send_button:
                    self.log(log_prefix, 'error')
                    self.log(f"  └─ Motivo: El chat no estuvo listo para enviar (timeout).", 'error')
                    return False, False
                if self.should_stop:
                    return False, False

                msg_to_send = message_to_send
                if not msg_to_send and not is_group_flag:
                    try:
                        if 'text=' in current_link:
                            msg_to_send = urllib.parse.unquote(current_link.split('text=')[1])
                    except Exception:
                        msg_to_send = None

                try:
                    needs_text_input = is_group_flag
                    if needs_text_input and msg_to_send is not None:
                        # --- Lógica robusta anti-audio (Modo Grupo) ---

                        # 1. Descartar el botón encontrado anteriormente (Microfono)
                        send_button = None

                        # 2. Asegurar el foco
                        try:
                            chat_field.click()
                            self._controlled_sleep(0.3)
                        except Exception as e:
                            self.log(f"Advertencia: no se pudo hacer clic en el campo: {e}", "warning")

                        # 3. Ciclo de intento de escritura y verificación
                        text_verified = False

                        # Intento 1: set_text
                        try:
                            chat_field.set_text(msg_to_send)
                            self._controlled_sleep(0.8) # Esperar a que la UI reaccione

                            # Verificación
                            curr_text = chat_field.get_text()
                            if curr_text and len(curr_text.strip()) > 0:
                                text_verified = True
                        except Exception:
                            pass

                        # Intento 2: keyevents (si falló el 1)
                        if not text_verified:
                            self.log("Texto no detectado tras set_text. Reintentando con pulsaciones...", "warning")
                            # Asegurar foco nuevamente
                            try: chat_field.click()
                            except: pass

                            self._write_message_with_keyevents(device, msg_to_send)
                            self._controlled_sleep(1.0)

                            # Verificación final
                            try:
                                curr_text = chat_field.get_text()
                                if curr_text and len(curr_text.strip()) > 0:
                                    text_verified = True
                            except:
                                pass

                        # 4. PUNTO CRÍTICO: Si no hay texto, ABORTAR
                        if not text_verified:
                            self.log(log_prefix, 'error')
                            self.log(f"  └─ Motivo: No se pudo escribir el mensaje en el campo. Se aborta para evitar enviar audio.", 'error')
                            return False, False

                        # 5. Buscar el botón de enviar (Avión)
                        # Reintentar varias veces porque el cambio de icono puede tardar
                        found_send = False
                        for _ in range(3):
                            new_send_btn = self._locate_message_send_button(ui_device, is_sms=False, wait_timeout=1)
                            if new_send_btn:
                                # Validación extra: Chequear descripción si es posible para evitar micrófono
                                try:
                                    desc = (new_send_btn.info.get('contentDescription') or "").lower()
                                    # Si dice "grabar" o "voice", es peligroso.
                                    if "grabar" in desc or "voice" in desc or "voz" in desc:
                                        self.log("Botón detectado como Micrófono. Esperando...", "warning")
                                        self._controlled_sleep(0.5)
                                        continue
                                except:
                                    pass

                                send_button = new_send_btn
                                found_send = True
                                break
                            self._controlled_sleep(0.5)

                        if not found_send:
                            self.log(log_prefix, 'error')
                            self.log("  └─ Motivo: No se encontró el botón de enviar (Avión) tras escribir.", 'error')
                            return False, False

                    # Usar el send_button (ya sea el original o el re-localizado)
                    if send_button:
                        send_button.click()
                    else:
                        self.log(log_prefix, 'error')
                        self.log(f"  └─ Motivo: No se pudo hacer clic (botón de enviar perdido).", 'error')
                        return False, False

                    self._controlled_sleep(1.0)

                    if self._detect_send_failure(ui_device):
                        self.log(log_prefix, 'error')
                        self.log(f"  └─ Motivo: WhatsApp reportó un fallo de envío (mensaje 'No se envió').", 'error')
                        return False, True

                    self.log(log_prefix, 'success')
                    self._controlled_sleep(0.6)
                    return True, False

                except Exception as e:
                    self.log(log_prefix, 'error')
                    self.log(f"  └─ Motivo: Error inesperado ({e}).", 'error')
                    return False, False

            for attempt_idx, current_link in enumerate(attempt_links):
                if self.should_stop:
                    return False

                if attempt_idx > 0:
                    self.log("Intentando con un número alternativo...", 'warning')
                    self.close_all_apps(device)
                    if self.should_stop:
                        return False

                success, should_retry = attempt_send(current_link, attempt_idx, is_group)
                if success:
                    return True

                if should_retry and attempt_idx < total_attempts - 1:
                    continue

                if should_retry and attempt_idx == total_attempts - 1:
                    self.log("✗ No se pudo enviar el mensaje tras agotar los números disponibles.", 'error')
                return False

            return False

        except Exception as e:
            self.log(f"✗ Error crítico en la lógica de envío: {e}", 'error')
            import traceback
            traceback.print_exc()
            return False

    def _detect_send_failure(self, ui_device, timeout=1.0):
        """Verifica si aparece el mensaje de error de WhatsApp/SMS indicando que no se envió."""
        failure_selectors = [
            dict(textMatches="(?i)no se envi[óo]"),
            dict(descriptionMatches="(?i)no se envi[óo]"),
            dict(textMatches="(?i)mensaje no enviado"),
            dict(descriptionMatches="(?i)mensaje no enviado"),
            dict(textMatches="(?i)reintentar"),
            dict(descriptionMatches="(?i)reintentar"),
            dict(resourceIdMatches="(?i).*retry.*"),
            dict(resourceIdMatches="(?i).*resend.*"),
        ]

        for selector in failure_selectors:
            try:
                node = ui_device(**selector)
                if timeout:
                    if node.wait(timeout=timeout):
                        return True
                elif node.exists:
                    return True
            except Exception:
                continue
        return False
    # --- ################################################################## ---
    # --- FIN
    # --- ################################################################## ---
    def run_cambiador(self):
        """Ejecuta la secuencia Cambiador en todos los dispositivos."""
        if not self.devices:
            messagebox.showwarning("Sin dispositivos", "No hay dispositivos conectados. Detecta dispositivos primero.", parent=self.root)
            return
        
        adb_exe = self.adb_path.get()
        if not adb_exe or not os.path.exists(adb_exe):
            messagebox.showerror("Error ADB", "No se encontró la ruta de ADB. Detecta dispositivos primero.", parent=self.root)
            return
        
        confirm = messagebox.askyesno(
            "Cambiador",
            f"Se ejecutará la secuencia Cambiador en {len(self.devices)} dispositivo(s).\n\n"
            "Secuencia:\n"
            "• Abrir configuración de WhatsApp Business\n"
            "• Navegar y ejecutar acciones (13 TABs + ENTERs)\n"
            "• Abrir configuración de WhatsApp Normal\n"
            "• Navegar y ejecutar acciones (13 TABs + ENTERs)\n\n"
            "¿Continuar?",
            parent=self.root
        )
        
        if not confirm:
            return
        
        self.log(f"Iniciando Cambiador en {len(self.devices)} dispositivo(s)...", 'info')
        
        for idx, device in enumerate(self.devices, 1):
            self.log(f"[{idx}/{len(self.devices)}] Procesando: {device}", 'info')
            
            try:
                # Secuencia para WhatsApp Business
                self.log(f"  → Procesando WhatsApp Business en {device}", 'info')
                
                # Abrir configuración de WhatsApp Business
                cmd = f'"{adb_exe}" -s {device} shell am start -a android.settings.APPLICATION_DETAILS_SETTINGS -d package:com.whatsapp.w4b'
                subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=5)
                time.sleep(3)  # Esperar 3 segundos para que la app se abra completamente
                
                # 13 TABs
                for i in range(13):
                    cmd = f'"{adb_exe}" -s {device} shell input keyevent KEYCODE_TAB'
                    subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=5)
                    time.sleep(0.05)
                
                # ENTER
                cmd = f'"{adb_exe}" -s {device} shell input keyevent KEYCODE_ENTER'
                subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=5)
                time.sleep(0.05)
                
                # TAB
                cmd = f'"{adb_exe}" -s {device} shell input keyevent KEYCODE_TAB'
                subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=5)
                time.sleep(0.05)
                
                # ENTER
                cmd = f'"{adb_exe}" -s {device} shell input keyevent KEYCODE_ENTER'
                subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=5)
                time.sleep(0.05)
                
                # Cerrar configuración
                cmd = f'"{adb_exe}" -s {device} shell am force-stop com.android.settings'
                subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=5)
                time.sleep(0.05)
                
                # Secuencia para WhatsApp Normal
                self.log(f"  → Procesando WhatsApp Normal en {device}", 'info')
                
                # Abrir configuración de WhatsApp Normal
                cmd = f'"{adb_exe}" -s {device} shell am start -a android.settings.APPLICATION_DETAILS_SETTINGS -d package:com.whatsapp'
                subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=5)
                time.sleep(3)  # Esperar 3 segundos para que la app se abra completamente
                
                # 13 TABs
                for i in range(13):
                    cmd = f'"{adb_exe}" -s {device} shell input keyevent KEYCODE_TAB'
                    subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=5)
                    time.sleep(0.05)
                
                # ENTER
                cmd = f'"{adb_exe}" -s {device} shell input keyevent KEYCODE_ENTER'
                subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=5)
                time.sleep(0.05)
                
                # TAB
                cmd = f'"{adb_exe}" -s {device} shell input keyevent KEYCODE_TAB'
                subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=5)
                time.sleep(0.05)
                
                # ENTER
                cmd = f'"{adb_exe}" -s {device} shell input keyevent KEYCODE_ENTER'
                subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=5)
                time.sleep(0.05)
                
                # Cerrar configuración
                cmd = f'"{adb_exe}" -s {device} shell am force-stop com.android.settings'
                subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=5)
                time.sleep(0.05)
                
                self.log(f"✓ Dispositivo {device} completado", 'success')
                
            except subprocess.TimeoutExpired:
                self.log(f"✗ Timeout en dispositivo {device}", 'error')
            except Exception as e:
                self.log(f"✗ Error en {device}: {str(e)}", 'error')
        
        self.log("✓ Cambiador completado en todos los dispositivos", 'success')
        messagebox.showinfo("Completado", f"Cambiador ejecutado en {len(self.devices)} dispositivo(s).", parent=self.root)
    
    def switch_whatsapp_account(self):
        """Cambia de cuenta en WhatsApp para todos los dispositivos."""
        if not self.devices:
            messagebox.showwarning("Sin dispositivos", "No hay dispositivos conectados. Detecta dispositivos primero.", parent=self.root)
            return
        
        adb_exe = self.adb_path.get()
        if not adb_exe or not os.path.exists(adb_exe):
            messagebox.showerror("Error ADB", "No se encontró la ruta de ADB. Detecta dispositivos primero.", parent=self.root)
            return
        
        confirm = messagebox.askyesno(
            "Cambiar Cuenta WhatsApp",
            f"Se ejecutará el cambio de cuenta en {len(self.devices)} dispositivo(s).\n\n"
            "Secuencia:\n"
            "• Abrir WhatsApp\n"
            "• Navegar al menú\n"
            "• Cambiar cuenta\n\n"
            "¿Continuar?",
            parent=self.root
        )
        
        if not confirm:
            return
        
        self.log(f"Iniciando cambio de cuenta en {len(self.devices)} dispositivo(s)...", 'info')
        
        for idx, device in enumerate(self.devices, 1):
            self.log(f"[{idx}/{len(self.devices)}] Procesando: {device}", 'info')
            
            # Usar la función _switch_account_for_device con delay de 0.4s (4x más lento)
            success = self._switch_account_for_device(device, delay=0.4)
            
            if success:
                self.log(f"✓ Dispositivo {device} completado", 'success')
            else:
                self.log(f"✗ Error en dispositivo {device}", 'error')
        
        self.log("✓ Cambio de cuenta completado en todos los dispositivos", 'success')
        messagebox.showinfo("Completado", f"Cambio de cuenta ejecutado en {len(self.devices)} dispositivo(s).", parent=self.root)
    
    def _switch_account_for_device(self, device, delay=0.1):
        """Cambia de cuenta en WhatsApp Normal para un dispositivo específico.
        
        Args:
            device: ID del dispositivo
            delay: Tiempo de espera entre comandos (default 0.1s para cambio automático rápido)
        """
        adb_exe = self.adb_path.get()
        if not adb_exe or not os.path.exists(adb_exe):
            self.log(f"Error: No se encontró ADB para cambiar cuenta en {device}", 'error')
            return False
        
        try:
            commands = [
                "shell am start -n com.whatsapp/.Main",
                "shell input keyevent KEYCODE_DPAD_UP",
                "shell input keyevent KEYCODE_DPAD_UP",
                "shell input keyevent KEYCODE_DPAD_RIGHT",
                "shell input keyevent KEYCODE_ENTER",
                "shell input keyevent KEYCODE_TAB",
                "shell input keyevent KEYCODE_TAB",
                "shell input keyevent KEYCODE_TAB",
                "shell input keyevent KEYCODE_TAB",
                "shell input keyevent KEYCODE_TAB",
                "shell input keyevent KEYCODE_TAB",
                "shell input keyevent KEYCODE_TAB",
                "shell input keyevent KEYCODE_ENTER"
            ]
            
            for cmd in commands:
                full_cmd = f'"{adb_exe}" -s {device} {cmd}'
                subprocess.run(full_cmd, shell=True, capture_output=True, text=True, timeout=5)
                time.sleep(delay)  # Tiempo configurable según el contexto
            
            self.log(f"✓ Cuenta cambiada en {device}", 'success')
            return True
            
        except subprocess.TimeoutExpired:
            self.log(f"✗ Timeout al cambiar cuenta en {device}", 'error')
            return False
        except Exception as e:
            self.log(f"✗ Error al cambiar cuenta en {device}: {str(e)}", 'error')
            return False


    def detect_phone_numbers_thread(self):
        """Inicia la detección de números de teléfono en un hilo separado."""
        if not self.devices:
            messagebox.showwarning("Sin dispositivos", "No hay dispositivos conectados. Detecta dispositivos primero.", parent=self.root)
            return
        threading.Thread(target=self._detect_and_prepare_phone_lines, args=(True,), daemon=True).start()

    def _infer_lines_from_installed_apps(self, log_errors=True):
        """Inferir líneas disponibles solo revisando paquetes instalados cuando no hay números detectados."""
        inferred = []
        adb_binary = self.adb_path.get() if hasattr(self, 'adb_path') else None

        if not adb_binary or not os.path.exists(adb_binary):
            return inferred

        for device_serial in self.devices:
            try:
                result = subprocess.run(
                    [adb_binary, '-s', device_serial, 'shell', 'pm', 'list', 'packages'],
                    capture_output=True,
                    text=True,
                    timeout=8,
                    encoding='utf-8',
                    errors='ignore'
                )
                if result.returncode != 0:
                    if log_errors:
                        self.log(f"No se pudo listar paquetes en {device_serial}.", 'warning')
                    continue

                packages = {line.split(':', 1)[1] for line in result.stdout.splitlines() if line.startswith('package:')}

                if 'com.whatsapp' in packages:
                    inferred.append({"device": device_serial, "type": "WhatsApp", "number": "(sin detectar)", "package": "com.whatsapp"})
                if 'com.whatsapp.w4b' in packages:
                    inferred.append({"device": device_serial, "type": "WhatsApp Business", "number": "(sin detectar)", "package": "com.whatsapp.w4b"})

            except Exception as e:
                if log_errors:
                    self.log(f"Error al inferir apps instaladas en {device_serial}: {e}", 'warning')

        return inferred

    def _detect_and_prepare_phone_lines(self, show_results_popup=False, auto_confirm=False):
        """
        Detecta todos los números de WA y WA Business en los dispositivos conectados
        y los almacena en self.detected_phone_lines.
        Retorna True si se encontraron líneas, False en caso contrario.
        """
        self.log("Iniciando detección de números de teléfono...", 'info')
        self.numbers_editor_start_requested = False
        detect_btn = getattr(self, 'detect_numbers_btn', None)
        if detect_btn and detect_btn.winfo_exists():
            self.root.after(0, detect_btn.configure, {'state': tk.DISABLED, 'text': "Detectando..."})

        self.detected_phone_lines = []
        temp_dir = tempfile.mkdtemp(prefix="hermes_ui_dump_")

        try:
            all_numbers = {}
            for device_serial in self.devices:
                self.log(f"Procesando dispositivo: {device_serial}", 'info')
                numbers = self._get_number_with_uiautomator2(device_serial, temp_dir)
                all_numbers[device_serial] = numbers

                if numbers.get("WhatsApp") not in ["No encontrado", "Error"]:
                    self.detected_phone_lines.append({"device": device_serial, "type": "WhatsApp", "number": numbers["WhatsApp"], "package": "com.whatsapp"})
                if numbers.get("WhatsApp Business") not in ["No encontrado", "Error"]:
                    self.detected_phone_lines.append({"device": device_serial, "type": "WhatsApp Business", "number": numbers["WhatsApp Business"], "package": "com.whatsapp.w4b"})

            snapshot_copy = {dev: numbers.copy() for dev, numbers in all_numbers.items()}
            self.last_detection_snapshot = snapshot_copy
            self._update_detected_numbers_summary(snapshot_copy)

            if show_results_popup:
                self.log("Detección de números finalizada.", 'success')
                summary_lines = self._format_detected_numbers_summary_lines(snapshot_copy)
                if summary_lines:
                    popup_text = "Líneas detectadas:\n" + "\n".join(f"· {line}" for line in summary_lines)
                else:
                    popup_text = "No se encontraron líneas de WhatsApp válidas."
                self.root.after(0, lambda: messagebox.showinfo("Detección Finalizada", popup_text, parent=self.root))

            if snapshot_copy:
                if auto_confirm:
                    self.log("Confirmación automática de números detectados (Modo Programado).", 'info')
                    self.numbers_editor_start_requested = True
                    # No mostramos el editor, pero aseguramos que el flag de "confirmado" esté activo
                else:
                    if hasattr(self, 'numbers_editor_closed_event'):
                        self.numbers_editor_closed_event.clear()
                    self.root.after(0, lambda data=snapshot_copy: self._show_detected_numbers_editor(data))
                    if hasattr(self, 'numbers_editor_closed_event'):
                        self.numbers_editor_closed_event.wait()
            else:
                if hasattr(self, 'numbers_editor_closed_event'):
                    self.numbers_editor_closed_event.set()

            return bool(self.detected_phone_lines)

        except Exception as e:
            self.log(f"Error durante la detección de números: {e}", 'error')
            if show_results_popup:
                self.root.after(0, lambda: messagebox.showerror("Error", f"Ocurrió un error inesperado:\n{e}", parent=self.root))
            return False
        finally:
            if hasattr(self, 'numbers_editor_closed_event') and not self.numbers_editor_closed_event.is_set():
                self.numbers_editor_closed_event.set()
            try:
                shutil.rmtree(temp_dir)
            except Exception as e:
                self.log(f"Error al eliminar el directorio temporal: {e}", "warning")
            detect_btn = getattr(self, 'detect_numbers_btn', None)
            if detect_btn and detect_btn.winfo_exists():
                self.root.after(0, detect_btn.configure, {'state': tk.NORMAL, 'text': "Detectar Números"})

    def _format_detected_numbers_summary_lines(self, snapshot=None):
        """Genera las líneas de resumen para los números detectados."""
        snapshot = snapshot if snapshot is not None else getattr(self, 'last_detection_snapshot', {})
        lines = []
        for device in sorted(snapshot.keys()):
            numbers = snapshot[device]
            segments = []
            for whatsapp_type in ("WhatsApp", "WhatsApp Business"):
                raw_value = numbers.get(whatsapp_type, "No encontrado")
                display_value = raw_value if raw_value not in (None, "") else "No encontrado"
                segments.append(f"{whatsapp_type}: {display_value}")
            if segments:
                lines.append(f"{device} → " + " | ".join(segments))
        return lines

    def _update_detected_numbers_summary(self, snapshot=None):
        """Actualiza la etiqueta principal con el resumen de números detectados."""
        lines = self._format_detected_numbers_summary_lines(snapshot)
        if lines:
            self.log("Resumen de líneas detectadas:", 'info')
            for line in lines:
                self.log(f"· {line}", 'info')
        else:
            self.log("No se encontraron líneas de WhatsApp válidas.", 'warning')

    def _close_numbers_editor(self, start_requested=False):
        """Cierra la ventana del editor de números (si existe)."""
        if start_requested:
            self.numbers_editor_start_requested = True
        else:
            self.numbers_editor_start_requested = False
        if self.numbers_editor_window is not None:
            try:
                self.numbers_editor_window.destroy()
            except tk.TclError:
                pass
        self.numbers_editor_window = None
        self.numbers_editor_entries = {}
        if hasattr(self, 'numbers_editor_closed_event'):
            self.numbers_editor_closed_event.set()

    def _save_numbers_editor(self, start_after_save=False):
        """Guarda los cambios realizados manualmente en la tabla de números detectados."""
        if not self.numbers_editor_entries:
            self._close_numbers_editor(start_requested=start_after_save)
            return

        updated_snapshot = {}
        updated_lines = []
        for device, entry_map in self.numbers_editor_entries.items():
            updated_snapshot[device] = {}
            for whatsapp_type, entry in entry_map.items():
                value = entry.get().strip()
                sanitized = value.replace(" ", "").replace("-", "").replace("(", "").replace(")", "")
                if sanitized:
                    updated_snapshot[device][whatsapp_type] = sanitized
                    package = 'com.whatsapp' if whatsapp_type == 'WhatsApp' else 'com.whatsapp.w4b'
                    updated_lines.append({
                        "device": device,
                        "type": whatsapp_type,
                        "number": sanitized,
                        "package": package
                    })
                else:
                    previous_value = self.last_detection_snapshot.get(device, {}).get(whatsapp_type, "No encontrado")
                    updated_snapshot[device][whatsapp_type] = previous_value if previous_value == "Error" else "No encontrado"

        self.detected_phone_lines = updated_lines
        self.last_detection_snapshot = updated_snapshot
        self._update_detected_numbers_summary(updated_snapshot)
        self.log("Números ajustados manualmente desde el editor.", 'info')
        self._close_numbers_editor(start_requested=start_after_save)
        if not start_after_save:
            messagebox.showinfo("Números actualizados", "Los cambios se guardaron correctamente.", parent=self.root)

    def _show_detected_numbers_editor(self, detection_results):
        """Muestra una tabla editable con los números detectados por dispositivo."""
        if not detection_results:
            if hasattr(self, 'numbers_editor_closed_event'):
                self.numbers_editor_closed_event.set()
            return

        self._close_numbers_editor()

        editor = ctk.CTkToplevel(self.root)
        editor.title("Editar números detectados")
        editor.resizable(False, False)
        editor.transient(self.root)
        editor.grab_set()
        editor.focus_force()
        editor.configure(fg_color=self.colors['bg'])

        width, height = 760, 470
        self._center_toplevel(editor, width, height)

        outer = ctk.CTkFrame(editor, fg_color="transparent")
        outer.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        card = ctk.CTkFrame(outer, fg_color=self.colors['bg_card'], corner_radius=22)
        card.pack(fill=tk.BOTH, expand=True)

        header = ctk.CTkFrame(card, fg_color="transparent")
        header.pack(fill=tk.X, padx=25, pady=(25, 10))
        header.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            header,
            text="Verifica las líneas detectadas",
            font=self.fonts['card_title'],
            text_color=self.colors['text']
        ).grid(row=0, column=0, sticky='w')

        header_actions = ctk.CTkFrame(header, fg_color="transparent")
        header_actions.grid(row=0, column=1, sticky='e')

        ctk.CTkButton(
            header_actions,
            text="Confirmar e iniciar",
            command=lambda: self._save_numbers_editor(start_after_save=True),
            font=self.fonts['button'],
            fg_color=self.colors.get('action_start', '#16A34A'),
            hover_color=self.hover_colors.get(
                'action_start',
                darken_color(self.colors.get('action_start', '#16A34A'), 0.18)
            ),
            text_color=self.colors['text_header_buttons'],
            height=40,
            corner_radius=10
        ).grid(row=0, column=0, sticky='e')

        instructions = ctk.CTkLabel(
            card,
            text=("Revisa y ajusta los números detectados antes de continuar. "
                  "Puedes abrir la calculadora del dispositivo para identificarlo."),
            font=self.fonts['setting_label'],
            text_color=self.colors['text_light'],
            wraplength=width - 120,
            justify='left'
        )
        instructions.pack(fill=tk.X, padx=25, pady=(0, 15))

        table_container = ctk.CTkFrame(card, fg_color="transparent")
        table_container.pack(fill=tk.BOTH, expand=True, padx=25, pady=(0, 20))

        # Preparar colores con valores predeterminados para evitar errores si faltan claves
        action_detect_color = self.colors.get('action_detect', '#2563EB')
        action_detect_hover = self.hover_colors.get('action_detect', darken_color(action_detect_color, 0.18))
        action_cancel_color = self.colors.get('action_cancel', '#DC2626')
        action_cancel_hover = self.hover_colors.get('action_cancel', darken_color(action_cancel_color, 0.18))
        action_start_color = self.colors.get('action_start', '#16A34A')
        action_start_hover = self.hover_colors.get('action_start', darken_color(action_start_color, 0.18))

        table = ctk.CTkScrollableFrame(
            table_container,
            fg_color=self.colors['bg'],
            corner_radius=14,
            height=260
        )
        table.pack(fill=tk.BOTH, expand=True)
        table.grid_columnconfigure(0, weight=1)
        table.grid_columnconfigure(1, weight=1)
        table.grid_columnconfigure(2, weight=1)

        headers = ["Dispositivo", "WhatsApp", "WhatsApp Business"]
        for col, header_text in enumerate(headers):
            ctk.CTkLabel(
                table,
                text=header_text,
                font=('Inter', 13, 'bold'),
                text_color=self.colors['text']
            ).grid(row=0, column=col, sticky='ew', padx=12, pady=(12, 6))

        self.numbers_editor_entries = {}
        for row_index, device in enumerate(sorted(detection_results.keys()), start=1):
            values = detection_results.get(device, {})

            device_button = ctk.CTkButton(
                table,
                text=device,
                command=lambda ds=device: self._open_device_calculator(ds),
                font=self.fonts['button_small'],
                fg_color=action_detect_color,
                hover_color=action_detect_hover,
                corner_radius=10,
                height=34
            )
            device_button.grid(row=row_index, column=0, sticky='ew', padx=12, pady=6)

            entries = {}
            for col_index, whatsapp_type in enumerate(("WhatsApp", "WhatsApp Business"), start=1):
                current_value = values.get(whatsapp_type, "No encontrado")
                placeholder = "Error" if current_value == "Error" else "No encontrado"
                entry = ctk.CTkEntry(
                    table,
                    placeholder_text=placeholder,
                    font=self.fonts['setting_label'],
                    corner_radius=10,
                    height=34,
                    border_width=1,
                    border_color="#d1d5db"
                )
                if current_value not in ("No encontrado", "Error", None, ""):
                    entry.insert(0, current_value)
                entry.grid(row=row_index, column=col_index, sticky='ew', padx=12, pady=6)
                entries[whatsapp_type] = entry

            self.numbers_editor_entries[device] = entries

        buttons_frame = ctk.CTkFrame(card, fg_color="transparent")
        buttons_frame.pack(fill=tk.X, padx=25, pady=(0, 25))
        buttons_frame.grid_columnconfigure(0, weight=1)
        buttons_frame.grid_columnconfigure(1, weight=1)
        buttons_frame.grid_columnconfigure(2, weight=1)

        cancel_btn = ctk.CTkButton(
            buttons_frame,
            text="Cancelar",
            command=self._close_numbers_editor,
            font=self.fonts['button'],
            fg_color=action_cancel_color,
            hover_color=action_cancel_hover,
            text_color=self.colors['text_header_buttons'],
            height=40
        )
        cancel_btn.grid(row=0, column=0, sticky='ew', padx=(0, 10))

        save_only_btn = ctk.CTkButton(
            buttons_frame,
            text="Guardar cambios",
            command=self._save_numbers_editor,
            font=self.fonts['button'],
            fg_color=action_detect_color,
            hover_color=action_detect_hover,
            text_color=self.colors['text_header_buttons'],
            height=40
        )
        save_only_btn.grid(row=0, column=1, sticky='ew', padx=5)

        start_btn = ctk.CTkButton(
            buttons_frame,
            text="Confirmar e iniciar",
            command=lambda: self._save_numbers_editor(start_after_save=True),
            font=self.fonts['button'],
            fg_color=action_start_color,
            hover_color=action_start_hover,
            text_color=self.colors['text_header_buttons'],
            height=40
        )
        start_btn.grid(row=0, column=2, sticky='ew', padx=(10, 0))

        self.numbers_editor_window = editor
        editor.protocol("WM_DELETE_WINDOW", self._close_numbers_editor)

    def _open_device_calculator(self, device_serial):
        """Abre la calculadora del dispositivo seleccionado mediante ADB."""
        adb_path = self.adb_path.get()
        if not adb_path or not os.path.exists(adb_path):
            self.auto_detect_adb()
            adb_path = self.adb_path.get()

        if not adb_path or not os.path.exists(adb_path):
            self.log("ADB no configurado. No se puede abrir la calculadora.", 'error')
            messagebox.showerror("ADB no configurado", "No se encontró 'adb.exe'. Configúralo antes de abrir la calculadora.", parent=self.root)
            return

        self.log(f"[{device_serial}] Abriendo calculadora para identificar el dispositivo...", 'info')

        intents = [
            ['-s', device_serial, 'shell', 'am', 'start', '-a', 'android.intent.action.MAIN', '-c', 'android.intent.category.APP_CALCULATOR'],
            ['-s', device_serial, 'shell', 'am', 'start', '-n', 'com.android.calculator2/.Calculator'],
            ['-s', device_serial, 'shell', 'am', 'start', '-n', 'com.sec.android.app.popupcalculator/.Calculator'],
            ['-s', device_serial, 'shell', 'am', 'start', '-n', 'com.miui.calculator/.cal.CalculatorActivity']
        ]

        for intent_args in intents:
            if self._run_adb_command(intent_args, timeout=5):
                self.log(f"[{device_serial}] Calculadora abierta correctamente.", 'success')
                return

        self.log(f"[{device_serial}] No se pudo abrir la calculadora.", 'warning')
        messagebox.showwarning("Calculadora no encontrada", f"No se pudo abrir la calculadora en {device_serial}.", parent=self.root)

    def _get_number_with_uiautomator2(self, device_serial, temp_dir):
        """Usa uiautomator2 para encontrar los números de teléfono en un dispositivo."""
        device_numbers = {"WhatsApp": "No encontrado", "WhatsApp Business": "No encontrado"}
        try:
            d = u2.connect(device_serial)
            # Desbloquear si es necesario (mejor si el usuario lo hace)
            d.unlock()

            # 1. Listar paquetes de WhatsApp según la selección del usuario
            wa_mode = self.whatsapp_mode.get()
            self.log(f"  Optimizando detección para el modo: {wa_mode}", 'info')

            packages_to_scan = []
            if wa_mode == "Normal":
                packages_to_scan.append("com.whatsapp")
            elif wa_mode == "Business":
                packages_to_scan.append("com.whatsapp.w4b")
            else: # Ambas o Todas
                packages_to_scan.append("com.whatsapp")
                packages_to_scan.append("com.whatsapp.w4b")

            # Verificar si los paquetes realmente existen en el dispositivo
            installed_packages_raw = d.shell('pm list packages').output
            installed_packages = installed_packages_raw.splitlines()

            final_packages_to_scan = [f"package:{p}" for p in packages_to_scan if f"package:{p}" in installed_packages]

            if not final_packages_to_scan:
                self.log(f"  No se encontraron las apps de WhatsApp seleccionadas en {device_serial}.", 'warning')
                return {"WhatsApp": "No encontrado", "WhatsApp Business": "No encontrado"}

            for pkg_full in final_packages_to_scan:
                pkg = pkg_full.split(':')[1]
                self.log(f"  Analizando paquete: {pkg} en {device_serial}", 'info')
                number = "No encontrado"
                try:
                    # 2. Abrir la app
                    d.app_start(pkg, stop=True)
                    d.wait_activity(".Main", timeout=12)
                    self.log(f"    {pkg} iniciado.", 'info')

                    # --- PASO A: Buscar en la pantalla principal de chats (Método Robusto)---
                    self.log("    Paso A: Buscando número en la pantalla principal...", 'info')
                    try:
                        time.sleep(2) # Esperar a que carguen los chats

                        # Obtener el volcado de la jerarquía de la UI como XML
                        xml_dump = d.dump_hierarchy()
                        root = ET.fromstring(xml_dump)
                        nodes = list(root.iter())

                        def _extract_number(text_value: str):
                            if not text_value:
                                return None
                            match_local = re.search(r'(\+[\d\s\-\(\)]+)', text_value)
                            if match_local:
                                return re.sub(r'[\s\-\(\)]', '', match_local.group(1))
                            return None

                        def _normalize_tu(text_value: str):
                            normalized = text_value.strip().lower()
                            return normalized.replace('ú', 'u')

                        def _find_number_with_tu(element, parent=None):
                            text_value = element.get('text') or ''
                            normalized_text = text_value.strip()
                            normalized_for_tu = _normalize_tu(normalized_text)

                            if normalized_for_tu and '(tu)' in normalized_for_tu:
                                # 1) Intentar en el mismo elemento
                                number_local = _extract_number(normalized_text)
                                if number_local:
                                    return number_local

                                # 2) Intentar en los hijos
                                for child in list(element):
                                    number_child = _extract_number(child.get('text') or '')
                                    if number_child:
                                        return number_child

                                # 3) Intentar en el padre y sus hijos (hermanos)
                                if parent is not None:
                                    number_parent = _extract_number(parent.get('text') or '')
                                    if number_parent:
                                        return number_parent

                                    for sibling in list(parent):
                                        if sibling is element:
                                            continue
                                        number_sibling = _extract_number(sibling.get('text') or '')
                                        if number_sibling:
                                            return number_sibling

                            for child in list(element):
                                result = _find_number_with_tu(child, element)
                                if result:
                                    return result

                            return None

                        primary_number = _find_number_with_tu(root)

                        if primary_number:
                            number = primary_number
                            self.log(f"    ¡Número principal detectado en la lista!: {number}", 'success')

                            if 'w4b' in pkg:
                                device_numbers['WhatsApp Business'] = number
                            else:
                                device_numbers['WhatsApp'] = number
                            d.app_stop(pkg)
                            continue # Pasa al siguiente paquete de WhatsApp

                        # Fallback: analizar todo el texto visible pero solo si se referencia a "Tú"
                        all_text_nodes = [node.get('text') for node in nodes if node.get('text')]
                        main_screen_text = " ".join(filter(None, all_text_nodes))

                        match = re.search(r'(\+[\d\s\-\(\)]+)(?=[^\+]*\(t[uú]\))', main_screen_text, flags=re.IGNORECASE)
                        if match:
                            number = re.sub(r'[\s\-\(\)]', '', match.group(1))
                            self.log(f"    ¡Número encontrado en la pantalla principal!: {number}", 'success')

                            if 'w4b' in pkg:
                                device_numbers['WhatsApp Business'] = number
                            else:
                                device_numbers['WhatsApp'] = number
                            d.app_stop(pkg)
                            continue # Pasa al siguiente paquete de WhatsApp
                        else:
                            self.log("    Número no encontrado en la pantalla principal. Continuando al Paso B...", 'info')

                    except Exception as e:
                        self.log(f"    Paso A falló con error: {e}. Se continuará con el Paso B.", 'warning')

                    # --- PASO B: Navegar a "Nuevo Chat" -> Contacto Propio -> Info ---
                    self.log("    Paso B: Navegando a 'Nuevo Chat'...", 'info')
                    new_chat_button = d(description="Nuevo chat")
                    if new_chat_button.wait(timeout=7):
                        new_chat_button.click()
                        time.sleep(1) # Reducido para acelerar

                        tu_regex = r"(?i).*\(t[uú]\)"
                        simple_tu_regex = r"(?i)t[uú]"

                        self.log("    Buscando contacto propio (Tú)...", 'info')
                        own_contact_element = d(textMatches=tu_regex)

                        if not own_contact_element.wait(timeout=7):
                            self.log("    '(Tú)' con paréntesis no visible todavía. Probando coincidencia laxa...", 'warning')
                            own_contact_element = d(textMatches=simple_tu_regex)

                        if own_contact_element.wait(timeout=7):
                            # Estrategia 1: Intentar extraer el número directamente de la lista
                            contact_text = own_contact_element.get_text(timeout=3)
                            if contact_text:
                                match = re.search(r'(\+[\d\s\-\(\)]+)', contact_text)
                                if match:
                                    number = re.sub(r'[\s\-\(\)]', '', match.group(1)) # Limpiar todos los símbolos no numéricos
                                    self.log(f"    Número encontrado directamente en la lista de contactos: {number}", 'success')
                                    # Asignar el número encontrado y saltar al siguiente paquete
                                    if 'w4b' in pkg:
                                        device_numbers['WhatsApp Business'] = number
                                    else:
                                        device_numbers['WhatsApp'] = number
                                    d.press("back") # Volver de la lista de contactos
                                    continue      # Ir al siguiente paquete de whatsapp

                            # Estrategia 2: Si no se encuentra el número, navegar a la pantalla de información
                            self.log("    Número no visible en la lista. Navegando a la info de contacto...", 'info')
                            own_contact_element.click()

                            self.log("    Esperando a que el chat muestre '(Tú)' en la cabecera...", 'info')
                            contact_header_element = d(textMatches=tu_regex)
                            header_ready = contact_header_element.wait(timeout=12)
                            if not header_ready:
                                self.log("    No apareció '(Tú)' en la cabecera tras abrir el chat.", 'error')
                                d.press("back")

                            if header_ready:
                                self.log("    Abriendo pantalla de información...", 'info')

                                # Estrategia de clic modificada para ser más robusta y evitar crashes.
                                clicked_successfully = False

                                # 1. El método más fiable: buscar por el texto visible "(Tú)".
                                self.log("    Intento 1: Buscando por texto '(Tú)'...", 'info')
                                contact_text_element = d(textMatches=tu_regex)
                                if contact_text_element.wait(timeout=7):
                                    try:
                                        contact_text_element.click()
                                        clicked_successfully = True
                                        self.log("    Éxito al hacer clic en elemento con texto '(Tú)'.", 'success')
                                    except Exception as e:
                                        self.log(f"    Elemento de texto encontrado, pero el clic falló: {e}", 'warning')

                                # 2. Si no funciona, intentar el selector original por resourceId.
                                if not clicked_successfully:
                                    self.log("    Intento 2: Buscando por resourceId 'conversation_contact_name'...", 'info')
                                    contact_name_in_toolbar = d(resourceId=f"{pkg}:id/conversation_contact_name")
                                    if contact_name_in_toolbar.wait(timeout=5):
                                        try:
                                            contact_name_in_toolbar.click()
                                            clicked_successfully = True
                                            self.log("    Éxito al hacer clic en 'conversation_contact_name'.", 'success')
                                        except Exception as e:
                                            self.log(f"    Elemento 'conversation_contact_name' encontrado, pero el clic falló: {e}", 'warning')

                                # 3. Como último recurso, hacer clic en toda la barra de herramientas.
                                if not clicked_successfully:
                                    self.log("    Intento 3: Buscando por resourceId 'toolbar'...", 'info')
                                    toolbar = d(resourceId=f"{pkg}:id/toolbar")
                                    if toolbar.wait(timeout=5):
                                        try:
                                            toolbar.click()
                                            clicked_successfully = True
                                            self.log("    Éxito al hacer clic en 'toolbar'.", 'success')
                                        except Exception as e:
                                            self.log(f"    Elemento 'toolbar' encontrado, pero el clic falló: {e}", 'warning')

                                if clicked_successfully:
                                    self.log("    En la pantalla de información, buscando número...", 'info')
                                    time.sleep(2) # Esperar a que cargue la pantalla de info

                                    # Bucle de desplazamiento inteligente con detección de final de página
                                    number_found_in_text = False
                                    max_scrolls = 10  # Límite de seguridad para evitar bucles infinitos
                                    for i in range(max_scrolls):
                                        self.log(f"    Intento de búsqueda y scroll #{i+1}...", 'info')

                                        # Capturar textos ANTES de desplazarse
                                        before_scroll_texts = {elem.get_text() for elem in d(className="android.widget.TextView") if elem.exists and elem.get_text()}
                                        current_view_text = " ".join(before_scroll_texts)

                                        # Buscar el número en la vista actual
                                        match = re.search(r'(\+[\d\s\-\(\)]+)', current_view_text)
                                        if match:
                                            number = re.sub(r'[\s\-\(\)]', '', match.group(1))
                                            self.log(f"    ¡Número encontrado!: {number}", 'success')
                                            number_found_in_text = True
                                            break

                                        # Si no se encuentra, realizar el desplazamiento
                                        self.log("    Número no encontrado, realizando 'fling'...", 'info')
                                        scrollable_view = d(scrollable=True)
                                        if scrollable_view.exists:
                                            scrollable_view.fling.forward()
                                        else:
                                            # Fallback a swipe si no hay un elemento "scrollable"
                                            width, height = d.info['displayWidth'], d.info['displayHeight']
                                            d.swipe(width / 2, height * 0.8, width / 2, height * 0.2, 0.5)

                                        time.sleep(1.5)

                                        # Capturar textos DESPUÉS de desplazarse y comparar
                                        after_scroll_texts = {elem.get_text() for elem in d(className="android.widget.TextView") if elem.exists and elem.get_text()}
                                    if before_scroll_texts == after_scroll_texts:
                                        self.log("    El contenido no cambió. Se ha llegado al final de la página.", 'info')
                                        # Realizar una última búsqueda por si acaso
                                        final_text = " ".join(after_scroll_texts)
                                        final_match = re.search(r'(\+[\d\s\-\(\)]+)', final_text)
                                        if final_match:
                                            number = re.sub(r'[\s\-\(\)]', '', final_match.group(1))
                                            self.log(f"    Número encontrado en la última vista: {number}", 'success')
                                            number_found_in_text = True
                                        break

                                if not number_found_in_text:
                                    self.log("    Búsqueda finalizada. No se encontró el número en la página.", 'warning')
                                    number = "No encontrado"
                            else:
                                self.log("    No se pudo encontrar un elemento para abrir la pantalla de info.", 'error')
                        else:
                            self.log("    No se encontró el contacto '(Tú)' en la lista.", 'error')

                except Exception as e:
                    self.log(f"    Error procesando {pkg}: {e}", 'error')

                # Asignar nombre amigable
                if 'w4b' in pkg:
                    device_numbers['WhatsApp Business'] = number
                else:
                    device_numbers['WhatsApp'] = number

                d.app_stop(pkg) # Limpiar para la siguiente app

        except Exception as e:
            self.log(f"  Fallo grave al conectar o procesar {device_serial}: {e}", 'error')
            return {"WhatsApp": "Error", "WhatsApp Business": "Error"}

        return device_numbers

    def close_all_apps(self, device):
        """Fuerza el cierre de WhatsApp y Google (MOD 25)."""
        self.log(f"Cerrando apps en {device}", 'info')
        targets = [
            "com.whatsapp.w4b",
            "com.whatsapp",
            "com.google.android.googlequicksearchbox",
            "com.google.android.apps.messaging",
            "com.samsung.android.messaging",
            "com.android.mms"
        ]
        for package in targets:
            close_args = ['-s', device, 'shell', 'am', 'force-stop', package]
            self._run_adb_command(close_args, timeout=5) # Usar la función helper, ignorar resultado

# --- Main ---
def main():
    """Función principal: Configura CTk y abre la app principal."""
    ctk.set_appearance_mode("Dark")
    ctk.set_default_color_theme("blue")
    root = ctk.CTk()

    # Iniciar directamente la aplicación principal
    app = Hermes(root)

    root.mainloop()

if __name__ == "__main__":
    main()
